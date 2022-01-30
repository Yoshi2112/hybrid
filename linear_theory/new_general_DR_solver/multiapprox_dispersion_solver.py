# -*- coding: utf-8 -*-
"""
Created on Fri Nov 27 13:30:20 2020

@author: Yoshi

Note: This script just copies the functions related to calculating the cold
dispersion/growth rates since the 'omura play' source script isn't a final
product
"""
import warnings, pdb, sys, os, time
sys.path.append('../')
import numpy             as np
import numba             as nb
import pandas            as pd
import matplotlib        as mpl
import matplotlib.pyplot as plt
import matplotlib.dates  as mdates
import matplotlib.colors as colors
import matplotlib.gridspec as gspec
import matplotlib.cm     as cm
import multiprocessing
import multiprocessing.sharedctypes
from   matplotlib.ticker       import MaxNLocator
from   scipy.interpolate       import splrep, splev
from   matplotlib.lines        import Line2D
from   scipy.optimize          import fsolve
from   scipy.special           import wofz
from   scipy.interpolate       import griddata
from   mpl_toolkits.axes_grid1 import make_axes_locatable

sys.path.append(os.environ['GOOGLE_DRIVE'] + '//Uni//PhD 2017//Data//Scripts//')
import crres_file_readers as cfr
import rbsp_fields_loader as rfl
import rbsp_file_readers  as rfr
import fast_scripts       as fscr
import analysis_scripts   as ascr
import extract_parameters_from_data as data
from   emperics          import geomagnetic_magnitude, sheely_plasmasphere


c  = 3e8
qp = 1.602e-19
mp = 1.673e-27
e0 = 8.854e-12                        # Permittivity of free space
mu0= 4e-7*np.pi


#%% DATA MANAGEMENT FUNCTIONS
def nearest_index(items, pivot):
    closest_val = min(items, key=lambda x: abs(x - pivot))
    for ii in range(len(items)):
        if items[ii] == closest_val:
            return ii
    sys.exit('Error: Unable to find index')

    
def extract_species_arrays(time_start, time_end, probe, rbsp_path='G://DATA//RBSP//',
                           cmp=[70, 20, 10], return_raw_ne=False, HM_filter_mhz=50,
                           nsec=None, HOPE_only=True):
    '''
    Data module only extracts the 3 component species dictionary from HOPE and RBSPICE 
    energetic measurements. This function creates the single axis arrays required to 
    go into the dispersion solver.
    
    All output values are in SI (/m3, etc.) except temperature, which is in eV
    
    Structure of each array (instr_species, time) where
     -- 0,1,2 are cold             H+, He+, O+
     -- 3,4,5 are HOPE    (warm)   "
     -- 6,7,8 are RBSPICE (warmer) "
    '''
    times, B0, cold_dens, hope_dens, hope_temp, hope_anis, spice_dens, spice_temp, spice_anis\
        = data.load_and_interpolate_plasma_params(time_start, time_end, probe, rbsp_path=rbsp_path,
                                                  HM_filter_mhz=HM_filter_mhz, nsec=nsec,
                                                  HOPE_only=HOPE_only)

    Nt       = times.shape[0]
    _density = np.zeros((9, Nt), dtype=float)
    _tper    = np.zeros((9, Nt), dtype=float)
    _ani     = np.zeros((9, Nt), dtype=float)

    _name    = np.array([   'cold $H^{+}$',    'cold $He^{+}$',    'cold $O^{+}$',
                            'HOPE $H^{+}$',    'HOPE $He^{+}$',    'HOPE $O^{+}$',
                         'RBSPICE $H^{+}$', 'RBSPICE $He^{+}$', 'RBSPICE $O^{+}$'])
    
    _mass    = np.array([1.0, 4.0, 16.0, 1.0, 4.0, 16.0, 1.0, 4.0, 16.0]) * mp
    _charge  = np.array([1.0, 1.0,  1.0, 1.0, 1.0,  1.0, 1.0, 1.0,  1.0]) * qp
    
    _density[0] = cold_dens * cmp[0]/100.;  _density[3] = hope_dens[0];  _density[6] = spice_dens[0];
    _density[1] = cold_dens * cmp[1]/100.;  _density[4] = hope_dens[1];  _density[7] = spice_dens[1];
    _density[2] = cold_dens * cmp[2]/100.;  _density[5] = hope_dens[2];  _density[8] = spice_dens[2];

    _tper[0] = 0.0; _tper[3] = hope_temp[0]; _tper[6] = spice_temp[0]
    _tper[1] = 0.0; _tper[4] = hope_temp[1]; _tper[7] = spice_temp[1]
    _tper[2] = 0.0; _tper[5] = hope_temp[2]; _tper[8] = spice_temp[2]

    _ani[0]  = 0.0; _ani[3]  = hope_anis[0]; _ani[6]  = spice_anis[0];
    _ani[1]  = 0.0; _ani[4]  = hope_anis[1]; _ani[7]  = spice_anis[1];
    _ani[2]  = 0.0; _ani[5]  = hope_anis[2]; _ani[8]  = spice_anis[2];

    if return_raw_ne == False:
        return times, B0, _name, _mass, _charge, _density, _tper, _ani
    else:
        return times, B0, _name, _mass, _charge, _density, _tper, _ani, cold_dens
    
    
def CRRES_extract_species_arrays(_time_start, _time_end, crres_path='G://DATA//CRRES//',
                           rc_portion=0.05, cold_comp=[70, 20, 10], hot_comp=[100., 0., 0.],
                           tperp = [30., 0., 0.], ani=[1.0, 0.0, 0.0],
                           return_raw_ne=False, HM_filter_mhz=50, nsec=None):
    '''
    Similar to the RBSP version, but with more tweaks because we have no integrated hot moments
    
    Important params:
        rc_portion -- Portion of electron density devoted to Ring Current (hot ions)
        cold_comp  -- Percentage composition of cold plasma component as [H, He, O] 
        hot_comp   -- Percentage composition of hot  plasma component as [H, He, O]
        tperp      -- Perpendicular temperature of hot plasma as [H, He, O]
        ani        -- Temperature anisotropy of hot plasma as [H, He, O]
        
    Tricky, since the magnetic field and cold density are really the only time-vary
    components that we have access to. It limits our investigation to direct-field
    or total_plasma (i.e. basically cold component) alterations of the growth rate.
    
    Might be able to use LEPA data as some sort of proxy? At least to direct 
    the temperatures/anisotropies
    '''
    times, B0, edens = data.load_CRRES_data(_time_start, _time_end,
                                            crres_path='G://DATA//CRRES//',
                                            nsec=nsec)
    cold_dens = (1. - rc_portion)*edens; hot_dens = edens - cold_dens
    
    _Nt      = times.shape[0]
    _density = np.zeros((6, _Nt), dtype=float)
    _tper    = np.zeros((6, _Nt), dtype=float)
    _ani     = np.zeros((6, _Nt), dtype=float)

    _name    = np.array([   'cold $H^{+}$',    'cold $He^{+}$',    'cold $O^{+}$',
                            'hot $H^{+}$',    'hot $He^{+}$',    'hot $O^{+}$'])
    
    _mass    = np.array([1.0, 4.0, 16.0, 1.0, 4.0, 16.0]) * mp
    _charge  = np.array([1.0, 1.0,  1.0, 1.0, 1.0,  1.0]) * qp
    
    _density[0] = cold_dens * cold_comp[0]*0.01;  _density[3] = hot_dens * hot_comp[0]*0.01
    _density[1] = cold_dens * cold_comp[1]*0.01;  _density[4] = hot_dens * hot_comp[0]*0.01
    _density[2] = cold_dens * cold_comp[2]*0.01;  _density[5] = hot_dens * hot_comp[0]*0.01
    
    _tper[3] = tperp[0]
    _tper[4] = tperp[1]
    _tper[5] = tperp[2]

    _ani[3]  = ani[0]
    _ani[4]  = ani[1]
    _ani[5]  = ani[2]
    
    if return_raw_ne == False:
        return times, B0, _name, _mass, _charge, _density, _tper, _ani
    else:
        return times, B0, _name, _mass, _charge, _density, _tper, _ani, cold_dens    

    
def create_species_array(B0, name, mass, charge, density, tper, ani,
                         remove_zero_density_species=False, pcyc_rad=None):
    '''
    For each ion species, total density is collated and an entry for 'electrons' added (treated as cold)
    Also output a PlasmaParameters dict containing things like alfven speed, density, hydrogen gyrofrequency, etc.
    
    Inputs must be in SI units: T, kg, C, /m3, eV, etc.
    Outputs are in SI units, with frequencies being angular frequencies
    
    -- NOTE --:: if pcyc_rad is defined (as it might be for some paper validations, normalization
                 checks, etc.) it will overwrite the value for B0 to prevent conflicting values
    ''' 
    nsp       = name.shape[0]
    e0        = 8.854e-12
    mu0       = 4e-7*np.pi
    q         = 1.602e-19
    me        = 9.101e-31
    mp        = 1.673e-27
    ne        = density.sum()
    
    if pcyc_rad is not None:
        B0 = pcyc_rad * mp / qp
    
    t_par = np.zeros(nsp); alpha_par = np.zeros(nsp)
    for ii in range(nsp):
        t_par[ii] = q*tper[ii] / (ani[ii] + 1)
        alpha_par[ii] = np.sqrt(2.0 * t_par[ii]  / mass[ii])
    
    # Create initial fields
    Species = np.array([], dtype=[('name', 'U20'),          # Species name
                                  ('mass', 'f8'),           # Mass in kg
                                  ('density', 'f8'),        # Species density in /m3
                                  ('tper', 'f8'),           # Perpendicular temperature in eV
                                  ('anisotropy', 'f8'),     # Anisotropy: T_perp/T_par - 1
                                  ('plasma_freq_sq', 'f8'), # Square of the plasma frequency
                                  ('gyrofreq', 'f8'),       # Cyclotron frequency
                                  ('vth_par', 'f8')])       # Parallel Thermal velocity

    # Insert species values into each
    for ii in range(nsp):
        if density[ii] != 0.0 or remove_zero_density_species == False:
            new_species = np.array([(name[ii], mass[ii], density[ii], tper[ii], ani[ii],
                                                    density[ii] * charge[ii] ** 2 / (mass[ii] * e0),
                                                    charge[ii]  * B0 / mass[ii],
                                                    alpha_par[ii])], dtype=Species.dtype)
            Species = np.append(Species, new_species)
    
    # Add cold electrons
    Species = np.append(Species, np.array([('Electrons', me, ne, 0, 0,
                                            ne * q ** 2 / (me * e0),
                                            -q  * B0 / me,
                                            0.)], dtype=Species.dtype))
    
    PlasParams = {}
    PlasParams['va']       = B0 / np.sqrt(mu0*(density * mass).sum())  # Alfven speed (m/s)
    PlasParams['n0']       = ne                                        # Electron number density (/m3)
    
    if pcyc_rad is None:
        PlasParams['pcyc_rad'] = q*B0 / mp                                 # Proton cyclotron frequency (rad/s)
        PlasParams['B0']       = B0                                        # Magnetic field value (T)
    else:
        PlasParams['pcyc_rad'] = q*B0 / mp                                 # Proton cyclotron frequency (rad/s)
        PlasParams['B0']       = B0                                        # Magnetic field value (T)

    return Species, PlasParams


def fix_names(path):
    # Create destination path
    new_dir = path + 'FIXED//'
    if not os.path.exists(new_dir): os.makedirs(new_dir)
    
    ii = 0
    for file in os.listdir(path):
        if not os.path.isdir(path+file):

            # Load existing data
            DR_file  = np.load(path+file)
            cmp      = DR_file['comp']
            k_np     = DR_file['all_k']
            WPDR_out = DR_file['all_WPDR']
            wCGR_out = DR_file['all_wCGR']
            wVg_out  = DR_file['all_wVg']
    
            times    = DR_file['times']
            B0       = DR_file['B0']
            name     = DR_file['name']
            mass     = DR_file['mass']
            charge   = DR_file['charge']
            density  = DR_file['density']
            tper     = DR_file['tper']
            ani      = DR_file['ani']
            cold_dens= DR_file['cold_dens']
            HM_filter_mhz= DR_file['HM_filter_mhz']
            
            # Re-save with right filename
            nsec = 5
            save_string = file[6:24]
            new_path = new_dir + 'DISPw_{}_cc_{:03}_{:03}_{:03}_{}sec.npz'.format(save_string,
                                                int(10*cmp[0]), int(10*cmp[1]), int(10*cmp[2]), nsec)
            
            np.savez(new_path, all_k=k_np, all_WPDR=WPDR_out, all_wCGR=wCGR_out, all_wVg=wVg_out,
                 comp=np.asarray(cmp), times=times, B0=B0, name=name, mass=mass, charge=charge,
                 density=density, tper=tper, ani=ani, cold_dens=cold_dens,
                 HM_filter_mhz=np.array([HM_filter_mhz]))
            ii += 1
            print('Fixed {} of {}'.format(ii, len(os.listdir(path))))
    return


def get_mag_data(_plot_start, _plot_end, _probe, _olap=0.95, _res=25.0,
                 _HM=True, HM_LP=50.0, HM_HP=None, 
                 _split_HM=False, _split_freq = 7.0,
                 transverse_only=False):
    '''
    Master function to calculate Pc1 spectra and HM. 
    split_HM keyword outputs two HM arrays, for above and below the split
    frequency (and below the LP frequency)
    LP/HP is applied to HM frequencies BEFORE split
    '''
    
    print('Loading magnetic field data')
    _times, _mags, _HM_mags, _dt, gyfreqs = \
            rfl.load_decomposed_magnetic_field(rbsp_path, _plot_start, _plot_end, _probe, 
                                       pad=3600, LP_B0=1.0, LP_HM=HM_LP, 
                                       get_gyfreqs=True, return_B0=False)
            
    print('Calculating dynamic spectra...')
    _xpow, _xtime, _xfreq = fscr.autopower_spectra(_times, _mags[:, 0], _plot_start, 
                                                         _plot_end, _dt, overlap=_olap, df=_res)
        
    _ypow, _ytime, _yfreq = fscr.autopower_spectra(_times, _mags[:, 1], _plot_start, 
                                                     _plot_end, _dt, overlap=_olap, df=_res)
    
    _zpow, _ztime, _zfreq = fscr.autopower_spectra(_times, _mags[:, 2], _plot_start, 
                                                     _plot_end, _dt, overlap=_olap, df=_res)
    
    if transverse_only == False:
        _power = (_xpow[:, :] + _ypow[:, :] + _zpow).T
    else:
        _power = (_xpow[:, :] + _ypow[:, :]).T
    
    # Bandpass and return HM if selected
    if _HM != True:
        return _xtime, _xfreq, _power, _times, gyfreqs
    else:
        print('Filtering HM mags')
        _st, _en = rfr.boundary_idx64(_times, _plot_start, _plot_end)
        
        # High-pass HM frequencies if flagged
        if HM_HP is not None:
            _HM_mags = ascr.clw_high_pass(_HM_mags, HM_HP, _dt)
        
        # Split HM frequencies if flagged
        if _split_HM != True:
            return _xtime, _xfreq, _power, _times[_st:_en], _HM_mags[_st:_en], gyfreqs[:, _st:_en]
        else:
            _lHM_mags = np.zeros(_HM_mags.shape, dtype=_HM_mags.dtype)
            for jj in range(3):
                _lHM_mags[:, jj] = ascr.clw_low_pass(_HM_mags[:, jj], _split_freq, _dt)
                _HM_mags[:, jj]  = _HM_mags[:, jj] - _lHM_mags[:, jj]
            return _xtime, _xfreq, _power, _times[_st:_en], _lHM_mags[_st:_en], _HM_mags[_st:_en], gyfreqs[:, _st:_en]


def get_power_spectrum(_times, _env, zeropad=False, zmult=1.0):
    '''
    Short function to calculate timeseries power spectrum, assuming real input
    to take advantage of FFT symmetry (directly, not with rfft(), since I'm 
    not sure how it deals with timeseries power/energy conservation)
    '''
    # Demean and zero-pad (if flagged)
    if zeropad == False:
        _tseries = _env - _env.mean()
    else:
        # Technically this length here could be changed to something else
        # or set to some condition depending on if _env.shape[0] < X
        _zeros   = np.zeros(int(zmult*_env.shape[0]))
        _tseries = _env - _env.mean()
        _tseries = np.concatenate((_zeros, _tseries, _zeros))
    
    print('Calculating power spectrum...')
    _nP   = _tseries.shape[0]
    _FFT  = np.fft.fft(_tseries)
    _nFFT = 2*_FFT[:_nP//2+1] / _env.shape[0]
    _PWR  = (_nFFT * np.conj(_nFFT)).real
    _dt   = (_times[1] - _times[0]) / np.timedelta64(1, 's')
    _df   = 1./(_nP * _dt)
    _freq = np.arange(_nP//2 + 1) * _df
    return _freq, _PWR


#%% CORE CALCULATION FUNCTIONS  
def remove_duplicates(PDR_solns):
    '''
    For a given series of solutions in N bands, zeros duplicate solutions
    Advanced: Detects the cyclotron frequencies (H, He, O) and zeros the 
    solution that doesn't fall in the requisite band. If this check doesn't
    work (for some reason), just delete the second (He or O) duplicate
    
    Just write it simple initially. Hard-code for 3 solutions only.
    '''
    N_dups = 0
    for _mm in range(PDR_solns.shape[0]):
        if PDR_solns[_mm, 0, 0] == PDR_solns[_mm, 1, 0]:
            PDR_solns[_mm, 1] *= np.nan; N_dups += 1
            
        if PDR_solns[_mm, 0, 0] == PDR_solns[_mm, 2, 0]:
            PDR_solns[_mm, 2] *= np.nan; N_dups += 1
            
        if PDR_solns[_mm, 1, 0] == PDR_solns[_mm, 2, 0]:
            PDR_solns[_mm, 2] *= np.nan; N_dups += 1
    return N_dups

def remove_bad_solutions(PDR_solns, ier):
    N_bad = 0
    for jj in range(PDR_solns.shape[1]):
        for ii in range(1, PDR_solns.shape[0]):
            if ier[ii, jj] == 5:
                PDR_solns[ii, jj] = np.nan
                N_bad += 1
    return N_bad

def Z(arg):
    '''
    Return Plasma Dispersion Function : Normalized Fadeeva function
    Plasma dispersion function related to Fadeeva function
    (Summers & Thorne, 1993) by i*sqrt(pi) factor.
    '''
    return 1j*np.sqrt(np.pi)*wofz(arg)

def Y(arg):
    return np.real(Z(arg))


def hot_dispersion_eqn(w, k, Species):
    '''
    Function used in scipy.fsolve minimizer to find roots of dispersion relation
    for hot plasma approximation.
    Iterates over each k to find values of w that minimize to D(wr, k) = 0
    
    In this case, w is a vector [wr, wi] and fsolve is effectively doing a multivariate
    optimization.
    
    type_out allows purely real or purely imaginary (coefficient only) for root
    finding. Set as anything else for complex output.
    
    FSOLVE OPTIONS :: If bad solution, return np.nan?
    
    Eqns 1, 13 of Chen et al. (2013) equivalent to those of Wang et al. (2016)
    '''
    wc = w[0] + 1j*w[1]
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        hot_sum = 0.0
        for ii in range(Species.shape[0]):
            sp = Species[ii]
            if sp['tper'] == 0:
                hot_sum += sp['plasma_freq_sq'] * wc / (sp['gyrofreq'] - wc)
            else:
                pdisp_arg   = (wc - sp['gyrofreq']) / (sp['vth_par']*k)
                pdisp_func  = Z(pdisp_arg)*sp['gyrofreq'] / (sp['vth_par']*k)
                brackets    = (sp['anisotropy'] + 1) * (wc - sp['gyrofreq'])/sp['gyrofreq'] + 1
                Is          = brackets * pdisp_func + sp['anisotropy']
                hot_sum    += sp['plasma_freq_sq'] * Is

    solution = (wc ** 2) - (c * k) ** 2 + hot_sum
    return np.array([solution.real, solution.imag])


def warm_dispersion_eqn(w, k, Species):
    '''    
    Function used in scipy.fsolve minimizer to find roots of dispersion relation
    for warm plasma approximation.
    Iterates over each k to find values of w that minimize to D(wr, k) = 0
    
    Eqn 14 of Chen et al. (2013)
    '''
    wr = w[0]
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        warm_sum = 0.0
        for ii in range(Species.shape[0]):
            sp = Species[ii]
            if sp['tper'] == 0:
                warm_sum   += sp['plasma_freq_sq'] * wr / (sp['gyrofreq'] - wr)
            else:
                pdisp_arg   = (wr - sp['gyrofreq']) / (sp['vth_par']*k)
                numer       = ((sp['anisotropy'] + 1)*wr - sp['anisotropy']*sp['gyrofreq'])
                Is          = sp['anisotropy'] + numer * Y(pdisp_arg) / (sp['vth_par']*k)
                warm_sum   += sp['plasma_freq_sq'] * Is
            
    solution = wr ** 2 - (c * k) ** 2 + warm_sum
    return np.array([solution, 0.0])


def cold_dispersion_eqn(w, k, Species):
    '''
    Function used in scipy.fsolve minimizer to find roots of dispersion relation
    for warm plasma approximation.
    Iterates over each k to find values of w that minimize to D(wr, k) = 0
    
    Eqn 19 of Chen et al. (2013)
    '''
    wr = w[0]
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        cold_sum = 0.0
        for ii in range(Species.shape[0]):
            cold_sum += Species[ii]['plasma_freq_sq'] * wr / (Species[ii]['gyrofreq'] - wr)
            
    solution = wr ** 2 - (c * k) ** 2 + cold_sum
    return np.array([solution, 0.0])


def get_warm_growth_rates(wr, k, Species):
    '''
    Calculates the temporal and convective linear growth rates for a plasma
    composition contained in Species for each frequency w. Assumes a cold
    dispersion relation is valid for k but uses a warm approximation in the
    solution for D(w, k).
    
    Equations adapted from Chen et al. (2013)
    '''    
    w_der_sum = 0.0
    k_der_sum = 0.0
    Di        = 0.0
    
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        
        for ii in range(Species.shape[0]):
            sp = Species[ii]
            
            # If cold
            if sp['tper'] == 0:
                w_der_sum += sp['plasma_freq_sq'] * sp['gyrofreq'] / (wr - sp['gyrofreq'])**2
                k_der_sum += 0.0
                Di        += 0.0
            
            # If hot
            else:
                zs           = (wr - sp['gyrofreq']) / (sp['vth_par']*k)
                Yz           = np.real(Z(zs))
                dYz          = -2*(1 + zs*Yz)
                A_bit        = (sp['anisotropy'] + 1) * wr / sp['gyrofreq']
                
                # Calculate frequency derivative of Dr (sums bit)
                w_der_outsd  = sp['plasma_freq_sq']*sp['gyrofreq'] / (wr*k*sp['vth_par'])
                w_der_first  = A_bit * Yz
                w_der_second = (A_bit - sp['anisotropy']) * wr * dYz / (k * sp['vth_par']) 
                w_der_sum   += w_der_outsd * (w_der_first + w_der_second)
        
                # Calculate Di (sums bit)
                Di_bracket = 1 + (sp['anisotropy'] + 1) * (wr - sp['gyrofreq']) / sp['gyrofreq']
                Di_after   = sp['gyrofreq'] / (k * sp['vth_par']) * np.sqrt(np.pi) * np.exp(- zs ** 2)
                Di        += sp['plasma_freq_sq'] * Di_bracket * Di_after
        
                # Calculate wavenumber derivative of Dr (sums bit)
                k_der_outsd  = sp['plasma_freq_sq']*sp['gyrofreq'] / (wr*k*k*sp['vth_par'])
                k_der_first  = A_bit - sp['anisotropy']
                k_der_second = Yz + zs * dYz
                k_der_sum   += k_der_outsd * k_der_first * k_der_second
    
    # Get and return ratio
    Dr_wder = 2*wr + w_der_sum
    Dr_kder = -2*k*c**2 - k_der_sum

    temporal_growth_rate   = - Di / Dr_wder
    group_velocity         = - Dr_kder / Dr_wder
    convective_growth_rate =   temporal_growth_rate / np.abs(group_velocity)
    return temporal_growth_rate, convective_growth_rate, group_velocity


def get_cold_growth_rates(wr, k, Species):
    '''
    Simplified version of the warm growth rate equation.
    
    From Chen et al. (2013)
    '''
    w_der_sum = 0.0
    Di        = 0.0
    
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        
        for ii in range(Species.shape[0]):
            sp         = Species[ii]
            w_der_sum += sp['plasma_freq_sq'] * sp['gyrofreq'] / (wr - sp['gyrofreq'])**2
            
            if sp['vth_par'] != 0.0:
                # Calculate Di (sums bit)
                zs         = (wr - sp['gyrofreq']) / (sp['vth_par']*k)
                Di_bracket = 1 + (sp['anisotropy'] + 1) * (wr - sp['gyrofreq']) / sp['gyrofreq']
                Di_after   = sp['gyrofreq'] / (k * sp['vth_par']) * np.sqrt(np.pi) * np.exp(- zs ** 2)
                Di        += sp['plasma_freq_sq'] * Di_bracket * Di_after
    
    # Get and return ratio
    Dr_wder = 2*wr + w_der_sum
    Dr_kder = -2*k*c**2

    temporal_growth_rate   = - Di / Dr_wder
    group_velocity         = - Dr_kder / Dr_wder
    convective_growth_rate =   temporal_growth_rate / np.abs(group_velocity)
    return temporal_growth_rate, convective_growth_rate, group_velocity


def get_dispersion_relation(Species, k, approx='warm', guesses=None, complex_out=True,
                            print_filtered=True, return_vg=False, force_3soln=False):
    '''
    All input values in SI units from the Species array
    
    Given a range of k, returns the real and imaginary parts of the plasma dispersion
    relation specified by the Species present.
    
    Type of dispersion relation controlled by 'approx' kwarg as:
        hot  :: Full dispersion relation for complex w = wr + i*gamma
        warm :: Small growth rate approximation that allows D(wr, k) = 0
        cold :: Dispersion relation used for wr, growth rate calculated as per warm
        
    Gamma is only solved for in the DR with the hot approx
    The other two require calls to another function with a wr arg.
    
    To do: Code a number of 'tries' when a solution fails to converge. Change
            the initial 'guess' for that k by using previous guess multiplied
            by a random number between 0.99-1.01 (i.e. 1% variation)
            
           Could also change to make sure solutions use previous best solution,
            if the previous solution ended up as a np.nan, and implement nan'ing
            solutions as they're solved to prevent completely losing convergence.
           
           Better way to calculate PDR for k = 0? Exceptions from a bunch of things,
           but surely there's an analytic way of doing it with the not-CPDR (or is 
           that valid?)
    
           Also need to write function to remove duplicate solutions, and check which
           band they should theoretically be in (not that it super matters, as long as
           there's only one solution). This most likely occurs when cold density is
           close to zero for a species, generally the warm species part isn't enough
           to sustain the wave in a band (or is this a result of the approximations?)
           
    More TODO: 
        -- Find a better way to solve for w from k instead of relying on fsolve. Is there
            a way to solve for each individual mode without the risk of jumping bands?
            I mean it's just the zero to an equation, but we don't want it to grab onto 
            the wrong one. But also how do you tell it if it's the wrong one or not?
        -- Find a way to solve for the R mode as well. Stick this as the first returned
            solution so that the number of ions doesn't matter.
    '''
    gyfreqs, counts = np.unique(Species['gyrofreq'], return_counts=True)
    
    # Remove electron count, 
    gyfreqs = gyfreqs[1:]
    N_solns = counts.shape[0] - 1
    
    # Check that 3 species are present if flagged
    if force_3soln == True and N_solns != 3:
        raise ValueError('force_3soln flag requires one of each (H, He, O) to be in Species array')

    # fsolve arguments
    eps    = 1.01           # Offset used to supply initial guess (since right on w_cyc returns an error)
    tol    = 1e-10          # Absolute solution convergence tolerance in rad/s
    fev    = 1000000        # Maximum number of iterations
    Nk     = k.shape[0]     # Number of wavenumbers to solve for
    
    # Solution and error arrays :: Two-soln array for wr, gamma. 
    # PDR_solns init'd as ones because 0.0 returns spurious root
    PDR_solns = np.ones( (Nk, N_solns, 2), dtype=np.float64)*0.01
    CGR_solns = np.zeros((Nk, N_solns   ), dtype=np.float64)
    VEL_solns = np.zeros((Nk, N_solns   ), dtype=np.float64)
    ier       = np.zeros((Nk, N_solns   ), dtype=int)
    msg       = np.zeros((Nk, N_solns   ), dtype='<U256')

    # Initial guesses (check this?)
    for ii in range(1, N_solns):
        PDR_solns[0, ii - 1]  = np.array([[gyfreqs[-ii - 1] * eps, 0.0]])
    
    if approx == 'hot':
        func = hot_dispersion_eqn
    elif approx == 'warm':
        func = warm_dispersion_eqn
    elif approx == 'cold':
        func = cold_dispersion_eqn
    else:
        sys.exit('ABORT :: kwarg approx={} invalid. Must be \'cold\', \'warm\', or \'hot\'.'.format(approx))
    
    # Define function to solve for (all have same arguments)
    if guesses is None or guesses.shape != PDR_solns.shape:
        for jj in range(N_solns):
            for ii in range(1, Nk):
                #if np.isnan(k[ii]):
                #    PDR_solns[ii, jj] = np.nan
                #else:
                    PDR_solns[ii, jj], infodict, ier[ii, jj], msg[ii, jj] =\
                        fsolve(func, x0=PDR_solns[ii - 1, jj], args=(k[ii], Species), xtol=tol, maxfev=fev, full_output=True)
            
            if False:
                # Solve for k[0] using initial guess of k[1]
                PDR_solns[0, jj], infodict, ier[0, jj], msg[0, jj] =\
                    fsolve(func, x0=PDR_solns[1, jj], args=(k[0], Species), xtol=tol, maxfev=fev, full_output=True)
            else:
                # Set k[0] as equal to k[1] (better for large Nk)
                PDR_solns[0, jj] = PDR_solns[1, jj]
    else:
        for jj in range(N_solns):
            for ii in range(1, Nk):
                PDR_solns[ii, jj], infodict, ier[ii, jj], msg[ii, jj] =\
                    fsolve(func, x0=guesses[ii, jj], args=(k[ii], Species), xtol=tol, maxfev=fev, full_output=True)

    N_bad = remove_bad_solutions(PDR_solns, ier)
    #N_dup = remove_duplicates(PDR_solns)
    if print_filtered == True:
        print(f'{N_bad} solutions filtered for {approx} approximation.')
        #print(f'{N_dup} duplicates removed.')

    # Solve for growth rate/convective growth rate here
    if approx == 'hot':
        CGR_solns *= np.nan
    elif approx == 'warm':
        for jj in range(N_solns):
            PDR_solns[:, jj, 1], CGR_solns[:, jj], VEL_solns[:, jj] = get_warm_growth_rates(PDR_solns[:, jj, 0], k, Species)
    elif approx == 'cold':
        for jj in range(N_solns):
            PDR_solns[:, jj, 1], CGR_solns[:, jj], VEL_solns[:, jj] = get_cold_growth_rates(PDR_solns[:, jj, 0], k, Species)

    # Convert to complex number if flagged, else return as (Nk, N_solns, 2) for real/imag components
    if complex_out == True:
        OUT_solns = np.zeros((Nk, N_solns   ), dtype=np.complex128)
        for ii in range(Nk):
            for jj in range(N_solns):
                OUT_solns[ii, jj] = PDR_solns[ii, jj, 0] + 1j*PDR_solns[ii, jj, 1]
    else:
        OUT_solns = PDR_solns
    
    # Zero out bands that have no density (i.e. no solution should be present)
    if force_3soln == True:
        #print('Forcing 3 species solution...')
        for lbl, ii in zip(['H^{+}', 'He^{+}', 'O^{+}'], range(3)):
            
            # Check if any species present
            cum_dens = 0.0
            for _xx in range(Species.shape[0]):
                if lbl in Species[_xx]['name']:
                    cum_dens += Species[_xx]['density']
                    
            if cum_dens == 0.0:
                OUT_solns[:, ii] *= 0.0
                CGR_solns[:, ii] *= 0.0
                    
    if return_vg == True:
        return OUT_solns, CGR_solns, VEL_solns
    else:
        return OUT_solns, CGR_solns


#%% DATA TIMESERIES CALCULATIONS
def get_DR_filepath(_save_dir, _time_start, _time_end, _cmp, _nsec):
    save_string = _time_start.astype(object).strftime('%Y%m%d_%H%M_') \
                + _time_end.astype(object).strftime('%H%M')
    
    h_pcen  = int(round(1e3*_cmp[0], 2))
    he_pcen = int(round(1e3*_cmp[1], 2))
    o_pcen  = int(round(1e3*_cmp[2], 2))
    
    # Assume that the 100% cold composition is hydrogen (for now)
    if any(cc == 100.0 for cc in _cmp):
       if _nsec is None:
           DR_path = _save_dir + 'DISPw_{}_cc_1000_000_000.npz'.format(save_string)
       else:
           DR_path = _save_dir + 'DISPw_{}_cc_1000_000_000_{}sec.npz'.format(save_string, _nsec)
    else:
        if _nsec is None:
            DR_path = _save_dir + 'DISPw_{}_cc_{:03}_{:03}_{:03}.npz'.format(save_string,
                                                    h_pcen, he_pcen, o_pcen)
        else:
            DR_path = _save_dir + 'DISPw_{}_cc_{:03}_{:03}_{:03}_{}sec.npz'.format(save_string,
                                                    h_pcen, he_pcen, o_pcen, _nsec)
    return DR_path


def get_DRs_for_data_timeseries(_save_dir, _time_start, _time_end, _probe, _pad, _cmp, 
                                kmin=0.0, kmax=1.0, Nk=1000, knorm=True,
                                _nsec=None, HM_filter_mhz=50):
    '''
    Calculate dispersion relation and temporal growth rate for all times between
    time_start and time_end. Nk is resolution in k-space for each solution
    (more = slower for each solutions). nsec is time cadence of interpolated
    satellite parameters.
    
    kmin, kmax :: Wavenumber range in units of c/va or /m
    Nk         :: Number of points to calculate wavenumber for
    knorm      :: Flag to specify units of kmin/kmax. True: c/va, False: /m
    
    suff kwarg for parameters varying only by flags.
    
    Filenames are arranged as 'DISP_YYYYMMDD_AAAA_BBBB_cc_JJ_KK_LL.npz'
        for Year, Month, Day
        for A, B    as the start/end UT times
        for J, K, L as the cold plasma composition assumption
        
    Should probably also save things like the filter
    '''
    DR_path = get_DR_filepath(_save_dir, _time_start, _time_end, _cmp, _nsec)
    
    # Cast as array just in case its a list
    _cmp = np.asarray(_cmp)
    
    if os.path.exists(DR_path) == False:
        times, B0, name, mass, charge, density, tper, ani, cold_dens = \
        extract_species_arrays(_time_start, _time_end, _probe, cmp=_cmp, 
                               return_raw_ne=True, nsec=_nsec, HM_filter_mhz=HM_filter_mhz)

        Nt         = times.shape[0]
        all_CPDR   = np.zeros((Nt, Nk, 3), dtype=np.complex128)
        all_WPDR   = np.zeros((Nt, Nk, 3), dtype=np.complex128)
        all_HPDR   = np.zeros((Nt, Nk, 3), dtype=np.complex128)
        all_k      = np.zeros((Nt, Nk)   , dtype=np.float64)
        
        # Get dispersion relations for each time if possible
        for ii in range(Nt):
            Species, PP = create_species_array(B0[ii], name, mass, charge, density[:, ii], tper[:, ii], ani[:, ii])

            all_k[ii]   = np.linspace(kmin, kmax, Nk, endpoint=False)
            # Convert k-extrema to /m if needed
            if knorm == True:
                all_k[ii] *= PP['pcyc_rad'] / PP['va']
                
            print('Calculating dispersion/growth relation for {}'.format(times[ii]))
            try:
                all_CPDR[ii], cold_CGR = get_dispersion_relation(Species, all_k[ii], approx='cold')
            except:
                print('COLD ERROR: Skipping to next time...')
                all_CPDR[ii, :, :] = np.ones((Nk, 3), dtype=np.complex128) * np.nan 
                
            try:                
                all_WPDR[ii], warm_CGR = get_dispersion_relation(Species, all_k[ii], approx='warm')
            except:
                print('WARM ERROR: Skipping to next time...')
                all_WPDR[ii, :, :] = np.ones((Nk, 3), dtype=np.complex128) * np.nan
                
            try:
                all_HPDR[ii],  hot_CGR = get_dispersion_relation(Species, all_k[ii], approx='hot' )
            except:
                print('HOT ERROR: Skipping to next time...')
                all_HPDR[ii, :, :] = np.ones((Nk, 3), dtype=np.complex128) * np.nan
        
        # Saves data used for DR calculation as well, for future reference (and plotting)
        if os.path.exists(save_dir) == False:
            os.makedirs(save_dir)
                
        print('Saving dispersion history...')
        np.savez(DR_path, all_CPDR=all_CPDR, all_WPDR=all_WPDR, all_HPDR=all_HPDR, all_k=all_k, comp=_cmp,
                 times=times, B0=B0, name=name, mass=mass, charge=charge, density=density, tper=tper,
                 ani=ani, cold_dens=cold_dens, HM_filter_mhz=np.array([HM_filter_mhz]))
    else:
        print('Dispersion results already exist, loading from file...')
        DR_file   = np.load(DR_path)
        
        all_k     = DR_file['all_k']
        all_CPDR  = DR_file['all_CPDR']
        all_WPDR  = DR_file['all_WPDR']
        all_HPDR  = DR_file['all_HPDR']
                
        times     = DR_file['times']
        B0        = DR_file['B0']
        name      = DR_file['name']
        mass      = DR_file['mass']
        charge    = DR_file['charge']
        density   = DR_file['density']
        tper      = DR_file['tper']
        ani       = DR_file['ani']
        cold_dens = DR_file['cold_dens']
    return all_k, all_CPDR, all_WPDR, all_HPDR, \
           times, B0, name, mass, charge, density, tper, ani, cold_dens


def get_DRs_chunked(Nk, kmin, kmax, knorm, times, B0, name, mass, charge, density, tper, ani,
                      k_dict, CPDR_dict, WPDR_dict, HPDR_dict, cCGR_dict, wCGR_dict, hCGR_dict,
                      cVg_dict, wVg_dict, hVg_dict, st=0, worker=None):
    '''
    Function designed to be run in parallel. All dispersion inputs as previous. 
    output_PDR arrays are shared memory
    
    Dispersion inputs (B0, times, density, etc.) are chunked so this function calculates
    them all.
    
    st is the first time index to be computed in the main array, gives position in
    output array to place results in
    
    Thus the array index in output_PDRs will be st+ii for ii in range(Nt)
    '''
    k_arr    = np.frombuffer(k_dict['arr']).reshape(k_dict['shape'])
    
    CPDR_arr = np.frombuffer(CPDR_dict['arr']).reshape(CPDR_dict['shape'])
    WPDR_arr = np.frombuffer(WPDR_dict['arr']).reshape(WPDR_dict['shape'])
    HPDR_arr = np.frombuffer(HPDR_dict['arr']).reshape(HPDR_dict['shape'])
    
    cCGR_arr = np.frombuffer(cCGR_dict['arr']).reshape(cCGR_dict['shape'])
    wCGR_arr = np.frombuffer(wCGR_dict['arr']).reshape(wCGR_dict['shape'])
    hCGR_arr = np.frombuffer(hCGR_dict['arr']).reshape(hCGR_dict['shape'])
    
    cVg_arr  = np.frombuffer(cVg_dict['arr']).reshape(cVg_dict['shape'])
    wVg_arr  = np.frombuffer(wVg_dict['arr']).reshape(wVg_dict['shape'])
    hVg_arr  = np.frombuffer(hVg_dict['arr']).reshape(hVg_dict['shape'])
    
    for ii in range(times.shape[0]):
        Species, PP = create_species_array(B0[ii], name, mass, charge, density[:, ii], tper[:, ii], ani[:, ii])
        this_k      = np.linspace(kmin, kmax, Nk, endpoint=False)
        
        # Convert k-extrema to /m if needed
        if knorm == True:
            this_k *= PP['pcyc_rad'] / PP['va']
        
        # Calculate dispersion relations if possible
        print('Worker', worker, '::', times[ii])
        #try:
        this_CPDR, this_cCGR, this_cVg = get_dispersion_relation(Species, this_k, approx='cold', complex_out=False,
                                                       print_filtered=False, return_vg=True)
        #except:
        #    print('COLD ERROR: Skipping', times)
        #    this_CPDR = np.ones((Nk, 3, 2), dtype=np.complex128) * np.nan 
            
        #try:            
        this_WPDR, this_wCGR, this_wVg = get_dispersion_relation(Species, this_k, approx='warm', complex_out=False,
                                                       print_filtered=False, return_vg=True)
        #except:
        #    print('WARM ERROR: Skipping', times)
        #    this_WPDR = np.ones((Nk, 3, 2), dtype=np.complex128) * np.nan
        
        #try:
        this_HPDR,  this_hCGR, this_hVg = get_dispersion_relation(Species, this_k, approx='hot' , complex_out=False,
                                                        print_filtered=False, return_vg=True)
        #except:
        #    print('HOT ERROR: Skipping', times)
        #    this_HPDR = np.ones((Nk, 3, 2), dtype=np.complex128) * np.nan
               
        # Dump to shared memory
        k_arr[   st+ii, :]       = this_k[...]
        
        CPDR_arr[st+ii, :, :, :] = this_CPDR[...]
        WPDR_arr[st+ii, :, :, :] = this_WPDR[...]
        HPDR_arr[st+ii, :, :, :] = this_HPDR[...]
        
        cCGR_arr[st+ii, :, :]    = this_cCGR[...]
        wCGR_arr[st+ii, :, :]    = this_wCGR[...]
        hCGR_arr[st+ii, :, :]    = this_hCGR[...]
        
        cVg_arr[st+ii, :, :]    = this_cVg[...]
        wVg_arr[st+ii, :, :]    = this_wVg[...]
        hVg_arr[st+ii, :, :]    = this_hVg[...]
    return


def get_all_DRs_parallel(_save_dir, _time_start, _time_end, _probe, _cmp, 
                    kmin=0.0, kmax=1.0, Nk=1000, knorm=True,
                    _nsec=None, HM_filter_mhz=50, N_procs=7,
                    suff=''):
    '''
    Calculates all 3 approximations for the linear dispersion relation given
    satellite data parameters. Parallelized for faster running.
    '''
    DR_path = get_DR_filepath(_save_dir, _time_start, _time_end, _cmp, _nsec)
    
    if os.path.exists(DR_path) == False:
        # Load data
        times, B0, name, mass, charge, density, tper, ani, cold_dens = \
        extract_species_arrays(_time_start, _time_end, _probe, cmp=np.asarray(_cmp), 
                               return_raw_ne=True, nsec=_nsec, HM_filter_mhz=HM_filter_mhz)
    
        Nt      = times.shape[0]
        
        procs   = []
        
        # Create raw shared memory arrays with shapes. Store in dict to send with each worker
        k_shape      = (Nt, Nk)
        k_shm        = multiprocessing.RawArray('d', Nt * Nk)
        k_dict       = {'arr': k_shm, 'shape': k_shape}
        
        CPDR_shape   = (Nt, Nk, 3, 2)
        CPDR_shm     = multiprocessing.RawArray('d', Nt*Nk*3*2)
        CPDR_dict    = {'arr': CPDR_shm, 'shape': CPDR_shape}
        
        WPDR_shape   = (Nt, Nk, 3, 2)
        WPDR_shm     = multiprocessing.RawArray('d', Nt*Nk*3*2)
        WPDR_dict    = {'arr': WPDR_shm, 'shape': WPDR_shape}
        
        HPDR_shape   = (Nt, Nk, 3, 2)
        HPDR_shm     = multiprocessing.RawArray('d', Nt*Nk*3*2)
        HPDR_dict    = {'arr': HPDR_shm, 'shape': HPDR_shape}
        
        # Also for convective growth rates
        cCGR_shape   = (Nt, Nk, 3)
        cCGR_shm     = multiprocessing.RawArray('d', Nt*Nk*3)
        cCGR_dict    = {'arr': cCGR_shm, 'shape': cCGR_shape}
        
        wCGR_shape   = (Nt, Nk, 3)
        wCGR_shm     = multiprocessing.RawArray('d', Nt*Nk*3)
        wCGR_dict    = {'arr': wCGR_shm, 'shape': wCGR_shape}
        
        hCGR_shape   = (Nt, Nk, 3)
        hCGR_shm     = multiprocessing.RawArray('d', Nt*Nk*3)
        hCGR_dict    = {'arr': hCGR_shm, 'shape': hCGR_shape}
        
        # And for group velocities
        cVg_shape   = (Nt, Nk, 3)
        cVg_shm     = multiprocessing.RawArray('d', Nt*Nk*3)
        cVg_dict    = {'arr': cVg_shm, 'shape': cVg_shape}
        
        wVg_shape   = (Nt, Nk, 3)
        wVg_shm     = multiprocessing.RawArray('d', Nt*Nk*3)
        wVg_dict    = {'arr': wVg_shm, 'shape': wVg_shape}
        
        hVg_shape   = (Nt, Nk, 3)
        hVg_shm     = multiprocessing.RawArray('d', Nt*Nk*3)
        hVg_dict    = {'arr': hVg_shm, 'shape': hVg_shape}
        
        
        # Create numpy view into shared memory
        k_np         = np.frombuffer(k_shm).reshape(k_shape)
        CPDR_np      = np.frombuffer(CPDR_shm).reshape(CPDR_shape)
        WPDR_np      = np.frombuffer(WPDR_shm).reshape(WPDR_shape)
        HPDR_np      = np.frombuffer(HPDR_shm).reshape(HPDR_shape)
        
        cCGR_np      = np.frombuffer(cCGR_shm).reshape(cCGR_shape)
        wCGR_np      = np.frombuffer(wCGR_shm).reshape(wCGR_shape)
        hCGR_np      = np.frombuffer(hCGR_shm).reshape(hCGR_shape)
        
        cVg_np       = np.frombuffer(cVg_shm).reshape(cVg_shape)
        wVg_np       = np.frombuffer(wVg_shm).reshape(wVg_shape)
        hVg_np       = np.frombuffer(hVg_shm).reshape(hVg_shape)
        
        
        # Split input data into a list of chunks
        time_chunks    = np.array_split(times,   N_procs)
        field_chunks   = np.array_split(B0,      N_procs)
        density_chunks = np.array_split(density, N_procs, axis=1)
        tper_chunks    = np.array_split(tper,    N_procs, axis=1)
        ani_chunks     = np.array_split(ani,     N_procs, axis=1)
    
        # Instatiate each process with a different chunk
        acc = 0; start = time.time()
        for xx in range(N_procs):
            print('Starting process', xx)
            proc = multiprocessing.Process(target=get_DRs_chunked,
                                        args=(Nk, kmin, kmax, knorm, time_chunks[xx],
                                        field_chunks[xx], name, mass, charge, density_chunks[xx],
                                        tper_chunks[xx], ani_chunks[xx],
                                        k_dict, CPDR_dict, WPDR_dict, HPDR_dict,
                                        cCGR_dict, wCGR_dict, hCGR_dict,
                                        cVg_dict, wVg_dict, hVg_dict),
                                        kwargs={'st':acc, 'worker':xx})
            procs.append(proc)
            proc.start()
            
            acc += time_chunks[xx].shape[0]
        
        # Complete processes
        for proc in procs:
            proc.join()
                
        print('All processes complete')
        end = time.time()
        print('Total parallel time = {}s'.format(str(end-start)))
        
        # Make output complex
        CPDR_out = np.zeros((Nt, Nk, 3), dtype=np.complex128)
        WPDR_out = np.zeros((Nt, Nk, 3), dtype=np.complex128)
        HPDR_out = np.zeros((Nt, Nk, 3), dtype=np.complex128)
    
        for ii in range(Nt):
            for jj in range(Nk):
                for kk in range(3):
                    CPDR_out[ii, jj, kk] = CPDR_np[ii, jj, kk, 0] + 1j * CPDR_np[ii, jj, kk, 1]
                    WPDR_out[ii, jj, kk] = WPDR_np[ii, jj, kk, 0] + 1j * WPDR_np[ii, jj, kk, 1]
                    HPDR_out[ii, jj, kk] = HPDR_np[ii, jj, kk, 0] + 1j * HPDR_np[ii, jj, kk, 1]
            
        cCGR_out = cCGR_np;     cVg_out = cVg_np
        wCGR_out = wCGR_np;     wVg_out = wVg_np
        hCGR_out = hCGR_np;     hVg_out = hVg_np
            
        # Saves data used for DR calculation as well, for future reference (and plotting)
        if os.path.exists(save_dir) == False:
            os.makedirs(save_dir)
                
        print('Saving dispersion history...')
        np.savez(DR_path, all_CPDR=CPDR_out, all_WPDR=WPDR_out, all_HPDR=HPDR_out,
                 all_k=k_np, all_cCGR=cCGR_np, all_wCGR=wCGR_np, all_hCGR=hCGR_np,
                 all_cVg=cVg_np, all_wVg=wVg_np, all_hVg=hVg_np, comp=np.asarray(_cmp),
                 times=times, B0=B0, name=name, mass=mass, charge=charge, density=density, tper=tper,
                 ani=ani, cold_dens=cold_dens, HM_filter_mhz=np.array([HM_filter_mhz]))
    else:
        print('Dispersion results already exist, loading from file...')
        DR_file   = np.load(DR_path)
        
        k_np      = DR_file['all_k']
        CPDR_out  = DR_file['all_CPDR']
        WPDR_out  = DR_file['all_WPDR']
        HPDR_out  = DR_file['all_HPDR']
        
        try:
            cCGR_out  = DR_file['all_cCGR']
            wCGR_out  = DR_file['all_wCGR']
            hCGR_out  = DR_file['all_hCGR']
        except:
            print('No convective growth rates found in file.')
            cCGR_out  = None
            wCGR_out  = None
            hCGR_out  = None
            
        try:
            cVg_out  = DR_file['all_cVg']
            wVg_out  = DR_file['all_wVg']
            hVg_out  = DR_file['all_hVg']
        except:
            print('No group velocities found in file.')
            cVg_out  = None
            wVg_out  = None
            hVg_out  = None
        
        times     = DR_file['times']
        B0        = DR_file['B0']
        name      = DR_file['name']
        mass      = DR_file['mass']
        charge    = DR_file['charge']
        density   = DR_file['density']
        tper      = DR_file['tper']
        ani       = DR_file['ani']
        cold_dens = DR_file['cold_dens']
        
    return k_np, CPDR_out, WPDR_out, HPDR_out, cCGR_out, wCGR_out, hCGR_out, \
           cVg_out, wVg_out, hVg_out,                                        \
           times, B0, name, mass, charge, density, tper, ani, cold_dens


def get_DRs_chunked_warm_only(Nk, kmin, kmax, knorm, times, B0, name, mass, charge, density, tper, ani,
                      k_dict, WPDR_dict, wCGR_dict, wVg_dict, st=0, worker=None, approx='warm'):
    '''
    Function designed to be run in parallel. All dispersion inputs as previous. 
    output_PDR arrays are shared memory
    
    Dispersion inputs (B0, times, density, etc.) are chunked so this function calculates
    them all.
    
    st is the first time index to be computed in the main array, gives position in
    output array to place results in
    
    Thus the array index in output_PDRs will be st+ii for ii in range(Nt)
    '''
    k_arr    = np.frombuffer(k_dict['arr']).reshape(k_dict['shape'])
    
    WPDR_arr = np.frombuffer(WPDR_dict['arr']).reshape(WPDR_dict['shape'])
    wCGR_arr = np.frombuffer(wCGR_dict['arr']).reshape(wCGR_dict['shape'])
    wVg_arr  = np.frombuffer(wVg_dict['arr']).reshape(wVg_dict['shape'])

    for ii in range(times.shape[0]):
        Species, PP = create_species_array(B0[ii], name, mass, charge, density[:, ii],
                                           tper[:, ii], ani[:, ii], remove_zero_density_species=False)
        this_k      = np.linspace(kmin, kmax, Nk, endpoint=False)
        
        # Convert k-extrema to /m if needed
        if knorm == True: this_k *= PP['pcyc_rad'] / PP['va']
        
        # Calculate dispersion relation
        print('Worker', worker, '::', times[ii])
        this_WPDR, this_wCGR, this_wVg = get_dispersion_relation(Species, this_k, approx=approx, complex_out=False,
                                                       print_filtered=False, return_vg=True)

        # Dump to shared memory
        k_arr[   st+ii, :]       = this_k[...]
        WPDR_arr[st+ii, :, :, :] = this_WPDR[...]
        wCGR_arr[st+ii, :, :]    = this_wCGR[...]
        wVg_arr[ st+ii, :, :]    = this_wVg[ ...]
    return


def get_all_DRs_warm_only(_save_dir, _time_start, _time_end, _probe, _cmp, 
                    kmin=0.0, kmax=1.0, Nk=1000, knorm=True,
                    _nsec=None, HM_filter_mhz=50, N_procs=7,
                    suff='', data_path='E://DATA//RBSP//', output=True,
                    approx='warm', load_only=False):
    '''
    As above, but for the warm approximation only (for speed). Similarly parallelized.
    
    Also modified so composition filenames are saved as tenths of an percentage, i.e.
    a composition of 97.5/0.5/2.0 will be saved with a suffix of 975_005_020
    
    This works for now because no species will have a density of 100%
    i.e. don't have to worry about 1000 (i.e. 100%) blowing out my 3 places.
    
    Maybe put xxx if 100% (since this will happen only for 0% He and 0% Oxygen)
    So it should really only be the one file, so if h_comp == 100.0...
    '''
    
    DR_path = get_DR_filepath(_save_dir, _time_start, _time_end, _cmp, _nsec)
    
    if os.path.exists(DR_path) == False:
        if load_only == True:
            print('No DR exists, returning null')
            return
        
        # Load data
        if _probe == 'crres':
            times, B0, name, mass, charge, density, tper, ani, cold_dens = \
            CRRES_extract_species_arrays(_time_start, _time_end, crres_path='G://DATA//CRRES//',
                           rc_portion=0.05, cold_comp=_cmp, hot_comp=[100., 0., 0.],
                           tperp = [30., 0., 0.], ani=[1.0, 0.0, 0.0],
                           return_raw_ne=True, HM_filter_mhz=50, nsec=None)
        else:
            times, B0, name, mass, charge, density, tper, ani, cold_dens = \
            extract_species_arrays(_time_start, _time_end, _probe, cmp=np.asarray(_cmp), 
                                   return_raw_ne=True, nsec=_nsec, HM_filter_mhz=HM_filter_mhz,
                                   rbsp_path=data_path, HOPE_only=True)
    
        Nt      = times.shape[0]
        
        procs   = []
        
        # Create raw shared memory arrays with shapes. Store in dict to send with each worker
        k_shape      = (Nt, Nk)
        k_shm        = multiprocessing.RawArray('d', Nt * Nk)
        k_dict       = {'arr': k_shm, 'shape': k_shape}

        WPDR_shape   = (Nt, Nk, 3, 2)
        WPDR_shm     = multiprocessing.RawArray('d', Nt*Nk*3*2)
        WPDR_dict    = {'arr': WPDR_shm, 'shape': WPDR_shape}

        wCGR_shape   = (Nt, Nk, 3)
        wCGR_shm     = multiprocessing.RawArray('d', Nt*Nk*3)
        wCGR_dict    = {'arr': wCGR_shm, 'shape': wCGR_shape}

        wVg_shape   = (Nt, Nk, 3)
        wVg_shm     = multiprocessing.RawArray('d', Nt*Nk*3)
        wVg_dict    = {'arr': wVg_shm, 'shape': wVg_shape}
        
        # Create numpy view into shared memory
        k_np         = np.frombuffer(k_shm).reshape(k_shape)
        WPDR_np      = np.frombuffer(WPDR_shm).reshape(WPDR_shape)
        wCGR_np      = np.frombuffer(wCGR_shm).reshape(wCGR_shape)
        wVg_np       = np.frombuffer(wVg_shm).reshape(wVg_shape)
        
        # Split input data into a list of chunks
        time_chunks    = np.array_split(times,   N_procs)
        field_chunks   = np.array_split(B0,      N_procs)
        density_chunks = np.array_split(density, N_procs, axis=1)
        tper_chunks    = np.array_split(tper,    N_procs, axis=1)
        ani_chunks     = np.array_split(ani,     N_procs, axis=1)
    
        # Instatiate each process with a different chunk
        acc = 0; start = time.time()
        for xx in range(N_procs):
            print('Starting process', xx)
            proc = multiprocessing.Process(target=get_DRs_chunked_warm_only,
                                        args=(Nk, kmin, kmax, knorm, time_chunks[xx],
                                        field_chunks[xx], name, mass, charge, density_chunks[xx],
                                        tper_chunks[xx], ani_chunks[xx],
                                        k_dict, WPDR_dict, wCGR_dict, wVg_dict),
                                        kwargs={'st':acc, 'worker':xx, 'approx':approx})
            procs.append(proc)
            proc.start()
            
            acc += time_chunks[xx].shape[0]
        
        # Complete processes
        for proc in procs:
            proc.join()
                
        print('All processes complete')
        end = time.time()
        print('Total parallel time = {}s'.format(str(end-start)))
        
        # Make output complex
        WPDR_out = np.zeros((Nt, Nk, 3), dtype=np.complex128)
    
        for ii in range(Nt):
            for jj in range(Nk):
                for kk in range(3):
                    WPDR_out[ii, jj, kk] = WPDR_np[ii, jj, kk, 0] + 1j * WPDR_np[ii, jj, kk, 1]
            
        wCGR_out = wCGR_np;     wVg_out = wVg_np
            
        # Saves data used for DR calculation as well, for future reference (and plotting)
        if os.path.exists(_save_dir) == False:
            os.makedirs(_save_dir)
                
        print('Saving dispersion history...')
        np.savez(DR_path, all_k=k_np, all_WPDR=WPDR_out, all_wCGR=wCGR_np, all_wVg=wVg_np,
                 comp=np.asarray(_cmp), times=times, B0=B0, name=name, mass=mass, charge=charge,
                 density=density, tper=tper, ani=ani, cold_dens=cold_dens,
                 HM_filter_mhz=np.array([HM_filter_mhz]))
    else:
        if output == False:
            return
        print('Dispersion results already exist, loading from file...')
        DR_file   = np.load(DR_path)
        
        _cmp      = DR_file['comp']
        k_np      = DR_file['all_k']
        WPDR_out  = DR_file['all_WPDR']
        wCGR_out  = DR_file['all_wCGR']
        wVg_out   = DR_file['all_wVg']

        times     = DR_file['times']
        B0        = DR_file['B0']
        name      = DR_file['name']
        mass      = DR_file['mass']
        charge    = DR_file['charge']
        density   = DR_file['density']
        tper      = DR_file['tper']
        ani       = DR_file['ani']
        cold_dens = DR_file['cold_dens']
        
    if output == True:
        return k_np,  WPDR_out, wCGR_out, wVg_out, \
               times, B0, name, mass, charge, density, tper, ani, cold_dens, _cmp
           
           
def calculate_warm_sweep(_rbsp_path, _save_dir, _time_start, _time_end, _probe, _nsec=5, N_procs=7):
    '''
    Code up a version of this that can handle when one of the densities are 
    zero - this will be the get_dispersion_relation_3sp hard-coded for 3 species
    only. Trick will just be getting the initial guesses/gyrofrequencies right,
    and then putting the solutions in the right band. Shouldn't be too hard.
    
    For now, just don't solve anything that doesn't involve all 3 ions.
    
    UPDATE: When solving for any species with zero density, use the hot approximation
    as this will allow the warm components to potentially support the wave propagation
    '''                
    #for he_comp in np.arange(1.0, 30.5, 1.0):
    for o_comp in np.arange(0.0, 10.5, 0.5):
            he_comp = 0.0
            h_comp  = 100. - o_comp - he_comp
            _cmp    = np.array([h_comp, he_comp, o_comp], dtype=float)
            
            if any(cc == 0 for cc in _cmp):
                approx='hot'
            else:
                approx='warm'

            print('Calculating timeseries growth rate for {}/{}/{} cold composition'.format(h_comp, he_comp, o_comp))
            get_all_DRs_warm_only(_save_dir, _time_start, _time_end, _probe, _cmp, 
                kmin=0.0, kmax=1.5, Nk=1000, knorm=True,
                _nsec=_nsec, HM_filter_mhz=50., N_procs=N_procs,
                suff='', data_path=_rbsp_path, output=False, approx=approx)
    return


def get_all_DRs_with_cutoffs(_save_dir, _time_start, _time_end, _probe,  
                    kmin=0.0, kmax=1.0, Nk=1000, knorm=True,
                    _nsec=5, HM_filter_mhz=50, N_procs=7,
                    he_frac=0.01, data_path='E://DATA//RBSP//', output=True,
                    approx='warm', load_only=False):
    '''
    As above, but for the warm approximation only (for speed). Similarly parallelized.
    Plasma composition derived from cutoffs for each point in time.
    
    Set helium composition and oxygen/proton composition sets from cutoffs.
    
    Need to use hot function because sometimes He/O is sometimes zero?
    
    TODO: Calculate density based of total percent, not just cold percentage, since
    cutoffs depend on the whole plasma density composition (but only a few percent
    difference)
    '''
    save_string = _time_start.astype(object).strftime('%Y%m%d_%H%M_') \
                          + _time_end.astype(object).strftime('%H%M')
    DR_path = _save_dir + 'DISPw_{}_withcutoffs_{}sec_{}.npz'.format(save_string, _nsec, approx)

    if os.path.exists(DR_path) == False:
        if load_only == True:
            print('No DR exists, returning null')
            return
        
        # Load data
        if _probe == 'crres':
            times, B0, name, mass, charge, density, tper, ani, cold_dens = \
            CRRES_extract_species_arrays(_time_start, _time_end, crres_path='G://DATA//CRRES//',
                           rc_portion=0.05, hot_comp=[100., 0., 0.],
                           tperp = [30., 0., 0.], ani=[1.0, 0.0, 0.0],
                           return_raw_ne=True, HM_filter_mhz=50, nsec=None)
        else:
            times, B0, name, mass, charge, density, tper, ani, cold_dens = \
            extract_species_arrays(_time_start, _time_end, _probe, 
                                   return_raw_ne=True, nsec=_nsec, HM_filter_mhz=HM_filter_mhz,
                                   rbsp_path=data_path, HOPE_only=True)
        Nt      = times.shape[0]
        procs   = []
        
        # Reset cold density composition based on cutoff file
        cutoff_dict = data.read_cutoff_file(cutoff_filename)
        o_fracs     = data.calculate_o_from_he_and_cutoff(cutoff_dict['CUTOFF_NORM'], he_frac)
        o_fracs     = np.interp(times.astype(np.int64),
                        cutoff_dict['CUTOFF_TIME'].astype(np.int64),
                        o_fracs)
        
        density[2, :] = cold_dens * o_fracs
        density[1, :] = cold_dens * he_frac
        density[0, :] = cold_dens - density[2, :] - density[1, :]

        # Create raw shared memory arrays with shapes. Store in dict to send with each worker
        k_shape      = (Nt, Nk)
        k_shm        = multiprocessing.RawArray('d', Nt * Nk)
        k_dict       = {'arr': k_shm, 'shape': k_shape}

        WPDR_shape   = (Nt, Nk, 3, 2)
        WPDR_shm     = multiprocessing.RawArray('d', Nt*Nk*3*2)
        WPDR_dict    = {'arr': WPDR_shm, 'shape': WPDR_shape}

        wCGR_shape   = (Nt, Nk, 3)
        wCGR_shm     = multiprocessing.RawArray('d', Nt*Nk*3)
        wCGR_dict    = {'arr': wCGR_shm, 'shape': wCGR_shape}

        wVg_shape    = (Nt, Nk, 3)
        wVg_shm      = multiprocessing.RawArray('d', Nt*Nk*3)
        wVg_dict     = {'arr': wVg_shm, 'shape': wVg_shape}
        
        # Create numpy view into shared memory
        k_np         = np.frombuffer(k_shm).reshape(k_shape)
        WPDR_np      = np.frombuffer(WPDR_shm).reshape(WPDR_shape)
        wCGR_np      = np.frombuffer(wCGR_shm).reshape(wCGR_shape)
        wVg_np       = np.frombuffer(wVg_shm).reshape(wVg_shape)
        
        # Split input data into a list of chunks
        time_chunks    = np.array_split(times,   N_procs)
        field_chunks   = np.array_split(B0,      N_procs)
        density_chunks = np.array_split(density, N_procs, axis=1)
        tper_chunks    = np.array_split(tper,    N_procs, axis=1)
        ani_chunks     = np.array_split(ani,     N_procs, axis=1)
    
        #  Instatiate each process with a different chunk
        acc = 0; start = time.time()
        for xx in range(N_procs):
            print('Starting process', xx)
            proc = multiprocessing.Process(target=get_DRs_chunked_warm_only,
                                        args=(Nk, kmin, kmax, knorm, time_chunks[xx],
                                        field_chunks[xx], name, mass, charge, density_chunks[xx],
                                        tper_chunks[xx], ani_chunks[xx],
                                        k_dict, WPDR_dict, wCGR_dict, wVg_dict),
                                        kwargs={'st':acc, 'worker':xx, 'approx':approx})
            procs.append(proc)
            proc.start()
            
            acc += time_chunks[xx].shape[0]
        
        # Complete processes
        for proc in procs:
            proc.join()
                
        print('All processes complete')
        end = time.time()
        print('Total parallel time = {}s'.format(str(end-start)))
        
        # Make output complex
        WPDR_out = np.zeros((Nt, Nk, 3), dtype=np.complex128)
        for ii in range(Nt):
            for jj in range(Nk):
                for kk in range(3):
                    WPDR_out[ii, jj, kk] = WPDR_np[ii, jj, kk, 0] + 1j * WPDR_np[ii, jj, kk, 1]
            
        wCGR_out = wCGR_np; wVg_out = wVg_np
            
        # Saves data used for DR calculation as well, for future reference (and plotting)
        if os.path.exists(_save_dir) == False:
            os.makedirs(_save_dir)
                
        print('Saving dispersion history...')
        np.savez(DR_path, all_k=k_np, all_WPDR=WPDR_out, all_wCGR=wCGR_np, all_wVg=wVg_np,
                 times=times, B0=B0, name=name, mass=mass, charge=charge,
                 density=density, tper=tper, ani=ani, cold_dens=cold_dens,
                 HM_filter_mhz=np.array([HM_filter_mhz]))
    else:
        if output == False:
            return
        print('Dispersion results already exist, loading from file...')
        DR_file   = np.load(DR_path)
        
        k_np      = DR_file['all_k']
        WPDR_out  = DR_file['all_WPDR']
        wCGR_out  = DR_file['all_wCGR']
        wVg_out   = DR_file['all_wVg']

        times     = DR_file['times']
        B0        = DR_file['B0']
        name      = DR_file['name']
        mass      = DR_file['mass']
        charge    = DR_file['charge']
        density   = DR_file['density']
        tper      = DR_file['tper']
        ani       = DR_file['ani']
        cold_dens = DR_file['cold_dens']

    if output == True:
        return k_np,  WPDR_out, wCGR_out, wVg_out, \
               times, B0, name, mass, charge, density, tper, ani, cold_dens
           

#%% KOZYRA FUNCTIONS
@nb.njit()
def convective_growth_rate_kozyra(field, ndensc, ndensw, ANI, temperp,
                                  norm_ampl=0, norm_freq=0, NPTS=1000, maxfreq=1.0):
    '''
    Calculates the convective growth rate S as per eqn. 6 of Kozyra (1984). Plasma parameters passed as 
    length 3 numpy.ndarrays, one for each H+, He+, O+
    
    INPUT:
        field  -- Magnetic field intensity, nT
        ndensc -- Cold plasma densities, /cm3
        ndensw -- Warm plasma densities, /cm3
        ANI    -- Temperature anisotropy of each species
        temperp-- Perpendicular temperature of species warm component, eV
        
    OPTIONAL:
        norm_ampl -- Flag to normalize growth rate to max value (0: No, 1: Yes). Default 0
        norm_freq -- Flag to normalize frequency to proton cyclotron units. Default 0
        NPTS      -- Number of sample points up to maxfreq. Default 500
        maxfreq   -- Maximum frequency to calculate for in proton cyclotron units. Default 1.0
        
    Blows up for unknown reasons sometimes (div0 error in numba):
    B0         = 161.0
    ndensc     = np.array([70.0, 0.0, 0.0])
    ndensw     = np.array([0.7, 0.0, 0.0])
    anisotropy = np.array([0.5, 0.0, 0.0])
    tperp      = np.array([20., 0.0, 0.0])
    '''
    # Perform input checks 
    N = ndensc.shape[0]
        
    # CONSTANTS
    PMASS   = 1.673E-27
    MUNOT   = 1.25660E-6
    EVJOULE = 6.242E18
    CHARGE  = 1.602E-19
    
    # OUTPUT PARAMS
    growth = np.zeros(NPTS)                         # Output growth rate variable
    x      = np.zeros(NPTS)                         # Input normalized frequency
    stop   = np.zeros(NPTS)                         # Stop band flag (0, 1)
    
    # Set here since these are constant. Thought these were wrong because Z is
    # usually atomic number, not charge state (+1 for all ions here)
    M    = np.zeros(N)
    M[0] = 1.0
    M[1] = 4.0
    M[2] = 16.0
    
    # LOOP PARAMS
    step  = maxfreq / float(NPTS)
    FIELD = field*1.0E-9                 # convert to Tesla
        
    NCOLD   = ndensc * 1.0E6
    NHOT    = ndensw * 1.0E6
    etac    = NCOLD / NHOT[0]            # Needs a factor of Z ** 2?
    etaw    = NHOT  / NHOT[0]            # Here too
    numer   = M * (etac+etaw)
    
    # Calculate either beta or tpar/tperp depending on inputs
    TPERP   = temperp / EVJOULE
    TPAR    = TPERP / (1.0 + ANI)
    bet     = NHOT*TPAR / (FIELD*FIELD/(2.0*MUNOT))        
    alpha   = np.sqrt(2.0 * TPAR / PMASS)
    
    # Loop for calculation of CGR
    for k in range(1, NPTS):
          x[k]   = k*step
          denom  = 1.0 - M*x[k]
          
          sum1  = 0.0
          prod2 = 1.0
          
          for i in range(N):
               prod2   *= denom[i]
               prod     = 1.0
               temp     = denom[i]
               denom[i] = numer[i]
               
               for j in range(N):
                   prod *= denom[j]
               
               sum1    += prod
               denom[i] = temp

          sum2 = 0.0
          arg4 = prod2 / sum1
    
          # Check for stop band.
          if (arg4 < 0.0) and (x[k] > 1.0/M[N-1]):
              growth[k] = 0.0
              stop[k]   = 1
          else:
             arg3 = arg4 / (x[k] ** 2)
             
             for i in range(N):
                if (NHOT[i] > 1.0E-3):
                     arg1  = np.sqrt(np.pi) * etaw[i] / ((M[i]) ** 2 * alpha[i])       # Outside term
                     arg1 *= ((ANI[i] + 1.0) * (1.0 - M[i]*x[k]) - 1.0)                # Inside square brackets (multiply by outside)
                     arg2  = (-etaw[i] / M[i]) * (M[i]*x[k] - 1.0) ** 2 / bet[i]*arg3
                     
                     sum2 += arg1*np.exp(arg2)
             
             growth[k] = sum2*arg3/2.0
    
    ###########################
    ### NORMALIZE AND CLEAN ###
    ###########################    
    for ii in range(NPTS):
        if (growth[ii] < 0.0):
            growth[ii] = 0.0
          
    if (norm_freq == 0):
        cyclotron  = CHARGE*FIELD/(2.0*np.pi*PMASS)
        x         *= cyclotron
          
    if (norm_ampl == 1):
        growth /= growth.max()
    else:
        growth *= 1e9
    return x, growth, stop

     
def get_all_CGRs_kozyra(_time_start, _time_end, _probe, _cmp=np.array([70, 20, 10]), 
                        fmax_pcyc=1.0, Nf=1000, _nsec=None, HM_filter_mhz=50, instr='HOPE'):
    '''
    Plugs satellite values into Kozyra growth rate equation (not mine, the validated old one)
    for each time in the moment timeseries. Do one at a time (HOPE, RBSPICE) and then add them
    together to see if they linearly work together. Could do a 3-plot of this in time, or some
    sort of 3-plot pcolormesh timeseries.
    
    Input params from data (needed units for CGR calc):
        field    -- Magnetic field intensity, nT
        ndensc   -- Cold plasma densities, /cm3
        ndensw   -- Warm plasma densities, /cm3
        ANI      -- Temperature anisotropy of each species
        temperp  -- Perpendicular temperature of species warm component, eV
        
    'instr' set to be HOPE or RBSPICE depending on which hot population is desired.
    '''
    save_string = _time_start.astype(object).strftime('%Y%m%d_%H%M_') \
                + _time_end.astype(object).strftime('%H%M')
    if _nsec is None:
        DR_path = save_dir + 'KCGR_{}_cc_{:03}_{:03}_{:03}.npz'.format(save_string, int(_cmp[0]), int(_cmp[1]), int(_cmp[2]))
    else:
        DR_path = save_dir + 'KCGR_{}_cc_{:03}_{:03}_{:03}_{}sec.npz'.format(save_string, int(_cmp[0]), int(_cmp[1]), int(_cmp[2]), _nsec)
    
    # Density.shape = 9, 201
    
    if os.path.exists(DR_path) == False:
        times, B0, name, mass, charge, density, tper, anisotropy, cold_dens = \
        extract_species_arrays(_time_start, _time_end, _probe, cmp=np.asarray(_cmp), 
                               return_raw_ne=True, nsec=_nsec, HM_filter_mhz=HM_filter_mhz)

        field  = B0*1e9
        ndensc = density[:3, :]*1e-6
        
        hope_ndensw   = density[3:6, :]*1e-6
        hope_tempperp = tper[3:6, :]
        hope_ANI      = anisotropy[3:6, :]

        spice_ndensw   = density[6:, :]*1e-6
        spice_tempperp = tper[6:, :]
        spice_ANI      = anisotropy[6:, :]


        # Still need to define frequency array since PCYC changes with B0 (which changes bins)
        Nt            = times.shape[0]
        all_CGR_HOPE  = np.zeros((Nt, Nf), dtype=np.float64)
        all_stop_HOPE = np.zeros((Nt, Nf), dtype=np.float64)
        all_CGR_SPICE = np.zeros((Nt, Nf), dtype=np.float64)
        all_stop_SPICE= np.zeros((Nt, Nf), dtype=np.float64)
        all_f         = np.zeros((Nt, Nf), dtype=np.float64)
        
        # Get dispersion relations for each time if possible
        for ii in range(Nt):
            print('Calculating Kozyra CGR relation for {}'.format(times[ii]))
            all_f[ii], all_CGR_HOPE[ii], all_stop_HOPE[ii] = convective_growth_rate_kozyra(field[ii], ndensc[:, ii],
                                                 hope_ndensw[:, ii], hope_ANI[:, ii], temperp=hope_tempperp[:, ii],
                                                 norm_ampl=0, NPTS=Nf, maxfreq=1.0)
            
            all_f[ii], all_CGR_SPICE[ii], all_stop_SPICE[ii] = convective_growth_rate_kozyra(field[ii], ndensc[:, ii],
                                                 spice_ndensw[:, ii], spice_ANI[:, ii], temperp=spice_tempperp[:, ii],
                                                 norm_ampl=0, NPTS=Nf, maxfreq=1.0)

        # Saves data used for DR calculation as well, for future reference (and plotting)
        if os.path.exists(save_dir) == False:
            os.makedirs(save_dir)
                
        print('Saving dispersion history...')
        np.savez(DR_path, all_f=all_f, all_stop_HOPE=all_stop_HOPE, all_CGR_HOPE=all_CGR_HOPE,
                 all_stop_SPICE=all_stop_SPICE,  all_CGR_SPICE=all_CGR_SPICE, comp=_cmp,
                 times=times, B0=B0, name=name, mass=mass, charge=charge, density=density, tper=tper,
                 ani=anisotropy, cold_dens=cold_dens, HM_filter_mhz=np.array([HM_filter_mhz]))
    else:
        print('Dispersion results already exist, loading from file...')
        DR_file   = np.load(DR_path)
        
        all_f          = DR_file['all_f']
        all_stop_HOPE   = DR_file['all_stop_HOPE']
        all_CGR_HOPE   = DR_file['all_CGR_HOPE']
        all_stop_SPICE = DR_file['all_stop_SPICE']
        all_CGR_SPICE  = DR_file['all_CGR_SPICE']
                
        times     = DR_file['times']
        B0        = DR_file['B0']
        name      = DR_file['name']
        mass      = DR_file['mass']
        charge    = DR_file['charge']
        density   = DR_file['density']
        tper      = DR_file['tper']
        anisotropy= DR_file['ani']
        cold_dens = DR_file['cold_dens']
    return all_f, all_CGR_HOPE, all_stop_HOPE, all_CGR_SPICE, all_stop_SPICE, \
           times, B0, name, mass, charge, density, tper, anisotropy, cold_dens

     
#%% VALIDATION PLOTS (FOR DIAGNOSTICS)
def validation_plots_chen_2013():
    '''
    Routine to easily compare output from dispersion solver to plots from
    Chen et al. (2013) by varying helium density and temperature.
    
    R-mode solution not incorporated into this version of the code (since
    we don't really care about R-waves).
    '''
    mp       = 1.673e-27                        # Proton mass (kg)
    qi       = 1.602e-19                        # Elementary charge (C)
    _n0      = 100e6                            # Electron number density in /m3                      
    _B0      = 144e-9                           # Background magnetic field in T
    nhh      = 0.03                             # Fraction hot hydrogen
    nHe      = 0.05                             # Fraction warm helium
    THe      = 100.0                             # Helium temperature (eV) -- Does this have to be altered for 'total temp'?
    
    _name    = np.array(['Hot H'  , 'Cold H'        , 'Cold He'])               # Species label
    _mass    = np.array([1.0      , 1.0             , 4.0      ]) * mp          # Mass   in proton masses (kg)
    _charge  = np.array([1.0      , 1.0             , 1.0      ]) * qi          # Change in elementary units (C)
    _density = np.array([nhh      , 1.0 - nhh - nHe , nHe      ]) * _n0         # Density as a fraction of n0 (/m3)
    _tpar    = np.array([25e3     , 0.0             , THe      ])               # Parallel temperature in eV
    _ani     = np.array([1.0      , 0.0             , 0.0      ])               # Temperature anisotropy
    _tper    = (_ani + 1) * _tpar                                               # Perpendicular temperature in eV
    
    _Spec, _PP = create_species_array(_B0, _name, _mass, _charge, _density, _tper, _ani)

    _kh  = np.sqrt(_n0 * qi ** 2 / (mp * e0))/c
    _Nk  = 1000
    _k   = np.linspace(0.0, 2.0*_kh, _Nk)
    pcyc = qi * _B0 / mp 

    cold_DR, cold_CGR = get_dispersion_relation(_Spec, _k, approx='cold')
    warm_DR, warm_CGR = get_dispersion_relation(_Spec, _k, approx='warm')
    hot_DR ,  hot_CGR = get_dispersion_relation(_Spec, _k, approx='hot' )

    # Plot the things
    k_norm    = _k / _kh
    fig, axes = plt.subplots(2, sharex=True)
    
    for jj in range(hot_DR.shape[1]):
        axes[0].semilogx(k_norm,  hot_DR[:, jj].real / pcyc, ls='-' , c='r', lw=0.75)
        axes[0].semilogx(k_norm, warm_DR[:, jj].real / pcyc, ls='--', c='b', lw=0.75)
        axes[0].semilogx(k_norm, cold_DR[:, jj].real / pcyc, ls='--', c='k', lw=0.75)
        axes[0].set_ylabel('$\omega_r/\Omega_h$', fontsize=16)
        
        axes[1].semilogx(k_norm,  hot_DR[:, jj].imag / pcyc, ls='-' , c='r', label='Full', lw=0.75)
        axes[1].semilogx(k_norm, warm_DR[:, jj].imag / pcyc, ls='--', c='b', label='Warm', lw=0.75)
        axes[1].semilogx(k_norm, cold_DR[:, jj].imag / pcyc, ls='--', c='k', label='Cold', lw=0.75)
        
    axes[1].set_ylabel('$\gamma/\Omega_h$', fontsize=16)
    axes[1].set_xlabel('$k_\parallel/k_h$', fontsize=16)
    
    axes[0].set_ylim(0.0, 1.0)
    axes[0].set_xlim(1e-2, 2)
    axes[1].set_ylim(-0.01, 0.06)
    axes[1].set_xlim(1e-2, 2)
    
    fig.subplots_adjust(hspace=0.01)
    return


def validation_plots_wang_2016():
    '''
    Routine to easily compare output from dispersion solver to plots from
    Wang et al. (2016) by varying many parameters. Solver is derived from
    Chen et al. (2013) but plasma is plasma.
    '''
    L_shell  = 4                                # L-shell at which magnetic field and density are calculated
    n0       = sheely_plasmasphere(L_shell)     # Plasma density, /m3
    B0       = geomagnetic_magnitude(L_shell)   # Background magnetic field, T
    mp       = 1.673e-27                        # Proton mass (kg)
    qi       = 1.602e-19                        # Elementary charge (C)
    
    # This all must add up to 1
    RC_ab= 0.1
    H_ab = 0.6
    He_ab= 0.2
    O_ab = 0.1
    
    name    = np.array(['Warm H'  , 'Cold H' , 'Cold He', 'Cold O'])
    mass    = np.array([1.0       , 1.0      , 4.0      , 16.0    ]) * mp
    charge  = np.array([1.0       , 1.0      , 1.0      ,  1.0    ]) * qi
    density = np.array([RC_ab     , H_ab     , He_ab    ,  O_ab,  ]) * n0
    tpar    = np.array([25e3      , 0.0      , 0.0      ,  0.0    ])
    ani     = np.array([2.0       , 0.0      , 0.0      ,  0.0    ])
    tper    = (ani + 1) * tpar

    Spec, PP = create_species_array(B0, name, mass, charge, density, tper, ani)
    
    knorm_fac             = PP['pcyc_rad'] / PP['va']
    k_vals                = np.linspace(0.0, 1.0, 1000, endpoint=False) * knorm_fac

    CPDR_solns,  cold_CGR = get_dispersion_relation(Spec, k_vals, approx='cold' )
    HPDR_solns,   hot_CGR = get_dispersion_relation(Spec, k_vals, approx='hot' )

    CPDR_solns /= PP['pcyc_rad']
    HPDR_solns /= PP['pcyc_rad']   
    k_vals     *= PP['va'] / PP['pcyc_rad']
    
    species_colors = ['r', 'b', 'g']
    
    print('Plotting solutions...')
    plt.ioff()
    plt.figure(figsize=(15, 10))
    ax1 = plt.subplot2grid((2, 2), (0, 0), rowspan=2)
    ax2 = plt.subplot2grid((2, 2), (0, 1), rowspan=2)
    
    for ii in range(CPDR_solns.shape[1]):
        ax1.plot(k_vals[1:], CPDR_solns[1:, ii].real,      c=species_colors[ii], linestyle='--', label='Cold')
        ax1.plot(k_vals[1:], HPDR_solns[1:, ii].real, c=species_colors[ii], linestyle='-',  label='Hot')

    ax1.set_title('Dispersion Relation')
    ax1.set_xlabel(r'$kv_A / \Omega_p$')
    ax1.set_ylabel(r'$\omega_r/\Omega_p$')
    ax1.set_xlim(k_vals[0], k_vals[-1])
    
    ax1.set_ylim(0, 1.0)
    ax1.minorticks_on()
    
    for ii in range(CPDR_solns.shape[1]):
        ax2.plot(k_vals[1:], HPDR_solns[1:, ii].imag, c=species_colors[ii], linestyle='-')

    ax2.set_title('Temporal Growth Rate')
    ax2.set_xlabel(r'$kv_A / \Omega_p$')
    ax2.set_ylabel(r'$\gamma/\Omega_p$')
    ax2.set_xlim(k_vals[0], k_vals[-1])
    ax2.set_ylim(-0.05, 0.05)
    
    ax2.minorticks_on()

    figManager = plt.get_current_fig_manager()
    figManager.window.showMaximized() 
    return


def validation_plots_fraser_1996():
    '''
    Checking CGR solutions against solutions calculated by Fraser et al. (1996)
    using the Kozyra code and equations.
        
    Verdict: Shape is correct, but amplitudes are not. Off by approximately 3.4-6.0
    Factor is not consistent with changing anisotropy (I think?)
    
    Lowest anisotropy (0.5) is also missing highest band? Looks like it's been 
    damped out to below 0.0. Need more literature to compare (convective) growth
    rates with. Run with it on a trial basis, maybe integrate the Samson code
    into this later on.
    '''
    mp         = 1.673e-27                        # Proton mass (kg)
    qi         = 1.602e-19                        # Elementary charge (C)
    _B0        = 300e-9                           # Background magnetic field in T
    A_style    = [':', '-', '--']                 # Line style for each anisotropy
    Ah         = [0.5]                  # Anisotropy values
    
    _name      = np.array(['Cold H', 'Cold He', 'Cold O', 'Warm H', 'Warm He', 'Warm O'])         # Species label
    _mass      = np.array([1.0     , 4.0      , 16.0    , 1.0     , 4.0      , 16.0    ]) * mp    # Mass   in proton masses (kg)
    _charge    = np.array([1.0     , 1.0      , 1.0     , 1.0     , 1.0      , 1.0     ]) * qi    # Change in elementary units (C)
    _density   = np.array([196.0   , 22.0     , 2.0     , 5.1     , 0.05     , 0.13    ]) * 1e6   # Density in cm3 (/m3)
    _tper      = np.array([0.0     , 0.0      , 0.0     , 30.0    , 10.0     , 10.0    ]) * 1e3   # Parallel temperature in keV (eV)
    
    plt.ioff()
    fig, axes = plt.subplots(ncols=2, nrows=1, figsize=(16, 10))
    
    for ii in range(len(Ah)):
        _ani       = np.array([0.0     , 0.0      , 0.0     , Ah[ii], 1.0      , 1.0     ])         # Temperature anisotropy
        _Spec, _PP = create_species_array(_B0, _name, _mass, _charge, _density, _tper, _ani)
    
        knorm_fac  = _PP['pcyc_rad'] / _PP['va']
        k_vals     = np.linspace(0.0, 2.0, 5000, endpoint=False) * knorm_fac
    
        # Ouputs are (k_vals, N_solns)
        cold_DR, cold_CGR = get_dispersion_relation(_Spec, k_vals, approx='warm')
        cold_DR /= 2*np.pi
    
        # Plot each solution
        for jj in range(cold_DR.shape[1]):
            axes[0].plot(k_vals             , cold_DR[:, jj].real, ls=A_style[ii] , c='k', lw=0.75)
            axes[1].plot(cold_DR[:, jj].real, cold_CGR[:, jj]*1e7, ls=A_style[ii] , c='k', lw=0.75)
        
    fsize = 12; lpad = 10
    axes[0].set_xlabel(r'$k_\parallel$', fontsize=fsize)
    axes[0].set_ylabel(r'$\omega_r$'   , fontsize=fsize, rotation=0, labelpad=lpad)
    
    axes[1].set_ylabel('$S$ \n $\\times 10^7 m^{-1}$', fontsize=fsize, rotation=0, labelpad=lpad)
    axes[1].set_xlabel(r'$\omega_r$', fontsize=fsize)
    
    axes[0].set_xlim(0.0, k_vals[-1])
    #axes[0].set_ylim(0.0, None)
    
    axes[1].set_xlim(0.0, None)
    #axes[1].set_ylim(0.0, None)
    plt.show()
    return


def validation_plots_omura2010():
    '''
    Using Omura parameters to test against his group velocities, just to see if
    the minus signs from Chen (2013) are ok or not.
    
    Original Omura functions in here for comparison (since we know they work).
    
    Group velocities equal to within around 0.02%
    '''
    def get_gamma_c(w, Species):
        # Electron bit (gyfreq is signed, so this is a minus)
        cold_sum = Species[-1]['plasma_freq_sq'] / Species[-1]['gyrofreq']
        
        # Sum over ion species
        for ii in range(Species.shape[0] - 1):
            cold_sum += Species[ii]['plasma_freq_sq'] / (Species[ii]['gyrofreq'] - w)
        return cold_sum

    def get_group_velocity(w, k, Species):
        gam_c = get_gamma_c(w, Species) 
        
        # Ions only?
        ion_sum = 0.0
        for ii in range(Species.shape[0] - 1):
            ion_sum += Species[ii]['plasma_freq_sq'] / (Species[ii]['gyrofreq'] - w) ** 2
            
        denom = gam_c + w*ion_sum
        
        Vg = np.zeros(w.shape)
        for mm in range(w.shape[1]):
            Vg[:, mm] = 2 * c * c * k / denom[:, mm]
        return Vg

    def get_resonance_velocity(w, k, PP):
        Vr = np.zeros(w.shape)
        for mm in range(w.shape[1]):
            Vr[:, mm] = (w[:, mm] - PP['pcyc_rad']) / k
        return Vr
    
    def get_phase_velocity(w, k):
        Vp = np.zeros(w.shape)
        for mm in range(w.shape[1]):
            Vp[:, mm] = w[:, mm] / k
        return Vp
    
    def get_velocities(w, k, Species, PP):
        Vg = get_group_velocity(w, k, Species)
        Vr = get_resonance_velocity(w, k, PP)
        Vp = get_phase_velocity(w, k)
        return Vg, Vp, Vr

    
    # Parameters in SI units (Note: Added the hot bit here. Is it going to break anything?) nh = 7.2
    kB    = 1.380649e-23
    pcyc  = 3.7 # Hz
    B0    = 2 * np.pi * mp * pcyc / qp
    
    Th_para  = (mp * (6e5)**2 / kB) / 11603.
    Th_perp  = (mp * (8e5)**2 / kB) / 11603.
    Ah       = Th_perp / Th_para - 1

    name    = np.array(['H'    , 'He'  , 'O'  , 'Hot H'])
    mass    = np.array([1.0    , 4.0   , 16.0 , 1.0    ]) * mp
    charge  = np.array([1.0    , 1.0   , 1.0  , 1.0    ]) * qp
    density = np.array([136.8  , 17.0  , 17.0 , 7.2    ]) * 1e6
    ani     = np.array([0.0    , 0.0   , 0.0  , Ah])
    tpar    = np.array([0.0    , 0.0   , 0.0  , Th_para])
    tper    = (ani + 1) * tpar
    
    _Spec, _PP = create_species_array(B0, name, mass, charge, density, tper, ani)
    
    wlength                     = 4e4
    kmax                        = 2 * np.pi / wlength
    k_vals                      = np.linspace(0.0, kmax*1.5, 5000, endpoint=False)
    cold_DR, cold_CGR, cold_Vg  = get_dispersion_relation(_Spec, k_vals, approx='cold', return_vg=True) 
    
    plt.ioff()
    
# =============================================================================
#     # Plot CPDR
#     fig, ax = plt.subplots()
#     for ss in range(cold_DR.shape[1]):
#         ax.plot(1e3*k_vals / (2*np.pi), cold_DR.real[:, ss] / (2 * np.pi), c='k')
#         
#     ax.set_xlim(0, 0.025)
#     ax.set_ylim(0, 3.5)
#     
#     ax.set_title('Cold Plasma Dispersion Relation')
#     ax.set_ylabel('f (Hz)', fontsize=14)
#     ax.set_xlabel(r'$\frac{1}{\lambda}$ ($km^{-1}$)', rotation=0, fontsize=14, labelpad=30)
# =============================================================================
    

    # Resonance and Group Velocities plot (Figure 4a,b validated)
    f_max  = 4.0
    w_vals = cold_DR.real
    f_vals = cold_DR.real / (2*np.pi)
    
    V_group, V_phase, V_resonance = get_velocities(w_vals, k_vals, _Spec, _PP)

# =============================================================================
#     fig, ax = plt.subplots()
#     ax.plot(f_vals, V_resonance/1e3)
#     ax.set_xlabel('f (Hz)', fontsize=14)
#     ax.set_ylabel('$V_R$\nkm/s', rotation=0, fontsize=14, labelpad=30)
#     ax.set_xlim(0, f_max)
#     ax.set_ylim(-1500, 0)
# =============================================================================

    fig, ax = plt.subplots()
    ax.plot(f_vals, V_group/1e3, c='k', label='Omura')
    ax.plot(f_vals, cold_Vg/1e3, c='r', label='Chen')
    ax.set_xlabel('f (Hz)', fontsize=14)
    ax.set_ylabel('$V_g$\nkm/s', rotation=0, fontsize=14, labelpad=30)
    ax.set_xlim(0, f_max)
    ax.set_ylim(0, 250)
    
# =============================================================================
#     fig, ax = plt.subplots()
#     ax.plot(f_vals, V_phase/1e3)
#     ax.set_xlabel('f (Hz)', fontsize=14)
#     ax.set_ylabel('$V_p$\nkm/s', rotation=0, fontsize=14, labelpad=30)
#     ax.set_xlim(0, f_max)
#     #ax.set_ylim(0, 250)
# =============================================================================

    plt.show()
    return
     

def hybrid_test_plot():
    '''
    Quick check to see if growth would be expected from a set of inputs or not
    
    # Might need revalidation? Or heavy ions do more than I thought.
    # Actually no, it's just not as well behaved if I don't set tper as 
    #    zero for the cold populations
    '''
    B0      = 200e-9                           # Background magnetic field, T
    mp      = 1.673e-27                        # Proton mass (kg)
    qi      = 1.602e-19                        # Elementary charge (C)
    #kB      = 1.380649e-23
    
    name    = np.array(['Warm H'  , 'Cold H' , 'Cold He', 'Cold O'])
    mass    = np.array([1.0       , 1.0      , 4.0      , 16.0    ]) * mp
    charge  = np.array([1.0       , 1.0      , 1.0      ,  1.0    ]) * qi
    density = np.array([20.       , 120.     , 40.      ,  20.    ]) * 1e6
    tper    = np.array([40e3      , 0.0      , 0.0      ,  0.0    ])
    ani     = np.array([1.0       , 0.0      , 0.0      ,  0.0    ])

    Spec, PP = create_species_array(B0, name, mass, charge, density, tper, ani)
    
    knorm_fac             = PP['pcyc_rad'] / PP['va']
    k_vals                = np.linspace(0.0, 2.0, 1000, endpoint=False) * knorm_fac

    CPDR_solns,  cold_CGR = get_dispersion_relation(Spec, k_vals, approx='cold' )
    WPDR_solns,  warm_CGR = get_dispersion_relation(Spec, k_vals, approx='warm' )
    HPDR_solns,   hot_CGR = get_dispersion_relation(Spec, k_vals, approx='hot' )

    CPDR_solns /= PP['pcyc_rad']
    WPDR_solns /= PP['pcyc_rad']
    HPDR_solns /= PP['pcyc_rad']   
    k_vals     *= PP['va'] / PP['pcyc_rad']

    species_colors = ['r', 'b', 'g']
    
    print('Plotting solutions...')
    plt.ioff()
    plt.figure(figsize=(15, 10))
    ax1 = plt.subplot2grid((2, 2), (0, 0), rowspan=2)
    ax2 = plt.subplot2grid((2, 2), (0, 1), rowspan=2)
    
    for ii in range(CPDR_solns.shape[1]):
        ax1.plot(k_vals[1:], CPDR_solns[1:, ii].real, c=species_colors[ii], linestyle='--', label='Cold')
        ax1.plot(k_vals[1:], WPDR_solns[1:, ii].real, c=species_colors[ii], linestyle=':',  label='Warm')
        ax1.plot(k_vals[1:], HPDR_solns[1:, ii].real, c=species_colors[ii], linestyle='-',  label='Hot')

    ax1.set_title('Dispersion Relation')
    ax1.set_xlabel(r'$kv_A / \Omega_p$')
    ax1.set_ylabel(r'$\omega_r/\Omega_p$')
    ax1.set_xlim(k_vals[0], k_vals[-1])
    
    ax1.set_ylim(0, 1.0)
    ax1.minorticks_on()
    
    for ii in range(HPDR_solns.shape[1]):
        ax2.plot(k_vals[1:], CPDR_solns[1:, ii].imag, c=species_colors[ii], linestyle='--', label='Cold')
        ax2.plot(k_vals[1:], WPDR_solns[1:, ii].imag, c=species_colors[ii], linestyle=':',  label='Warm')
        ax2.plot(k_vals[1:], HPDR_solns[1:, ii].imag, c=species_colors[ii], linestyle='-',  label='Hot')

    glim = 0.2
    ax2.set_title('Temporal Growth Rate')
    ax2.set_xlabel(r'$kv_A / \Omega_p$')
    ax2.set_ylabel(r'$\gamma/\Omega_p$')
    ax2.set_xlim(k_vals[0], k_vals[-1])
    ax2.set_ylim(-glim, glim)
    
    ax2.minorticks_on()

    figManager = plt.get_current_fig_manager()
    figManager.window.showMaximized() 
    return


#%% PLOTTING FUNCTIONS
def create_band_legend(fn_ax, labels, colors):
    legend_elements = []
    for label, color in zip(labels, colors):
        legend_elements.append(Line2D([0], [0], color=color, lw=1, label=label))
        
    new_legend = fn_ax.legend(handles=legend_elements, loc='upper right')
    return new_legend


def create_type_legend(fn_ax, labels, linestyles, type_alpha):
    legend_elements = []
    for label, style, alpha in zip(labels, linestyles, type_alpha):
        legend_elements.append(Line2D([0], [0], color='k', lw=1, label=label, linestyle=style, alpha=alpha))
        
    new_legend = fn_ax.legend(handles=legend_elements, loc='upper left')
    return new_legend


def set_figure_text(ax, ii, field, name, mi, qi, ni, t_perp, A, ne):
    '''
    To Do:
        -- Add 'cold composition' section with percentages
    '''    
    #pdb.set_trace()
    TPER_kev = t_perp * 1e-3
    DENS_cc  = ni * 1e-6
    
    font    = 'monospace'
    fsize   = 9
    top     = 1.0               # Top of text block
    left    = 1.10              # Left boundary of text block
    v_space = 0.02              # Vertical spacing between lines
    
    ax.text(left + 0.08, top - 0.02, 'B0  = {:5.2f} nT'.format(field*1e9),           transform=ax.transAxes, fontsize=fsize, fontname=font)
    ax.text(left + 0.08, top - 0.04, 'n0  = {:5.2f} cm3'.format(ne*1e-6), transform=ax.transAxes, fontsize=fsize, fontname=font)

    pp_top = top - 0.10         # Top of plasma params list
    
    # Plasma Population List
    ax.text(left, pp_top + 0.7*v_space, r'     SPECIES      DENS     TPER       A ', transform=ax.transAxes, fontsize=fsize, fontname=font)
    ax.text(left, pp_top          ,     r'                   CM3      KEV ', transform=ax.transAxes, fontsize=fsize, fontname=font)

    for jj in range(name.shape[0]):
        pname = name[jj].replace('$', '')
        pname = pname.replace('{', '')
        pname = pname.replace('}', '')
        pname = pname.replace('^', '')
        
        ax.text(left, pp_top - (jj+1)*v_space, '{:>12}   {:>7.3f}   {:>7.2f}   {:>6.3f}'.format(pname, round(DENS_cc[jj], 3), round(TPER_kev[jj], 2), round(A[jj], 3)) , transform=ax.transAxes, fontsize=fsize, fontname=font)
    return


def solve_and_plot_one_time(_time_start, _time_end, _probe, _cmp, 
                                kmin=0.0, kmax=1.0, Nk=5000, knorm=True,
                                _nsec=None, suff='', HM_filter_mhz=50):
    
    ii        = 0
    
    save_string = _time_start.astype(object).strftime('%Y%m%d_%H%M_') \
                + _time_end.astype(object).strftime('%H%M')
                
    # This assumes the data was processed correctly in this save file
    DR_path = save_dir + 'DISP_{}_cc_{:03}_{:03}_{:03}{}.npz'.format(save_string, int(_cmp[0]), int(_cmp[1]), int(_cmp[2]), suff)
    DR_file = np.load(DR_path)
            
    times     = DR_file['times']
    B0        = DR_file['B0']
    name      = DR_file['name']
    mass      = DR_file['mass']
    charge    = DR_file['charge']
    density   = DR_file['density']
    tper      = DR_file['tper']
    ani       = DR_file['ani']
    #cold_dens = DR_file['cold_dens']
    
    print('Calculating dispersion/growth relation for {}'.format(times[ii]))
    Species, PP = create_species_array(B0[ii], name, mass, charge, density[:, ii], tper[:, ii], ani[:, ii])
    k_vals      = np.linspace(kmin, kmax, Nk, endpoint=False)
    pcyc        = PP['pcyc_rad']

    # Calculate dispersion relation 3 ways
    CPDR, cold_CGR = get_dispersion_relation(Species, k_vals*PP['pcyc_rad'] / PP['va'], approx='cold')
    WPDR, warm_CGR = get_dispersion_relation(Species, k_vals*PP['pcyc_rad'] / PP['va'], approx='warm')
    HPDR,  hot_CGR = get_dispersion_relation(Species, k_vals*PP['pcyc_rad'] / PP['va'], approx='hot' )
    
    print('Plotting...')
    species_colors = ['r', 'b', 'g']
    
    plt.ioff()
    plt.figure(figsize=(15, 10))
    ax1 = plt.subplot2grid((2, 2), (0, 0), rowspan=2)
    ax2 = plt.subplot2grid((2, 2), (0, 1), rowspan=2)
    #pdb.set_trace()
    for ii in range(CPDR.shape[1]):
        ax1.plot(k_vals, CPDR[:, ii].real/pcyc, c=species_colors[ii], linestyle='--', label='Cold')
        ax1.plot(k_vals, WPDR[:, ii].real/pcyc, c=species_colors[ii], linestyle='-',  label='Warm')
        ax1.plot(k_vals, HPDR[:, ii].real/pcyc, c=species_colors[ii], linestyle=':',  label='Hot')

    ax1.set_title('Dispersion Relation')
    ax1.set_ylabel(r'$\omega/\Omega_H$', rotation=0)
    ax1.set_xlabel(r'$k v_A / \Omega_H$')
    ax1.set_xlim(k_vals[0], k_vals[-1])
    
    ax1.set_ylim(0, 1.0)
    ax1.minorticks_on()
    
# =============================================================================
#     type_label  = ['Cold Plasma Approx.', 'Hot Plasma Approx.']
#     type_style  = ['--', '-']
#     type_legend = create_type_legend(ax1, type_label, type_style)
#     ax1.add_artist(type_legend)
#     
#     band_labels = [r'$H^+$', r'$He^+$', r'$O^+$']
#     band_legend = create_band_legend(ax2, band_labels, species_colors)
#     ax2.add_artist(band_legend)
# =============================================================================
    
    for ii in range(CPDR.shape[1]):
        ax2.plot(k_vals, CPDR[:, ii].imag/pcyc, c=species_colors[ii], linestyle='--', label='Cold')
        ax2.plot(k_vals, WPDR[:, ii].imag/pcyc, c=species_colors[ii], linestyle='-',  label='Warm')
        ax2.plot(k_vals, HPDR[:, ii].imag/pcyc, c=species_colors[ii], linestyle=':',  label='Hot')

    ax2.set_title('Temporal Growth Rate')
    ax2.set_ylabel(r'$\gamma/\Omega_H$', rotation=0)
    ax2.set_xlabel(r'$k v_A / \Omega_H$')
    ax2.set_xlim(k_vals[0], k_vals[-1])
    ax2.set_ylim(-0.05, 0.05)
    ax2.minorticks_on()

    figManager = plt.get_current_fig_manager()
    figManager.window.showMaximized() 
    return


def plot_max_growth_rate_with_time(times, k_vals, all_CPDR, all_WPDR, all_HPDR,
                                   save=False, norm_w=False, B0=None,
                                   ccomp=[70, 20, 10], suff='', ignore_damping=True,
                                   plot_pc1=False, plot_pearls=False):
    plot_dir = save_dir + '//MAX_GR_PLOTS{}//'.format(suff)
    if os.path.exists(plot_dir) == False:
        os.makedirs(plot_dir)
        
    save_string = time_start.astype(object).strftime('%Y%m%d_%H%M_') \
                + time_end.astype(object).strftime('%H%M')
    
    tick_label_size = 14
    mpl.rcParams['xtick.labelsize'] = tick_label_size 
    
    species_colors = ['r', 'b', 'g']
    band_labels    = [r'$H^+$', r'$He^+$', r'$O^+$']
    fontsize       = 18
    
    # Convert to mHz from rad/s
    #CPDR   = all_CPDR / (2*np.pi)
    WPDR   = all_WPDR / (2*np.pi)
    #HPDR   = all_HPDR / (2*np.pi)
    
    if plot_pc1 == True:
        fft_times, fft_freqs, pc1_power = \
                data.get_pc1_spectra(rbsp_path, time_start, time_end, probe,
                                     pc1_res=10.0, overlap=0.99, high_pass_mHz=50.0)
    else:
        fft_times, fft_freqs, pc1_power = None, None, None
        
    if plot_pearls == True:
        print('Plotting pearls...')
        pidx, pearl_times, NULL = rfr.get_pearl_times(time=time_start, crres=False, custom_txtname=None)

    
    #for PDR, lbl in zip([CPDR, WPDR, HPDR], ['cold', 'warm', 'hot']):
    for PDR, lbl in zip([WPDR], ['warm']):
        Nt    = times.shape[0]
        max_f = np.zeros((Nt, 3))
        max_k = np.zeros((Nt, 3))
        max_g = np.zeros((Nt, 3))
            
        # Extract max k and max growth rate for each time, band
        # Must be an issue here?
        for ii in range(Nt):
            for jj in range(3):
                    if any(np.isnan(PDR.real[ii, 1:, jj]) == True):
                        max_f[ii, jj] = np.nan
                        max_k[ii, jj] = np.nan
                        max_g[ii, jj] = np.nan
                    else:
                        max_idx       = np.where(PDR.imag[ii, 1:, jj] == PDR.imag[ii, 1:, jj].max())[0][0]
                        max_f[ii, jj] = PDR.real[ii, max_idx, jj]
                        max_k[ii, jj] = k_vals[ii, max_idx]
                        max_g[ii, jj] = PDR.imag[ii, max_idx, jj]
    
        plt.ioff()

        if plot_pc1 == True:
            fig, [ax2, ax1] = plt.subplots(2, figsize=(13, 6), sharex=True,
                                           gridspec_kw={'height_ratios': [1, 2]})
        
            ax2.pcolormesh(fft_times, fft_freqs, pc1_power.T, vmin=-5, vmax=0, cmap='nipy_spectral',
                           shading='flat')
        
            ax2.set_xlim(plot_start, plot_end)
            ax2.set_ylim(0, f_max)
            
            if plot_pearls == True:
                for this_time in pearl_times:
                    ax2.axvline(this_time, c='k', ls='--', alpha=0.50)
            
        else:
            fig, ax1 = plt.subplots(figsize=(13, 6))
        
        for ii in range(3):
            ax1.plot(times, 1e3*max_g[:, ii], color=species_colors[ii], label=band_labels[ii], marker='o')
        
        fig.suptitle('EMIC Temporal Growth Rate :: RBSP-A :: {}/{}/{} :: {} Approx.'.format(*ccomp, lbl), fontsize=fontsize+4)
        ax1.set_xlabel('Time (UT)', fontsize=fontsize)
        ax1.set_ylabel(r'$\gamma$ ($\times 10^{-3} s^{-1}$)', fontsize=fontsize)
        ax1.legend(loc='upper left', prop={'size': fontsize}) 
        
        ax1.xaxis.set_major_locator(mdates.MinuteLocator(interval=5))
        ax1.xaxis.set_major_formatter(mdates.DateFormatter('%H:%M'))

        # Set xlim to show either just pearls, or whole event
        ax1.set_xlim(plot_start, plot_end)
        figsave_path = plot_dir + 'LT_' + save_string + '_{:03}_{:03}_{:03}_{}.png'.format(ccomp[0], ccomp[1], ccomp[2], lbl)
    
        # Set ylim to show either just positive growth rate or everything
        if ignore_damping == True:
            ax1.set_ylim(0, None)
    
        if plot_pearls == True:
            for this_time in pearl_times:
                for ax in [ax1, ax2]:
                    ax.axvline(this_time, c='k', ls='--', alpha=0.50)
                
        #for this_time in vlines:
        #    for ax in [ax1, ax2]:
        #        ax.axvline(np.datetime64(this_time), c='k', ls='--', alpha=0.50)
    
        #for this_span in vspan:
        #    for ax in [ax1, ax2]:
        #        ax.axvspan(this_span[0], this_span[1], color='green', alpha=0.2)
    
        if save == True:
            print('Saving {}'.format(figsave_path))
            fig.savefig(figsave_path, bbox_inches='tight')
            plt.close('all')
        else:
            ax1.xaxis.set_major_formatter(mdates.DateFormatter('%H:%M:%S'))
            fig.autofmt_xdate()
            figManager = plt.get_current_fig_manager()
            figManager.window.showMaximized()
            plt.show()
    return max_k, max_g


def plot_max_CGR_with_time(times, k_vals, all_cCGR, all_wCGR, all_hCGR,
                            save=False, norm_w=False, B0=None,
                            ccomp=[70, 20, 10], suff='', ignore_damping=True,
                            plot_pc1=False, plot_pearls=False):
    '''
    UNFINISHED...
    '''
    plot_dir = save_dir + '//MAX_GR_PLOTS{}//'.format(suff)
    if os.path.exists(plot_dir) == False:
        os.makedirs(plot_dir)
        
    save_string = time_start.astype(object).strftime('%Y%m%d_%H%M_') \
                + time_end.astype(object).strftime('%H%M')
    
    tick_label_size = 14
    mpl.rcParams['xtick.labelsize'] = tick_label_size 
    
    species_colors = ['r', 'b', 'g']
    band_labels    = [r'$H^+$', r'$He^+$', r'$O^+$']
    fontsize       = 18
        
    if plot_pc1 == True:
        fft_times, fft_freqs, pc1_power = \
                data.get_pc1_spectra(rbsp_path, time_start, time_end, probe,
                                     pc1_res=25.0, overlap=0.99)
    else:
        fft_times, fft_freqs, pc1_power = None, None, None
        
    if plot_pearls == True:
        print('Plotting pearls...')
        pidx, pearl_times, NULL = rfr.get_pearl_times(time=time_start, crres=False, custom_txtname=None)

    
    for CGR, lbl in zip([all_cCGR, all_wCGR, all_hCGR], ['cold', 'warm', 'hot']):
        Nt    = times.shape[0]
        max_k = np.zeros((Nt, 3))
        max_g = np.zeros((Nt, 3))
            
        # Do CGR bands die? Assume yes
        for ii in range(Nt):
            for jj in range(3):
                    if any(np.isnan(CGR.real[ii, 1:, jj]) == True):
                        max_k[ii, jj] = np.nan
                        max_g[ii, jj] = np.nan
                    else:
                        max_idx       = np.where(CGR.imag[ii, 1:, jj] == CGR.imag[ii, 1:, jj].max())[0][0]
                        max_k[ii, jj] = k_vals[ii, max_idx]
                        max_g[ii, jj] = CGR[ii, max_idx, jj]
    
        plt.ioff()

        if plot_pc1 == True:
            fig, [ax2, ax1] = plt.subplots(2, figsize=(13, 6), sharex=True,
                                           gridspec_kw={'height_ratios': [1, 2]})
        
            ax2.pcolormesh(fft_times, fft_freqs, pc1_power.T, vmin=-7, vmax=1, cmap='jet',
                           shading='flat')
        
            ax2.set_xlim(plot_start, plot_end)
            ax2.set_ylim(0, f_max)
            
            if plot_pearls == True:
                for this_time in pearl_times:
                    ax2.axvline(this_time, c='k', ls='--', alpha=0.50)
            
        else:
            fig, ax1 = plt.subplots(figsize=(13, 6))
        
        for ii in range(3):
            ax1.plot(times, max_g[:, ii], color=species_colors[ii], label=band_labels[ii], marker='o')
        
        fig.suptitle('EMIC Temporal Growth Rate :: RBSP-A :: {}/{}/{} :: {} Approx.'.format(*ccomp, lbl), fontsize=fontsize+4)
        ax1.set_xlabel('Time (UT)', fontsize=fontsize)
        ax1.set_ylabel(r'$\gamma/V_g$ m^{-1}$)', fontsize=fontsize)
        ax1.legend(loc='upper left', prop={'size': fontsize}) 
        
        ax1.xaxis.set_major_locator(mdates.MinuteLocator(interval=5))
        ax1.xaxis.set_major_formatter(mdates.DateFormatter('%H:%M'))

        # Set xlim to show either just pearls, or whole event
        ax1.set_xlim(plot_start, plot_end)
        figsave_path = plot_dir + 'LT_' + save_string + '_{:03}_{:03}_{:03}_{}.png'.format(ccomp[0], ccomp[1], ccomp[2], lbl)
    
        # Set ylim to show either just positive growth rate or everything
        if ignore_damping == True:
            ax1.set_ylim(0, None)
    
        if plot_pearls == True:
            for this_time in pearl_times:
                ax1.axvline(this_time, c='k', ls='--', alpha=0.50)
    
        if save == True:
            print('Saving {}'.format(figsave_path))
            fig.savefig(figsave_path, bbox_inches='tight')
            plt.close('all')
        else:
            ax1.xaxis.set_major_formatter(mdates.DateFormatter('%H:%M:%S.%f'))
            fig.autofmt_xdate()
            figManager = plt.get_current_fig_manager()
            figManager.window.showMaximized()
            plt.show()
    return max_k, max_g


def plot_growth_rates_2D_warm(_rbsp_path, _time_start, _time_end, _probe,
                                 approx='warm', save=True, _cmp=[70, 20, 10],
                                 _nsec=None):
    '''
    Main function that downloads the data, queries the growth rates, and 
    plots a 2D colorplot of the temporal and convective growth rates. Might
    also include plugins to other data such as the Pc1 power to more directly
    look at the change. Also include growth rates?
    
    To Do: Add a normalize flag? At each point, this will normalize f by the 
    proton cylcotron frequency (which is varying in time). This may give slightly
    different results for when there are large perturbations in the background
    field due to the presence of HM?
    
    Also option to normalize based on set B rather than B0 at each time (equivalent
    to absolute measure)
    
    norm: None       (no normalization applied)
          'local'    (normalized based on B0 at each time)
          'absolute' (normalized based on B0 of norm_B0, default 200nT)
    
    How to set growth rate min/max?
    
    Calculate wavenumber range based on average B0 and density for time period, times 1.5
    
    To do:
        -- Plot 2D growth rate and check
        -- Cycle through compositions to find closest one. Can we make a minimization problem?
        -- Plot with Pc1 spectrum in the top, 3 solns with k vs. t vs. GR, then combine for vs. freq
        -- Do one plot each for TGR, CGR. Maybe one for Vg as well?
        -- Start storing values like nsec, comp, filter_freq, probe etc. as a dict
        
    GR output is (time, k, band)
    To accumulate growth over bands, just add? Because linear like Kozyra et al. (1984)
    Interpolation should be taken with a grain of salt, especially with the presence of
    stop bands. Might be better to just code up some solver that converts from w -> k -> GR.
    (e.g. like Kozyra does, using cold plasma approx for k)
    
    How to normalize k axis:
        1) Set k as independent of each time (i.e. calculate k values outside time loop) OR
        2) Use 2D time array so that both k_vals and times are shape(times, k) (Done this one)
    
    To Do: 
        Code up separate function for parameter search, don't need a new plot every time, just
        dump the data.
        
        Zero out any branches for ions that have zero density - these will just be copies
        of another solution
    '''
    k_vals,  WPDR, CGR, Vg, times, B0, name, mass, charge, density,\
        tper, ani, cold_dens, CMP = get_all_DRs_warm_only(_time_start, _time_end, _probe, _cmp, 
        kmin=0.0, kmax=1.5, Nk=1000, knorm=True,
        nsec=_nsec, HM_filter_mhz=50, N_procs=1,
        suff='', data_path=_rbsp_path)

    # Remove nan's at start of arrays (Just copy for now, do smarter later)
    WPDR[:, 0, :] = WPDR[:, 1, :]
    CGR[ :, 0, :] = CGR[ :, 1, :]
    Vg[  :, 0, :] = Vg[  :, 1, :]
    
    # Scale values for plotting/interpolating
    k_vals  *= 1e6
    TGR      = WPDR.imag*1e3
    freqs    = WPDR.real / (2*np.pi)
    max_f    = freqs[np.isnan(freqs) == False].max()
    
    time_2D = np.zeros(k_vals.shape, dtype=times.dtype)
    ti2d    = np.zeros(k_vals.shape, dtype=float)
    for ii in range(k_vals.shape[1]):
        time_2D[:, ii] = times[:]
        ti2d[   :, ii] = np.arange(times.shape[0])
        
    # Interpolate the frequency space values or load from file
    save_string = _time_start.astype(object).strftime('%Y%m%d_%H%M_') \
                + _time_end.astype(object).strftime('%H%M')
    fGR_path    = save_dir + 'fGRw_{}_cc_{:03}_{:03}_{:03}_{}sec.npz'.format(
                     save_string, int(_cmp[0]), int(_cmp[1]), int(_cmp[2]), _nsec)
    
    if os.path.exists(fGR_path) == False:
        time_interp = np.arange(times.shape[0], dtype=float)
        freq_interp = np.linspace(0.0, max_f, 1000)
        xi, yi      = np.meshgrid(time_interp, freq_interp)

        TGRi  = np.zeros(k_vals.shape).flatten()
        for ii in range(TGR.shape[2]):
            x = ti2d.flatten()
            y = freqs[:, :, ii].flatten()
            z = TGR[:, :, ii].flatten()
            print('Interpolating species', ii)
            TGRi[:] += griddata((x, y), z, (xi.flatten(), yi.flatten()),
                                method='cubic', fill_value=0.0)
        TGRi = TGRi.reshape(xi.shape)
        print('Saving growth rate interpolation...')
        np.savez(fGR_path, times=times, freq_interp=freq_interp, TGRi=TGRi)
    else:
        print('Growth rate interpolation already exist, loading from file...')
        fGR_file    = np.load(fGR_path)
        times       = fGR_file['times']
        freq_interp = fGR_file['freq_interp']
        TGRi        = fGR_file['TGRi']
    
    # Load and calculate Pc1 spectra
    ti, fac_mags, dt, gyfreqs = rfl.load_magnetic_field(_rbsp_path,
                        _time_start, _time_end, _probe, pad=3600,
                        LP_B0=1.0, get_gyfreqs=False, return_raw=False,
                        wave_HP=None, wave_LP=None)
    
    print('Calculating autospectra...')
    overlap=0.99; pc1_res=25.0; f_max=0.5
    pc1_xpower, pc1_xtimes, pc1_xfreq = fscr.autopower_spectra(ti, fac_mags[:, 0], _time_start, 
                                                     _time_end, dt, overlap=overlap, df=pc1_res)
    
    pc1_ypower, pc1_ytimes, pc1_yfreq = fscr.autopower_spectra(ti, fac_mags[:, 1], _time_start, 
                                                     _time_end, dt, overlap=overlap, df=pc1_res)
    
    pc1_perp_power = np.log10(pc1_xpower[:, :] + pc1_ypower[:, :]).T
    
    # Set limits        
    kmin     = 0.0
    kmax     = 40.0#k_vals.max()

    TGR_max  = TGR[np.isnan(TGR) == False].max()
    TGR_min  = 0.25*TGR_max
    
    # Plot the things
    plt.ioff()
    fig1, axes1 = plt.subplots(5, figsize=(16, 10))
    
    axes1[0].set_title('Temporal Growth Rate :: RBSP-{} :: {}'.format(_probe.upper(), date_string))
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        im1a = axes1[0].pcolormesh(pc1_xtimes, pc1_xfreq, pc1_perp_power,
                                   vmin=-5, vmax=0, cmap='jet')
        
        im2a = axes1[1].pcolormesh(time_2D, k_vals, TGR[:, :, 0],
                                   vmin=TGR_min, vmax=TGR_max, cmap='viridis')
        im3a = axes1[2].pcolormesh(time_2D, k_vals, TGR[:, :, 1],
                                   vmin=TGR_min, vmax=TGR_max, cmap='viridis')
        im4a = axes1[3].pcolormesh(time_2D, k_vals, TGR[:, :, 2],
                                   vmin=TGR_min, vmax=TGR_max, cmap='viridis')
        
        im5a = axes1[4].pcolormesh(times, freq_interp, TGRi,
                                   vmin=TGR_min, vmax=TGR_max, cmap='viridis')
    
    labels = ['f (Hz)',
              'k ($10^6$/m)\n$H^+$',
              'k ($10^6$/m)\n$He^+$',
              'k ($10^6$/m)\n$O^+$',
              'f (Hz)']
    
    for ax, im, lbl in zip(axes1, [im1a, im2a, im3a, im4a, im5a], labels):
        ax.set_xlim(_time_start, _time_end)
        
        divider = make_axes_locatable(ax)
        cax     = divider.append_axes("right", size="2%", pad=0.5)
        fig1.colorbar(im, cax=cax, label=r'$\gamma \times 10^3 s^{-1}$', orientation='vertical', extend='both')
        ax.set_ylabel(lbl, rotation=0, labelpad=30)
        
        if ax != axes1[-1]:
            ax.set_xticklabels([])
            
        if ax == axes1[0] or ax == axes1[-1]:
            ax.set_ylim(0.0, f_max)
        else:
            ax.set_ylim(kmin, kmax)
            
    axes1[-1].set_xlabel('Time (UT)')
    axes1[-1].xaxis.set_major_locator(mdates.MinuteLocator(interval=10))
    axes1[-1].xaxis.set_major_formatter(mdates.DateFormatter('%H:%M'))
    
    fig1.subplots_adjust(hspace=0.10)

    if save==False:
        plt.show()
    else:
        fig1.savefig(save_dir + 'TGR_{}'.format(save_string))
        print('Plot saved for {}'.format(save_string))
        plt.close('all')
    return


def plot_growth_rates_2D_sweep(_rbsp_path, _time_start, _time_end, _probe,
                                 approx='warm', save=True, nsec=None, log=False):
    '''
    Main function that downloads the data, queries the growth rates, and 
    plots a 2D colorplot of the temporal and convective growth rates. Might
    also include plugins to other data such as the Pc1 power to more directly
    look at the change. Also include growth rates?
    
    To Do: Add a normalize flag? At each point, this will normalize f by the 
    proton cylcotron frequency (which is varying in time). This may give slightly
    different results for when there are large perturbations in the background
    field due to the presence of HM?
    
    Also option to normalize based on set B rather than B0 at each time (equivalent
    to absolute measure)
    
    norm: None       (no normalization applied)
          'local'    (normalized based on B0 at each time)
          'absolute' (normalized based on B0 of norm_B0, default 200nT)
    
    How to set growth rate min/max?
    
    Calculate wavenumber range based on average B0 and density for time period, times 1.5
    
    To do:
        -- Plot 2D growth rate and check
        -- Cycle through compositions to find closest one. Can we make a minimization problem?
        -- Plot with Pc1 spectrum in the top, 3 solns with k vs. t vs. GR, then combine for vs. freq
        -- Do one plot each for TGR, CGR. Maybe one for Vg as well?
        -- Start storing values like nsec, comp, filter_freq, probe etc. as a dict
        
    GR output is (time, k, band)
    To accumulate growth over bands, just add? Because linear like Kozyra et al. (1984)
    Interpolation should be taken with a grain of salt, especially with the presence of
    stop bands. Might be better to just code up some solver that converts from w -> k -> GR.
    (e.g. like Kozyra does, using cold plasma approx for k)
    
    How to normalize k axis:
        1) Set k as independent of each time (i.e. calculate k values outside time loop) OR
        2) Use 2D time array so that both k_vals and times are shape(times, k) (Done this one)
    
    To Do: 
        Code up separate function for parameter search, don't need a new plot every time, just
        dump the data.
        
        Zero out any branches for ions that have zero density - these will just be copies
        of another solution
        
    Also to do:
        - Break this up into load and plot routines, currently very messy.
    '''
    plot_dir = save_dir + 'EVENT_LT_LINEAR_SCALE//'
    if not os.path.exists(plot_dir): os.makedirs(plot_dir)
    
    save_string = _time_start.astype(object).strftime('%Y%m%d_%H%M_') \
                + _time_end.astype(object).strftime('%H%M')
    
    pc1_xtimes, pc1_xfreq, pc1_power = get_mag_data(_time_start, _time_end, _probe,
                 _olap=0.95, _res=25.0,
                 _HM=False, HM_LP=50.0, HM_HP=None, 
                 _split_HM=False, _split_freq = 7.0,
                 transverse_only=False)
    
    
    # Load/Calculate each growth rate in sweep
    for he_comp in np.arange(0.5, 30., 0.5):
        for o_comp in np.arange(0.5, 10., 0.5):
            h_comp = 100. - o_comp - he_comp
            print('Composition: {}/{}/{}'.format(h_comp, he_comp, o_comp))
            comp = np.array([h_comp, he_comp, o_comp], dtype=float)
    
            DR_path = get_DR_filepath(save_dir, _time_start, _time_end, comp, nsec)
            
            if os.path.exists(DR_path):
                k_vals,  WPDR, CGR, Vg, times, B0, name, mass, charge, density,\
                    tper, ani, cold_dens, cmp = get_all_DRs_warm_only(_time_start, _time_end, _probe, comp, 
                    kmin=0.0, kmax=1.5, Nk=1000, knorm=True,
                    nsec=nsec, HM_filter_mhz=50, N_procs=1,
                    suff='', data_path=rbsp_path)
            else:
                print('No growth rate file for composition {}/{}/{}'.format(h_comp, he_comp, o_comp))
                continue

            # Remove nan's at start of arrays (Just copy for now, do smarter later)
            WPDR[:, 0, :] = WPDR[:, 1, :]
            CGR[ :, 0, :] = CGR[ :, 1, :]
            Vg[  :, 0, :] = Vg[  :, 1, :]
            
            # Scale values for plotting/interpolating
            k_vals  *= 1e6
            TGR      = WPDR.imag*1e3
            freqs    = WPDR.real / (2*np.pi)
            max_f    = freqs[np.isnan(freqs) == False].max()
            
            time_2D = np.zeros(k_vals.shape, dtype=times.dtype)
            ti2d    = np.zeros(k_vals.shape, dtype=float)
            for ii in range(k_vals.shape[1]):
                time_2D[:, ii] = times[:]
                ti2d[   :, ii] = np.arange(times.shape[0])
                
            # Interpolate the frequency space values or load from file
            fGR_path = save_dir + 'fGRw_{}_cc_{:03}_{:03}_{:03}_{}sec.npz'.format(
                          save_string, int(10*cmp[0]), int(10*cmp[1]),
                                       int(10*cmp[2]), nsec)
            
            if os.path.exists(fGR_path) == False:
                time_interp = np.arange(times.shape[0], dtype=float)
                freq_interp = np.linspace(0.0, max_f, 1000)
                xi, yi      = np.meshgrid(time_interp, freq_interp)
        
                TGRi  = np.zeros(k_vals.shape).flatten()
                CGRi  = np.zeros(k_vals.shape).flatten()
                
                try:
                    for ii in range(TGR.shape[2]):
                        x = ti2d.flatten()
                        y = freqs[:, :, ii].flatten()
                        z = TGR[:, :, ii].flatten()
                        print('Interpolating species TGR', ii)
                        TGRi[:] += griddata((x, y), z, (xi.flatten(), yi.flatten()),
                                            method='cubic', fill_value=0.0)
                    TGRi = TGRi.reshape(xi.shape)
                    
                    for ii in range(CGR.shape[2]):
                        x = ti2d.flatten()
                        y = freqs[:, :, ii].flatten()
                        z = CGR[:, :, ii].flatten()
                        print('Interpolating species CGR', ii)
                        CGRi[:] += griddata((x, y), z, (xi.flatten(), yi.flatten()),
                                            method='cubic', fill_value=0.0)
                    CGRi = CGRi.reshape(xi.shape)
                    print('Saving growth rate interpolation...')
                    np.savez(fGR_path, times=times, freq_interp=freq_interp, TGRi=TGRi, CGRi=CGRi)
                except:
                    pass
            else:
                print('Growth rate interpolation already exist, loading from file...')
                fGR_file    = np.load(fGR_path)
                times       = fGR_file['times']
                freq_interp = fGR_file['freq_interp']
                TGRi        = fGR_file['TGRi']
                CGRi        = fGR_file['CGRi']
            
            
            # Set limits        
            kmin     = 0.0
            kmax     = 40.0#k_vals.max()
        
            TGR_max  = TGR[np.isnan(TGR) == False].max()
            #TGR_min  = None
            
            # Plot the things
            plt.ioff()
            fig1, axes1 = plt.subplots(5, figsize=(16, 10))
            
            axes1[0].set_title('Temporal Growth Rate :: RBSP-{} :: {} :: Cold Composition {}/{}/{}'.format(
                                        probe.upper(), date_string, (cmp[0]), (cmp[1]),
                                          (cmp[2])))
            with warnings.catch_warnings():
                warnings.simplefilter("ignore")
                im1a = axes1[0].pcolormesh(pc1_xtimes, pc1_xfreq, pc1_power,
                                           vmin=1e-7, vmax=1e1, cmap='jet')
                
                im2a = axes1[1].pcolormesh(time_2D, k_vals, TGR[:, :, 0], cmap='viridis', vmin=0.0)
                                           #vmin=TGR_min, vmax=TGR_max, )
                im3a = axes1[2].pcolormesh(time_2D, k_vals, TGR[:, :, 1], cmap='viridis', vmin=0.0)
                                           #vmin=TGR_min, vmax=TGR_max, )
                im4a = axes1[3].pcolormesh(time_2D, k_vals, TGR[:, :, 2], cmap='viridis', vmin=0.0)
                                           #vmin=TGR_min, vmax=TGR_max, )
                
                try:
                    im5a = axes1[4].pcolormesh(times, freq_interp, TGRi, cmap='viridis', vmin=0.0, vmax=TGR_max)
                                               #vmin=TGR_min, vmax=TGR_max, )
                except:
                    pass
            
            labels = ['f (Hz)',
                      'k ($10^6$/m)\n$H^+$',
                      'k ($10^6$/m)\n$He^+$',
                      'k ($10^6$/m)\n$O^+$',
                      'f (Hz)']
            
            for ax, im, lbl in zip(axes1, [im1a, im2a, im3a, im4a, im5a], labels):
                if ax is axes1[0]:
                    clabel = '$P_\perp$ $(nT^2/Hz)$'
                else:
                    clabel = r'$\gamma \times 10^3 s^{-1}$'
                    
                try:
                    ax.set_xlim(_time_start, _time_end)
                    
                    divider = make_axes_locatable(ax)
                    cax     = divider.append_axes("right", size="2%", pad=0.5)
                    fig1.colorbar(im, cax=cax, label=clabel, orientation='vertical', extend='both')
                    ax.set_ylabel(lbl, rotation=0, labelpad=30)
                    
                    if ax != axes1[-1]:
                        ax.set_xticklabels([])
                    else:
                        ax.set_xlabel('Time (UT)')
                        #ax.xaxis.set_major_locator(mdates.MinuteLocator(interval=10))
                        ax.xaxis.set_major_formatter(mdates.DateFormatter('%H:%M'))
                        
                    if ax == axes1[0] or ax == axes1[-1]:
                        ax.set_ylim(0.0, f_max)
                    else:
                        ax.set_ylim(kmin, kmax)
                except:
                    pass
            
            fig1.subplots_adjust(hspace=0.10)
        
            if save==False:
                plt.show()
            else:
                fig1.savefig(plot_dir + 'TGR_{}_cc_{:03}_{:03}_{:03}_{}sec.png'.format(
                             save_string, int(10*cmp[0]), int(10*cmp[1]),
                                          int(10*cmp[2]), nsec))
                print('Plot saved.')
                plt.close('all')
    return


def plot_max_TGR_CGR_timeseries(rbsp_path, time_start, time_end, probe, pad, norm=None, norm_B0=200.):
    '''
    Note: Because this calculates in frequency space (using cold k) the 'max/min' values
    in the returned arrays are nan's because they are in a stop band.
    '''
    # Frequencies over which to solve (determines frequency cadence)
    Nf    = 1000
    f_max = 1.2
    f_min = 0.0
    freqs = np.linspace(f_max, f_min, Nf)
    w     = 2 * np.pi * freqs
    
    # Create species array for each time (determines time cadence)    
    times, B0, name, mass, charge, density, tper, ani, cold_dens = \
    extract_species_arrays(time_start, time_end, probe,
                           rbsp_path='G://DATA//RBSP//', 
                           cmp=np.array([70, 20, 10]),
                           return_raw_ne=True,
                           HM_filter_mhz=50,
                           nsec=None)
        
    # Initialize empty arrays for GR returns
    Nt      = times.shape[0]
    max_TGR = np.zeros((Nt, 3), dtype=np.float64)
    max_CGR = np.zeros((Nt, 3), dtype=np.float64)
        
    for ii in range(times.shape[0]):
        Species, PP = create_species_array(B0[ii], name, mass, charge, density[:, ii],
                                            tper[:, ii], ani[:, ii])
        
        pcyc   = qp * B0[ii] / (2 * np.pi * mp)
        H_idx  = nearest_index(freqs, pcyc)
        He_idx = nearest_index(freqs, 0.25*pcyc)
        O_idx  = nearest_index(freqs, 0.0625*pcyc)
        
        TGR, CGR, wVg = get_warm_growth_rates(w, Species)
        
        # Mask nan's
        TGR[np.isnan(TGR) == True] = 0.0
        CGR[np.isnan(CGR) == True] = 0.0
        
        try:
            # H-band max growth
            if TGR[H_idx:He_idx].shape[0] == 0:
                max_TGR[ii, 0] = np.nan
                max_CGR[ii, 0] = np.nan
            else:
                max_TGR[ii, 0] = TGR[H_idx:He_idx].max()
                max_CGR[ii, 0] = CGR[H_idx:He_idx].max()
                
            # He-band max growth
            if TGR[He_idx: O_idx].shape[0] == 0:
                max_TGR[ii, 1] = np.nan
                max_CGR[ii, 1] = np.nan
            else:
                max_TGR[ii, 1] = TGR[He_idx: O_idx].max()
                max_CGR[ii, 1] = CGR[He_idx: O_idx].max()
                
            # O-band max growth
            if TGR[O_idx:].shape[0] == 0:
                max_TGR[ii, 2] = np.nan
                max_CGR[ii, 2] = np.nan
            else:
                max_TGR[ii, 2] = TGR[O_idx:].max()
                max_CGR[ii, 2] = CGR[O_idx:].max()
        except:
            pdb.set_trace()

    # Just Temporal Growth Rates in single overlaid plot
    if False:
        fig, ax = plt.subplots(sharex=True)
        ax.plot(times, max_TGR[:, 0], c='r')
        ax.plot(times, max_TGR[:, 1], c='b')
        ax.plot(times, max_TGR[:, 2], c='green')
        
    # Pc1 spectra and TGR/CGR (with log and linear) in two figures
    if False:
        ti, fac_mags, fac_elec, dt, e_flag, gyfreqs = rfl.load_both_fields(rbsp_path, time_start, time_end, probe, pad=3600)
        
        pc1_xpower, pc1_xtimes, pc1_xfreq = fscr.autopower_spectra(ti, fac_mags[:, 0], time_start, 
                                                     time_end, dt, overlap=0.99, df=25.0)
    
        pc1_ypower, pc1_ytimes, pc1_yfreq = fscr.autopower_spectra(ti, fac_mags[:, 1], time_start, 
                                                         time_end, dt, overlap=0.99, df=25.0)
        
        pc1_perp_power = np.log10(pc1_xpower[:, :] + pc1_ypower[:, :])
        
        plt.ioff()
        # Temporal Growth Rate
        fig1, axes1 = plt.subplots(3, figsize=(16,10), sharex=True)
        
        axes1[0].pcolormesh(pc1_xtimes, pc1_xfreq, pc1_perp_power.T, vmin=-7, vmax=1, cmap='jet')
        axes1[0].set_ylabel('Frequency (Hz)')
        axes1[0].set_ylim(f_min, f_max)
        axes1[0].set_title('Temporal Growth Rate (s) vs. Time :: {} :: Cold-k Approximation'.format(date_string))
        
        axes1[1].semilogy(times, max_TGR[:, 0], c='r', label='$H^{+}$')
        axes1[1].semilogy(times, max_TGR[:, 1], c='b', label='$He^{+}$')
        axes1[1].semilogy(times, max_TGR[:, 2], c='green', label='$O^{+}$')
        axes1[1].legend()
        axes1[1].set_ylabel('$\log_{10}(\gamma)$', rotation=0)
        
        axes1[2].plot(times, max_TGR[:, 0], c='r', label='$H^{+}$')
        axes1[2].plot(times, max_TGR[:, 1], c='b', label='$He^{+}$')
        axes1[2].plot(times, max_TGR[:, 2], c='green', label='$O^{+}$')
        axes1[2].legend()
        axes1[2].set_ylabel('$\gamma$', rotation=0)
        
        axes1[2].set_xlim(time_start, time_end)
        
        # Convective Growth Rate
        fig2, axes2 = plt.subplots(3, figsize=(16,10), sharex=True)
        
        axes2[0].pcolormesh(pc1_xtimes, pc1_xfreq, pc1_perp_power.T, vmin=-7, vmax=1, cmap='jet')
        axes2[0].set_ylabel('Frequency (Hz)')
        axes2[0].set_ylim(f_min, f_max)
        axes2[0].set_title('Temporal Growth Rate (s) vs. Time :: {} :: Cold-k Approximation'.format(date_string))
        
        axes2[1].semilogy(times, max_CGR[:, 0], c='r', label='$H^{+}$')
        axes2[1].semilogy(times, max_CGR[:, 1], c='b', label='$He^{+}$')
        axes2[1].semilogy(times, max_CGR[:, 2], c='green', label='$O^{+}$')
        axes2[1].legend()
        axes2[1].set_ylabel('$\log_{10}(S)$', rotation=0)
        
        axes2[2].plot(times, max_CGR[:, 0], c='r', label='$H^{+}$')
        axes2[2].plot(times, max_CGR[:, 1], c='b', label='$He^{+}$')
        axes2[2].plot(times, max_CGR[:, 2], c='green', label='$O^{+}$')
        axes2[2].legend()
        axes2[2].set_ylabel('$S$', rotation=0)
        
        axes2[2].set_xlim(time_start, time_end)
        
        plt.show()
        
    
    # Linear TGR and CGR with input parameters for calculation (looking for spikes!)
    if True:
        plt.ioff()
        for max_arr, title in zip([max_TGR, max_CGR], ['TGR', 'CGR']):
            fig, axes = plt.subplots(9, figsize=(8,20), sharex=True)
            
            axes[0].set_title('Max. {} for {} :: Derived from Satellite Data'.format(title, date_string))
            axes[0].plot(times, max_arr[:, 0], c='r')
            axes[0].plot(times, max_arr[:, 1], c='b')
            axes[0].plot(times, max_arr[:, 2], c='green')
            
            axes[1].plot(times, B0*1e9)
            axes[1].set_ylabel('B0\nnT')
            
            axes[2].plot(times, density[0]*1e-6, c='r')
            axes[2].plot(times, density[1]*1e-6, c='b')
            axes[2].plot(times, density[2]*1e-6, c='green')
            axes[2].set_ylabel('Cold $n_i$\n/cc')
            
            axes[3].plot(times, density[3]*1e-6, c='r')
            axes[3].plot(times, density[4]*1e-6, c='b')
            axes[3].plot(times, density[5]*1e-6, c='green')
            axes[3].set_ylabel('HOPE $n_i$\n/cc')
            
            axes[4].plot(times, density[6]*1e-6, c='r')
            axes[4].plot(times, density[7]*1e-6, c='b')
            axes[4].plot(times, density[8]*1e-6, c='green')
            axes[4].set_ylabel('RBSPICE $n_i$\n/cc')
            
            axes[5].plot(times, tper[3]*1e-3, c='r')
            axes[5].plot(times, tper[4]*1e-3, c='b')
            axes[5].plot(times, tper[5]*1e-3, c='green')
            axes[5].set_ylabel('HOPE $T_\perp$\nkeV')
            
            axes[6].plot(times, tper[6]*1e-3, c='r')
            axes[6].plot(times, tper[7]*1e-3, c='b')
            axes[6].plot(times, tper[8]*1e-3, c='green')
            axes[6].set_ylabel('RBSPICE $T_\perp$\nkeV')
            
            axes[7].plot(times, ani[3], c='r')
            axes[7].plot(times, ani[4], c='b')
            axes[7].plot(times, ani[5], c='green')
            axes[7].set_ylabel('HOPE $A_i$')
            
            axes[8].plot(times, ani[6], c='r')
            axes[8].plot(times, ani[7], c='b')
            axes[8].plot(times, ani[8], c='green')
            axes[8].set_ylabel('RBSPICE $A_i$')
            
            fig.subplots_adjust(hspace=0)
            fig.align_ylabels()
            
            for ax in axes:
                ax.set_xlim(time_start, time_end)
                
            if False:
                plt.show()
            else:
                svpath = save_dir + '{}_stackplot.png'.format(title)
                print('Plot saved as {}'.format(svpath))
                fig.savefig(svpath)
                plt.close('all')
    return


def plot_dispersion_and_growth(ax_disp, ax_growth, k_vals, CPDR, WPDR, HPDR, w_cyc,
                               norm_w=False, norm_k=False, save=False, savepath=None, alpha=1.0):
    '''
    Plots the CPDR and WPDR nicely as per Wang et al 2016. Can plot multiple
    dispersion/growth curves for varying parameters.
    This is mainly just a function to be called by plot_all_DRs.
    
    INPUT:
        k_vals     -- Wavenumber values in /m or normalized to p_cyc/v_A
        CPDR_solns -- Cold-plasma frequencies in Hz or normalized to p_cyc
        WPDR_solns -- Warm-plasma frequencies in Hz or normalized to p_cyc. 
                   -- .real is dispersion relation, .imag is growth rate vs. k
                   
    To do: Add normalisations for w, k
    '''
    # Identifiers
    species_colors = ['r', 'b', 'g']
    band_labels    = [r'$H^+$', r'$He^+$', r'$O^+$']
    
    type_label = ['Cold Plasma Approx.', 'Warm Plasma Approx.', 'Hot Plasma Approx.', 'Cyclotron Frequencies']
    type_style = [':', '--', '-', '-']
    type_alpha = [1.0, 1.0, 1.0, 0.5]
    
    #######################
    ### DISPERSION PLOT ###
    #######################
    for ii in range(3):
        ax_disp.plot(k_vals[1:]*1e6, CPDR[1:, ii].real, c=species_colors[ii], linestyle='--', label='Cold')
        ax_disp.plot(k_vals[1:]*1e6, HPDR[1:, ii].real, c=species_colors[ii], linestyle='-',  label='Warm')
        ax_disp.axhline(w_cyc[ii] / (2 * np.pi), c='k', linestyle='-', alpha=type_alpha[-1])
    
    type_legend = create_type_legend(ax_disp, type_label, type_style, type_alpha)
    ax_disp.add_artist(type_legend)
    
    ax_disp.set_title('Dispersion Relation')
    ax_disp.set_xlabel(r'$k (\times 10^{-6} m^{-1})$')
    ax_disp.set_ylabel(r'$\omega${}'.format(' (Hz)'))
    
    ax_disp.set_xlim(0, k_vals[-1]*1e6)
    ax_disp.set_ylim(0, w_cyc[0] * 1.1 / (2 * np.pi))
    
    ########################
    ### GROWTH RATE PLOT ###
    ########################
    band_legend = create_band_legend(ax_growth, band_labels, species_colors)
    ax_growth.add_artist(band_legend)
    
    for ii in range(3):
        ax_growth.plot(k_vals[1:]*1e6, HPDR[1:, ii].imag, c=species_colors[ii], linestyle='-',  label='Growth')
    ax_growth.axhline(0, c='k', linestyle=':')
    
    ax_growth.set_title('Temporal Growth Rate')
    ax_growth.set_xlabel(r'$k (\times 10^{-6}m^{-1})$')
    ax_growth.set_ylabel(r'$\gamma (s^{-1})$')
    ax_growth.set_xlim(0, k_vals[-1]*1e6)
    
    ax_growth.set_ylim(None, None)
    
    y_thres_min = -0.05;  y_thres_max = 0.05
    if ax_growth.get_ylim()[0] < y_thres_min:
        y_min = y_thres_min
    else:
        y_min = y_thres_min
        
    if ax_growth.get_ylim()[0] > y_thres_max:
        y_max = None
    else:
        y_max = y_thres_max
    
    ax_growth.set_ylim(y_min, y_max)
    
    ax_disp.minorticks_on()
    ax_growth.minorticks_on() 
    
    ax_growth.yaxis.set_label_position("right")
    ax_growth.yaxis.tick_right()
    return





def plot_all_DRs(all_k, all_CPDR, all_WPDR, all_HPDR, times, B, name, mi, qi, ni, t_perp, A, ne,
                 suff='', HM_filter_mhz=50, overwrite=False, save=True, figtext=True):
    '''
    CPDR is cold approx
    WPDR is Chen's warm approx of 2013
    HPDR is the fully kinetic hot approx used in both Chen 2013 and Wang 2016.
    '''
    Nt = times.shape[0]
    
    figsave_dir = os.path.join(save_dir, 'all_DRs' + suff)
    if os.path.exists(figsave_dir) == False:
        os.makedirs(figsave_dir)

    for ii in range(Nt):
        save_string = times[ii].astype(object).strftime('%Y%m%d_%H%M%S')
                
        figsave_path = figsave_dir + 'linear_{}_{}.png'.format(save_string, ii)
        
        if os.path.exists(figsave_path) == True and overwrite == False:
            print('Plot already done, skipping...')
            continue
        
        # Convert frequencies to Hz (from rad/s for calculations)
        time   = times[ii]
        k_vals = all_k[ii]
        CPDR   = all_CPDR[ii] / (2*np.pi)
        WPDR   = all_WPDR[ii] / (2*np.pi)
        HPDR   = all_HPDR[ii] / (2*np.pi)

        w_cyc  = qi * B[ii] / mi 
        
        plt.ioff()
        fig, [ax1, ax2] = plt.subplots(ncols=2, figsize=(16, 10))

        fig.text(0.34, 0.974, '{}'.format(time))

        plot_dispersion_and_growth(ax1, ax2, k_vals, CPDR, WPDR, HPDR, w_cyc,
                                   norm_w=False, norm_k=False, save=False, 
                                   savepath=figsave_path)    

        plt.setp(ax2.get_xticklabels()[0], visible=False)

        # Change this to work with Species array
        if figtext == True:
            set_figure_text(ax2, ii, B[ii], name, mi, qi, ni[:, ii], t_perp[:, ii], A[:, ii], ne[ii])
        
        fig.tight_layout()
        fig.subplots_adjust(wspace=0, hspace=0, right=0.75)
        
        if save == True:
            print('Saving {}'.format(figsave_path))
            fig.savefig(figsave_path)
            plt.close('all')
        else:
            # Only shows the first one
            figManager = plt.get_current_fig_manager()
            figManager.window.showMaximized()
            break
    return


def plot_all_CGRs(all_k, all_cCGR, all_wCGR, all_hCGR, times, B, name, mi, qi, ni, t_perp, A, ne,
                 suff='', HM_filter_mhz=50, overwrite=False, save=True, figtext=True,
                 ylim=None):
    '''
    Cold, warm, hot approximations to the convective growth rate. None for hCGR because I
    haven't worked out how to calculate that yet.
    
    Just a simple timeseries for each time, since there's no dispersion function to go
    along with it.
    
    But perhaps could include Vg later on? And maybe even do a triple plot of group velocity, 
    temporal growth rate, and convective growth rate, for comparison. Would require outputting
    and saving the group velocity, and re-calculating for each event.
    
    all_CGR :: (times, k, N_solns)
    '''
    Nt = times.shape[0]
    
    figsave_dir = os.path.join(save_dir, 'all_CGRs' + suff)
    if os.path.exists(figsave_dir) == False:
        os.makedirs(figsave_dir)
        
    # Legend properties
    species_colors = ['r', 'b', 'g']
    band_labels    = [r'$H^+$', r'$He^+$', r'$O^+$']
    
    type_label = ['Cold Plasma Approx.', 'Warm Plasma Approx.']
    type_style = ['--', '-']
    type_alpha = [1.0, 1.0]


    for tt in range(Nt):
        save_string = times[tt].astype(object).strftime('%Y%m%d_%H%M%S')
        figsave_path = figsave_dir + 'linear_{}_{}.png'.format(save_string, tt)
        
        if os.path.exists(figsave_path) == True and overwrite == False:
            print('Plot already done, skipping...')
            continue
        
        # Pick specific time just to be easier
        time   = times[tt]
        k_vals = all_k[tt]
        cCGR   = all_cCGR[tt] * 1e9
        wCGR   = all_wCGR[tt] * 1e9

        plt.ioff()
        fig, ax = plt.subplots(figsize=(16, 10))

        # TODO : Check if this works fine with CGR
        if figtext == True:
            set_figure_text(ax, tt, B[tt], name, mi, qi, ni[:, tt],
                            t_perp[:, tt], A[:, tt], ne[tt])
        
        # Plot the actual CGRs
        for ii in range(3):
            ax.plot(k_vals[1:]*1e6, cCGR[1:, ii], c=species_colors[ii], linestyle='--', label='Cold')
            ax.plot(k_vals[1:]*1e6, wCGR[1:, ii], c=species_colors[ii], linestyle='-',  label='Warm')
    
        # Add legends to show species type and approximations
        type_legend = create_type_legend(ax, type_label, type_style, type_alpha)
        ax.add_artist(type_legend)
        
        band_legend = create_band_legend(ax, band_labels, species_colors)
        ax.add_artist(band_legend)
        
        ax.set_title('Convective Growth Rate :: {}'.format(time))
        ax.set_xlabel(r'$k (\times 10^{-6} m^{-1})$')
        ax.set_ylabel(r'$S (\times 10^{9} m^{-1})$')
        
        ax.set_xlim(0, k_vals[-1]*1e6)
        ax.set_ylim(0, ylim)
        
        fig.tight_layout()
        fig.subplots_adjust(wspace=0, hspace=0, right=0.75)
        
        if save == True:
            figsave_path = figsave_dir + '//CGR_{}_{}.png'.format(save_string, tt)
            print('Saving {}'.format(figsave_path))
            fig.savefig(figsave_path)
            plt.close('all')
        else:
            # Only shows the first one
            figManager = plt.get_current_fig_manager()
            figManager.window.showMaximized()
            break
    return


def plot_residuals(Species, PP, k, w_vals, lbl='', approx='hot'):
    '''
    k_vals :: Wavenumber in 1D array, indpt variable
    w_vals :: Solutions for each k in several bands. 2D array of shape (Nk, N_solns)
    lbl    :: Plot label
    approx :: Hot, warm, cold
    
    Function to back-substitude calculated solutions into respective equations in order
    to test their accuracy.
    '''
    print('Plotting residuals...')
    residuals = np.zeros((w_vals.shape[0], w_vals.shape[1], 2), dtype=float)
    
    for jj in range(w_vals.shape[1]):
        for ii in range(w_vals.shape[0]):
            if approx == 'cold':
                w_arg = w_vals[ii, jj].real
                residuals[ii, jj, 0] = cold_dispersion_eqn(w_arg, k[ii], Species)
            elif approx == 'warm':
                w_arg = w_vals[ii, jj].real
                residuals[ii, jj, 0] = warm_dispersion_eqn(w_arg, k[ii], Species)
            elif approx == 'hot':
                w_arg = np.array([w_vals[ii, jj].real, w_vals[ii, jj].imag])
                residuals[ii, jj]    =  hot_dispersion_eqn(w_arg, k[ii], Species)
        
    species_clrs = ['r', 'b', 'g']
        
    residuals  /= PP['pcyc_rad']
    k_vals      = k*PP['va'] / PP['pcyc_rad']
    
    # Plot here
    plt.ioff()
    fig, ax = plt.subplots(nrows=1, ncols=2)
  
    for ii in range(w_vals.shape[1]):
        ax[0].plot(k_vals, residuals[:, ii, 0], c=species_clrs[ii])
        ax[0].set_title('Dispersion Relation Residuals')
        ax[0].set_ylabel(r'$\omega_r/\Omega_p$')
        
        ax[1].plot(k_vals, residuals[:, ii, 1], c=species_clrs[ii])
        ax[1].set_title('Growth Rate Residuals')
        ax[1].set_ylabel(r'$\gamma/\Omega_p$')
        
        for axes in ax:
            axes.set_xlim(k_vals[0], k_vals[-1])
            axes.minorticks_on()
            axes.set_xlabel(r'$kv_A / \Omega_p$')
            

    figManager = plt.get_current_fig_manager()
    figManager.window.showMaximized() 
    return


def plot_normalized_pc1(_time_start, _time_end, _probe, save=True, save_interpolation=False,
                        zero_pad=False, zpad_mult=1.0):
    '''
    Generate standard spectra of Pc1, but then normalize each time's frequency
    bin to the cyclotron frequency
    
    Ok, so it didn't work for the 25/07/2013 event, but still a good idea
    Does that make it an IPDP? Rather than a real pearl? Am I barking up the
    wrong tree?
    '''  
    pc1_save_dir= '{}//NORMALIZED_PC1//EVENT_{}//'.format(ext_drive, date_string)
    if not os.path.exists(pc1_save_dir): os.makedirs(pc1_save_dir)
    
    plot_factor   = 0.3
    pc1_res       = 35.0                         # FFT resolution
    norm_freq     = np.arange(0.0, 1.2, 0.005)   # Normalized frequency grid
    
    ## LOAD DATA##
    pc1_times, pc1_freq, pc1_power, times, gyfreqs = get_mag_data(_time_start, _time_end, _probe,
                     _olap=0.95, _res=pc1_res,
                     _HM=False, HM_LP=50.0, HM_HP=None, 
                     _split_HM=False, _split_freq = 7.0,
                     transverse_only=False)
    
    ## PROCESS
    print('Normalizing frequency spectrum...')
    norm_power = np.zeros((norm_freq.shape[0], pc1_times.shape[0]), dtype=pc1_power.dtype)
    pcyc_freqs = np.zeros((pc1_times.shape[0]), dtype=pc1_power.dtype)
    
    for ii in range(pc1_times.shape[0]):
        
        # Get cyclotron frequency for this time
        idx           = np.where(abs(times - pc1_times[ii]) == np.min(abs(times - pc1_times[ii])))[0][0] 
        pcyc_freqs[ii] = gyfreqs[0, idx]
        
        # Get normalized frequency
        this_norm_freq = pc1_freq.copy() / pcyc_freqs[ii]
        
        # Interpolate this power spectrum back onto uniform normalized grid
        coeffs_power      = splrep(this_norm_freq, pc1_power[:, ii])
        power_out         = splev(norm_freq, coeffs_power)
        norm_power[:, ii] = power_out
        
        if save_interpolation:
            plt.ioff()
            fig0, ax = plt.subplots(2, figsize=(16, 9))
            
            ax[0].set_title('Raw FFT :: {}'.format(pc1_times[ii]))
            ax[0].semilogy(pc1_freq, pc1_power[:, ii], c='k', lw=1.5)
            ax[0].axvline(gyfreqs[0, idx], color='k', ls='--')
            ax[0].axvline(gyfreqs[1, idx], color='yellow', ls='--')
            ax[0].axvline(gyfreqs[2, idx], color='red', ls='--')
            ax[0].set_xlim(0.0, 1.2*gyfreqs[0, idx])
            
            ax[1].set_title('Normalized FFT :: $f_H = %.2f Hz$' % pcyc_freqs[ii])
            ax[1].semilogy(this_norm_freq, pc1_power[:, ii], marker='o', c='k', label='Raw Normalized')
            ax[1].scatter(norm_freq, power_out, marker='x', color='r', label='Interpolation')
            ax[1].axvline(1.0000, color='k', ls='--')
            ax[1].axvline(0.2500, color='yellow', ls='--')
            ax[1].axvline(0.0625, color='red', ls='--')
            ax[1].set_xlim(0.0, 1.2)
            ax[1].legend(loc='upper right')
        
            print('Saving time {:04}'.format(ii))
            fig0.savefig(pc1_save_dir + 'normalized_fft_time{:04}.png'.format(ii))
            plt.close('all')
        
    
    ## PLOT RESULTS ##
    plt.ioff()
    fig, axes = plt.subplots(2, sharex=True, figsize=(16, 9))
    
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        im1 = axes[0].pcolormesh(pc1_times, pc1_freq, pc1_power, cmap='jet',
                        norm=colors.LogNorm(vmin=1e-7, vmax=1e1))
        
        im2 = axes[1].pcolormesh(pc1_times, norm_freq, norm_power, cmap='jet',
                        norm=colors.LogNorm(vmin=1e-7, vmax=1e1))
        
        fig.colorbar(im1, ax=axes[0], extend='both').set_label(
                    '$|P|$\n$nT^2/Hz$', fontsize=12, rotation=0, labelpad=30)
        
        fig.colorbar(im2, ax=axes[1], extend='both').set_label(
                    '$|P|$\n$nT^2/Hz$', fontsize=12, rotation=0, labelpad=30)
    
    axes[0].set_title('Raw FFT :: {}'.format(pc1_times[ii]))
    axes[0].plot(times, gyfreqs[0], c='white')
    axes[0].plot(times, gyfreqs[1], c='yellow')
    axes[0].plot(times, gyfreqs[2], c='red')
    axes[0].scatter(pc1_times, pcyc_freqs, c='k', marker='x', label='Norm. constant')
    
    axes[0].set_ylabel('$f$\n(Hz)', rotation=0, fontsize=14, labelpad=20)
    axes[0].set_ylim(0.0, plot_factor*gyfreqs[0, idx])
    axes[0].set_xlim(_time_start, _time_end)
    
    axes[1].set_title('Normalized FFT')
    axes[1].axhline(1.0000, color='white')
    axes[1].axhline(0.2500, color='yellow')
    axes[1].axhline(0.0625, color='red')
    
    axes[1].set_ylabel('$f/f_{H^+}$\n(Hz)', rotation=0, fontsize=14, labelpad=20)
    axes[1].set_ylim(0.0, plot_factor)
    axes[1].set_xlim(_time_start, _time_end)
    axes[1].set_xlabel('Time (UT)')
    
    # Show image
    if not save:
        axes[-1].xaxis.set_major_formatter(mdates.DateFormatter('%H:%M'))
        fig.autofmt_xdate()
        figManager = plt.get_current_fig_manager()
        figManager.window.showMaximized()
        plt.show()
    else:
        print('Saving dynamic spectra comparison')
        fig.savefig(pc1_save_dir + 'normalized_fft_comparison.png')
        plt.close('all')
    return


def plot_cold_ion_via_cutoff(_cutoffs, title_suff=None):
    '''
    Calculates a range of possible n_He and n_O cold ion concentrations
    based on the cutoff values given in the list _cutoffs.
    
    Input:
        _cutoffs  :: List of (normalized) cutoff frequencies
                     Will be plot on n_He vs. n_O graph
        
    Note: Do single value first then expand to list. Or do type check?
    
    TODO:CHECK THIS
    '''    
    He_min = 0.0;   He_max = 0.5
    O_min  = 0.0;   O_max  = 0.5
    
    He_array = np.linspace(He_min, He_max, 101, endpoint=True)
    
    plt.ioff()
    fig, ax = plt.subplots(figsize=(16, 10))
    for co in _cutoffs:
        m_top   = He_array - He_array*(1. - co)/(1. - 4*co) - co
        m_bot   = (1. - co)/(1. - 16.*co) - 1.
        O_array = m_top/m_bot
        
        ax.plot(He_array, O_array, label='$\hat{\omega}_{co} = %.3f$' % co)
        #ax.scatter(0.095, 0.095, marker='x', c='red', label='Omura CLUSTER')

    ax.plot(He_array, He_array, c='k', ls='--', label='$n_{He^+}=n_{O^+}$', alpha=0.3)
    ax.legend()
    
    ax.set_xlim(He_min, He_max)
    ax.set_xlabel('$\\frac{n_{He^+}}{n_e}$', fontsize=18)
    ax.set_ylabel('$\\frac{n_{O^+}}{n_e}$', rotation=0, fontsize=18)
    ax.set_ylim(O_min, O_max)
    ax.set_title('Heavy Ion Concentration via EMIC Cutoff Observations', fontsize=18)
    if title_suff is not None:
        ax.set_title(ax.get_title() + title_suff)
    plt.show()
    return


def power_spectrum_CGR_comparison():
    '''
    Alexa's version plotting them over each other as a function of L shell
    I'm assuming she only changed the magnetic field? But we can do one better
    And use the density/field information to calculate parameters at each time
    Assuming some constant cold composition and some constant ring current
    
    TODO: Compare against the Chen et al. (2013) version of this, should be 
    identical (just like the Fraser, 1996 ones)
    
    Sweep through ion compositions (0.5 ):
        Cold helium: 0-30%
        Cold oxygen: 0-10%  # Use cutoff to work out what these values should be
    '''
    
    
    if True:
        _crres_path = '%s//DATA//CRRES//' % ext_drive
        _time_start = np.datetime64('1991-07-17T20:15:00.000000')
        _time_end   = np.datetime64('1991-07-17T21:00:00.000000')
    else:
        _crres_path = '%s//DATA//CRRES//' % ext_drive
        _time_start = np.datetime64('1991-08-12T22:10:00.000000')
        _time_end   = np.datetime64('1991-08-12T23:15:00.000000')
        
    # Load CRRES data
    den_times, den_dict = cfr.get_crres_density(_crres_path, _time_start, _time_end, pad=0)
    
    times, B0, HM, dB, E0, HMe, dE, S, B, E = cfr.get_crres_fields(_crres_path,
                  _time_start, _time_end, pad=0, E_ratio=5.0, rotation_method='vector', 
                  output_raw_B=True, interpolate_nan=None, B0_LP=1.0,
                  Pc1_LP=5000, Pc1_HP=100, Pc5_LP=30, Pc5_HP=None, dEx_LP=None)
    
    # Calculate power spectrum
    mag_freq, mag_power_x = get_power_spectrum(times, dB[:, 0])
    mag_freq, mag_power_y = get_power_spectrum(times, dB[:, 1])
    mag_freq, mag_power_z = get_power_spectrum(times, dB[:, 2])
    mag_power = (mag_power_x + mag_power_y + mag_power_z)
    
    # Smooth power spectrum (Maybe do rolling mean instead)
    #smoothed_power = gaussian_smooth(mag_power)
    smoothed_power = pd.DataFrame(mag_power).rolling(11).mean()
    
    # Low-pass total field to avoid aliasing (Assume 8.5 second cadence)
    B_dt  = 1.0 / 32.0
    nyq   = 1.0 / (2.0 * 8.5) 
    for ii in range(3):
        B[:, ii] = ascr.clw_low_pass(B[:, ii].copy(), nyq, B_dt, filt_order=4)
        
    # Take magnitude and interpolate            
    B_mag    = np.sqrt(B[:, 0] ** 2 + B[:, 1] ** 2 + B[:, 2] ** 2)
    B_interp = np.interp(den_times.astype(np.int64), times.astype(np.int64), B_mag)  
    edens    = den_dict['NE_CM3']
    max_gyfreq = qp * B_interp.max() * 1e-9 / (2 * np.pi * mp)
    
    ring_currents = np.arange(0.05, 5.0, 0.05)
    heliums = np.arange(0, 30., 0.25)
    oxygens = np.arange(0, 30., 0.25)
    
    count = 0
    save_path = 'D://Google Drive//Uni//PhD 2017//Josh PhD Share Folder//Thesis//Data_Plots//19910717_CRRES//CGRS//'
    for mm in range(ring_currents.shape[0]):
        for nn in range(heliums.shape[0]):
            for oo in range(oxygens.shape[0]):
                
                fname = save_path+f'EMIC_CGR_JUL171991_{count:05}.png'
                
                if os.path.exists(fname):
                    print(f'Skipping plot {count:05}...')
                    count += 1
                    continue
                else:
                    nc_o  = oxygens[oo]
                    nc_he = heliums[nn]
                    nc_h  = 100. - oxygens[oo] - heliums[nn]
                    
                    nh_h = ring_currents[mm]
                    
                    print(f'Calculating growth rate plot for {nc_h}/{nc_he}/{nc_o}, {nh_h}')
                        
                    # Set compositions used for loop
                    rc_percent = ring_currents[mm]               # Percentage of density that is ring current
                    ncold = np.array([nc_h, nc_he, nc_o], dtype=float)  # Percentage Split cold density into components
                    nwarm = np.array([nh_h, 0, 0], dtype=float)   # Percentage Split warm density into components
                    
                    ANI   = np.array([1.0, 0.0, 0.0], dtype=float) # Anisotropy of RC ions
                    tperp = np.array([40e3, 0, 0], dtype=float)    # Perpendicular temperature of RC ions (eV)
                    
                    plt.ioff()
                    fig, [ax, cax] = plt.subplots(nrows=1, ncols=2, figsize=(16, 10),
                                     gridspec_kw={'width_ratios':[2, 0.05]})
                    
                    ax.set_title('EMIC CGR :: {} :: {:.2f}/{:.2f}/{:.2f} cold, {:.2f} hot protons'.format(
                        _time_start.astype(object).strftime('%Y-%m-%d'),
                        nc_h, nc_he, nc_o, nh_h))
                    ax.set_xlabel('Frequency (Hz)')
                    ax.set_ylabel('Growth Rate\n$(/cm^{-1})$', rotation=0, labelpad=20)
                    ax.set_xlim(0.0, max_gyfreq)
                    
                        
                    # Specify color values for time
                    time0  = _time_start.astype(np.int64)
                    time1  = _time_end.astype(np.int64)
                    norm   = mpl.colors.Normalize(vmin=time0, vmax=time1, clip=False)
                    mapper = cm.ScalarMappable(norm=norm, cmap=cm.viridis)
                    
                    for ii in range(den_times.shape[0]):
                        if ii%2 == 0:
                            #print('Calculating growth rates for {}'.format(den_times[ii]))
                            # Set parameters for function input
                            field  = B_interp[ii]                                       # nT
                            ndensc = (1. - rc_percent*1e-2) * edens[ii] * ncold*1e-2   # /cm3
                            ndensw =       rc_percent*1e-2  * edens[ii] * nwarm*1e-2   # /cm3
                    
                            # Calculate growth rate
                            freq, growth, stop = convective_growth_rate_kozyra(field, ndensc, ndensw, ANI, tperp,
                                                          norm_ampl=0, norm_freq=0, NPTS=1000, maxfreq=1.0)
                            
                            # Add line to plot
                            clr = mapper.to_rgba(den_times[ii].astype(np.int64))
                            
                            ax.plot(freq, growth, c=clr)
                            
                    # Add colorbar
                    label_every = 20
                    cbar    = fig.colorbar(mapper, cax=cax, orientation='vertical',
                                           ticks=den_times[::label_every].astype(np.int64))
                    
                    cbar.set_label('Time (UT)', labelpad=20)
                    
                    for label in cbar.ax.get_yminorticklabels():
                        label.set_visible(False)
                        
                    tlabels = [this_time.astype(object).strftime('%H:%M:%S') for this_time in den_times[::label_every]]
                    cbar.ax.set_yticklabels(tlabels)
                    
                    #fig.subplots_adjust(wspace=0.25)
                    
                    # Add FFT overlay
                    ax2 = ax.twinx()
                    ax2.plot(mag_freq, mag_power, c='k', lw=0.5)
                    ax2.plot(mag_freq, smoothed_power, c='k', lw=1.5)
                    ax2.set_ylabel('Power\n$(nT^2/Hz)$', rotation=0, labelpad=20)
                
                    fig.savefig(fname)
                    plt.close('all')
                    #plt.show()
                    count += 1
    return


def parameter_search_2D():
    '''
    Create two plots:
        -- Find major frequency from data (spectrogram plus FFT over whole event time)
        -- 2D plot over (T, A) space, since hot ion density makes little difference except growth rate
    Assume some small fraction of RC - i.e. 1-5%
    '''
    if True:
        _crres_path = '%s//DATA//CRRES//' % ext_drive
        _time_start = np.datetime64('1991-07-17T20:15:00.000000')
        _time_end   = np.datetime64('1991-07-17T21:00:00.000000')
        f_event     = 0.165     # Event mean/strongest frequency
        f_max       = 0.4
        
        # Note: These values vary wildly
        const_ne = 40.0 # Cold electron density (/cc)
        const_B0 = 65.0 # nT
        
        ER_lims    = [2.5, 25]    # Resonant energy limits (keV), equivalent to E_para (keV)
    else:
        _crres_path = '%s//DATA//CRRES//' % ext_drive
        _time_start = np.datetime64('1991-08-12T22:10:00.000000')
        _time_end   = np.datetime64('1991-08-12T23:15:00.000000')
        
    # Load CRRES data
    den_times, den_dict = cfr.get_crres_density(_crres_path, _time_start, _time_end, pad=0)
    times, B0, HM, dB, E0, HMe, dE, S, B, E = cfr.get_crres_fields(_crres_path,
                  _time_start, _time_end, pad=0, E_ratio=5.0, rotation_method='vector', 
                  output_raw_B=True, interpolate_nan=None, B0_LP=1.0,
                  Pc1_LP=5000, Pc1_HP=100, Pc5_LP=30, Pc5_HP=None, dEx_LP=None)
    
# =============================================================================
#     if False:
#         B_total = np.sqrt(B[:, 0]**2 + B[:, 1]**2 + B[:, 2]**2)
#         plt.plot(times, B_total)
#         plt.xlim(_time_start, _time_end)
#         plt.show()
#         sys.exit()
# =============================================================================
        
    # Calculate power spectrum
    mag_freq, mag_power_x = get_power_spectrum(times, dB[:, 0])
    mag_freq, mag_power_y = get_power_spectrum(times, dB[:, 1])
    mag_freq, mag_power_z = get_power_spectrum(times, dB[:, 2])
    mag_power = (mag_power_x + mag_power_y + mag_power_z)
    smoothed_power = pd.DataFrame(mag_power).rolling(11).mean()
    
    # Heavy ion fractions
    nc_He = 0.1
    nc_O  = 0.05
    nc_H  = 1. - nc_He - nc_O
    nc_RC = 0.01
    
    # Gyrofrequencies
    gyfreq = qp * B0*1e-9 / (2*np.pi*mp * np.array([1.0, 4.0, 16.0]))
    
    # Set arrays
    ndensc = np.array([nc_H*(1-nc_RC) , nc_He , nc_O])*const_ne
    ndensw = np.array([nc_H*   nc_RC  , 0.0   , 0.0 ])*const_ne
    ANI    = np.array([0.0            , 0.0   , 0.0 ])
    tperp  = np.array([0.0            , 0.0   , 0.0 ])
    
    # Do parameter search in (T_para, A) space
    nE = 500; Epara_axis = np.linspace(ER_lims[0], ER_lims[1], nE)*1e3
    nA = 500; anis_axis  = np.linspace(0.0, 2.0, nE)
    max_freq   = np.zeros((nE, nA), dtype=np.float64)
    max_growth = np.zeros((nE, nA), dtype=np.float64)
    
    for ii in range(nE):
        print('Energy:', Epara_axis[ii], 'keV')
        for jj in range(nA):
            
            # Set parameters for this loop
            ANI[0]   = anis_axis[jj]
            tpara    = Epara_axis[ii]
            tperp[0] = tpara * (anis_axis[jj]+1)
            
            # Calculate growth rate
            freq, growth, stop = convective_growth_rate_kozyra(
                const_B0, ndensc, ndensw, ANI, tperp,
                norm_ampl=0, norm_freq=0, NPTS=1000, maxfreq=1.0)
            
            he_low_idx, he_hi_idx = ascr.boundary_idx64(freq, gyfreq[2], gyfreq[1]) 
            peak_idx  = growth[he_low_idx: he_hi_idx].argmax() + he_low_idx

            max_freq[ii, jj] = freq[peak_idx]
            max_growth[ii, jj] = growth[peak_idx]
            
            pdb.set_trace()
    
    # Plot results:
    fig, axes = plt.subplots(nrows=2, ncols=3, gridspec_kw={
                 'width_ratios':[1, 0.01],
                 'height_ratios':[1, 3]})
    
    axes[0, 0].plot(mag_freq, mag_power,      c='k', lw=0.75, alpha=0.75)
    axes[0, 0].plot(mag_freq, smoothed_power, c='k', lw=1.5)
    axes[0, 0].set_ylabel('Power\n$(nT^2/Hz)$', rotation=0, labelpad=20)
    axes[0, 0].set_xlim(0.0, f_max)
    axes[0, 0].axvline(f_event, c='k', ls='--', alpha=0.5)
    axes[0, 1].set_visible(False)
    
    im = axes[1, 0].pcolormesh(Epara_axis, anis_axis, max_freq.T, shading='auto', 
                          norm=colors.Normalize(vmin=0.0, vmax=0.7), cmap='jet')
    fig.colorbar(im, cax=axes[1, 1], extend='both').set_label(
            'Frequency of max S', fontsize=10, rotation=0, labelpad=30)
    
    im = axes[2, 0].pcolormesh(Epara_axis, anis_axis, max_growth.T, shading='auto', 
                          norm=colors.Normalize(), cmap='jet')
    fig.colorbar(im, cax=axes[2, 1], extend='both').set_label(
            'Growth Rate at max Frequency', fontsize=10, rotation=0, labelpad=30)
    
    for ax in axes[1:, 0]:
        ax.set_xlim(None, None)
        ax.set_ylim(None, None)
        
    if False:
        pass
    else:
        plt.show()
    return


#%% THESIS PLOT FUNCTIONS
def thesis_plot_summaries(_rbsp_path, _time_start, _time_end, _probe):
    '''
    For each value of He going up in 1% incremenets up to 30%, with O going up
    to 10% in increments of 0.5%, plots a 4 axes figure containing:
        -- Frequency (w) at max growth rate (for just He band?)
        -- Value at max growth rate
        -- Frequency (w) at max CGR
        -- Value at max CGR
    The frequencies are calculated by finding the k of max TGR/CGR and using
    the dispersion relation to find that w(k).
    
    -> For each He concentration
    ----> For every O concentration in that He bin
    -------> For each time in a specific composition
    -----------> Find max GR and corresponding k value
    -----------> Calculate the frequency matching this k by using the dispersion relation
    
    IDEA: Plot table of max growth rates for each time, maybe on a grid of 
    He vs O concentration. This can go in a folder somewhere. Not sure how to
    turn that into a timeseries.
    '''    
    nsec = 5
    save_string = _time_start.astype(object).strftime('%Y%m%d_%H%M_') \
                + _time_end.astype(object).strftime('%H%M')
    
    he_frac = np.arange(1.0, 30.5, 1.0)*1e-2
    o_frac  = np.arange(0.5, 10.5, 0.5)*1e-2
    _cmp    = np.array([1. - he_frac[1] - o_frac[1], he_frac[1], o_frac[1]], dtype=float)
    
    # Load sample (first) DR to get times
    k_np,  WPDR_out, wCGR_out, wVg_out, times, B0, name, mass, charge, density,\
        tper, ani, cold_dens, _cmp = get_all_DRs_warm_only(save_dir, _time_start,
        _time_end, _probe, _cmp, _nsec=5, HM_filter_mhz=50., N_procs=7, suff='',
        data_path=_rbsp_path, load_only=True)
    
    # Initialize empty arrays for GR returns (Times, species, value)
    # Values are 0: Max growth rate, 1: k at max growth, 2: frequency at max growth
    Nt      = times.shape[0]
    max_TGR = np.zeros((Nt, 3, 3), dtype=np.float64)
    
    # Collect max value timeseries, plot each line as its collected
    # Currently set to have 21 oxygen bins and 31 helium bins
    # So plot 31 lines on each oxygen plot, with 7 axes in each of 3 figs
    plt.ioff()
    n_figs = 3 ; n_axes = 7
    norm   = mpl.colors.Normalize(vmin=he_frac[0]*100., vmax=he_frac[-1]*100.)
    mapper = cm.ScalarMappable(norm=norm, cmap=cm.jet)
    
    fsize=10
    
    print('Plotting summaries...')
    # Each axes corresponds to a new oxygen concentration
    for mm in range(n_figs):
        
        # Colorbar to show He percentage
        fig = plt.figure(figsize=(8.00, 0.5*11.00))
        gs  = gspec.GridSpec(7, 2, figure=fig, width_ratios=[1, 0.01], hspace=0.0, wspace=0.05)
        cax = fig.add_subplot(gs[:, 1])
        
        for nn in range(n_axes):
            oxy_idx = mm*n_axes + nn
            o_comp  = o_frac[oxy_idx]
            
            # Set axes values
            ax = fig.add_subplot(gs[nn, 0])
            
            o_lbl = round(o_comp*1e2, 1)
            ax.set_ylabel(f'{o_lbl:.1f}% $O^+$', rotation=0, labelpad=25,
                          fontsize=fsize)#\n($\gamma \\times 10^3 s^{-1}$)
            ax.set_xlim(_time_start, _time_end)
            if nn == n_axes - 1:
                ax.set_xlabel('Time (UT)', fontsize=fsize)
                ax.xaxis.set_major_formatter(mdates.DateFormatter('%H:%M'))
            elif nn == 0:
                ax.set_title(f'Max. $\gamma$ :: {title_string} :: Heavy Ion Parameter Search',
                             fontsize=fsize+2)
                ax.set_xticklabels([])
            else:
                ax.set_xticklabels([])
            cbar = fig.colorbar(mapper, cax=cax, orientation='vertical')
            cbar.set_label(label='$He^{+}(\%)$', size='large', weight='bold',
                           rotation=90, labelpad=10, fontsize=fsize+2)
            
            ax.axvline(time_start, c='k', ls='--', alpha=0.5)
            ax.axvline(time_end  , c='k', ls='--', alpha=0.5)
            
            for he_comp in he_frac:
                h_comp = 1. - o_comp - he_comp
                _cmp   = np.array([h_comp, he_comp, o_comp], dtype=float)
                
                DR_path = get_DR_filepath(save_dir, _time_start, _time_end, _cmp, nsec)
                
                # Reminder: DRs are (time, wavenumber, band)
                if os.path.exists(DR_path):
                    k_vals,  WPDR, CGR, Vg, times, B0, name, mass, charge, density,\
                        tper, ani, cold_dens, cmp = get_all_DRs_warm_only(save_dir, 
                          _time_start, _time_end, probe, _cmp, 
                        kmin=0.0, kmax=1.5, Nk=1000, knorm=True,
                        _nsec=nsec, HM_filter_mhz=50, N_procs=1,
                        suff='', data_path=rbsp_path, load_only=True)
                else:
                    print('No growth rate file for composition {}/{}/{}'.format(h_comp, he_comp, o_comp))
                    continue
        
                # Remove nan's at start of arrays (Just copy for now, do smarter later)
                WPDR[:, 0, :] = WPDR[:, 1, :]
                Vg[  :, 0, :] = Vg[  :, 1, :]
                
                # Scale values for plotting/interpolating
                k_vals  *= 1e6
                TGR      = WPDR.imag*1e3
                freqs    = WPDR.real / (2*np.pi)    # Frequencies in Hz
                
                # Filter growth rates
                TGR[np.isnan(TGR) == True] = 0.0
                freqs[np.isnan(freqs) == True] = 0.0
                
                for ii in range(Nt):
                    for sp in range(3):
                        max_idx_TGR        = TGR[ii, :, sp].argmax()    # Index of maximum GR
                        max_TGR[ii, sp, 0] = TGR[ii, :, sp].max()       # Maximum growth rate
                        max_TGR[ii, sp, 1] = k_vals[ii, max_idx_TGR]    # Wavenumber at maximum GR
                        max_TGR[ii, sp, 2] = freqs[ii, max_idx_TGR, sp] # Frequency  at maximum GR
                
                # Add this line
                clr = mapper.to_rgba(he_comp*100.)
                ax.plot(times, max_TGR[:, 1, 0], c=clr)
            ax.set_ylim(0.0, glim)
                
        # Save Figure
        fig.tight_layout()
        fig.align_ylabels()
        
        fig_name = f'TGR_SUMMARIES_{save_string}_fig{mm}.png'
        fig.savefig(plot_path + fig_name, dpi=200)
        plt.close('all')
        print('Plot saved:', fig_name)
        break
    return


def thesis_plot_2D_growth_rates_with_time(_rbsp_path, _plot_start, _plot_end, _probe,
                                 approx='warm', save=True, nsec=None, log=False,
                                 _time_start=None, _time_end=None, plot_mag=True):
    '''
    plot_start and plot_end used because linear calculations done with these
    (time_start, time_end are less than these, so filenames go with the bigger)
    These define the files and existing calculations
    Use time_start, time_end to define limitations of plot (probably just event)
    
    Plot growth rates and show with magnetic field data.
     - Plot the Pc1
     - Want to have the 3 species k vs. t
    
    How best to illustrate for multiple species?
        - Get max k, find w(k_max)
        - Plot with time
        - Maybe overlay contour spectra above some power (that shows the packeting)
    '''
    plot_dir = plot_path
    #plot_dir = 'E://2D_LINEAR_THEORY//NEW_2D_GROWTH_RATES//'

    if not os.path.exists(plot_dir): os.makedirs(plot_dir)
    
    if _time_start is None: _time_start = _plot_start
    if _time_end is None: _time_end = _plot_end
    
    if plot_mag == True:
        pc1_xtimes, pc1_xfreq, pc1_power, _, _ = get_mag_data(_time_start, _time_end, probe,
                     _olap=0.95, _res=25.0,
                     _HM=False, HM_LP=50.0, HM_HP=None, 
                     _split_HM=False, _split_freq = 7.0,
                     transverse_only=False)
    
    # Load/Calculate each growth rate in sweep
    for he_comp in np.arange(1.0, 30.5, 1.0)*1e-2:
        for o_comp in np.arange(0.5, 10.5, 0.5)*1e-2:
            print('Loading...')
            if he_comp == 0.30 and o_comp == 0.06:
                pass
            else:
                continue
            h_comp = 1. - o_comp - he_comp
            print('Composition: {}/{}/{}'.format(h_comp, he_comp, o_comp))
            comp = np.array([h_comp, he_comp, o_comp], dtype=float)
            
            save_string = _plot_start.astype(object).strftime('%Y%m%d_%H%M_') \
                        + _plot_end.astype(object).strftime('%H%M')

            DR_path = get_DR_filepath(save_dir, _plot_start, _plot_end, comp, nsec)
            
            if os.path.exists(DR_path):
                k_vals,  WPDR, CGR, Vg, times, B0, name, mass, charge, density,\
                    tper, ani, cold_dens, cmp = get_all_DRs_warm_only(save_dir, 
                      _plot_start, _plot_end, probe, comp, 
                    kmin=0.0, kmax=1.5, Nk=1000, knorm=True,
                    _nsec=nsec, HM_filter_mhz=50, N_procs=1,
                    suff='', data_path=rbsp_path)
            else:
                pdb.set_trace()
                print('No growth rate file for composition {}/{}/{}'.format(h_comp, he_comp, o_comp))
                continue

            # Remove nan's at start of arrays (Just copy for now, do smarter later)
            WPDR[:, 0, :] = WPDR[:, 1, :]
            CGR[ :, 0, :] = CGR[ :, 1, :]
            Vg[  :, 0, :] = Vg[  :, 1, :]
            
            # Scale values for plotting/interpolating
            k_vals  *= 1e6
            TGR      = WPDR.imag*1e3
            CGR      = CGR*1e9
            freqs    = WPDR.real / (2*np.pi)
            max_f    = freqs[np.isnan(freqs) == False].max()
            
            time_2D = np.zeros(k_vals.shape, dtype=times.dtype)
            ti2d    = np.zeros(k_vals.shape, dtype=float)
            for ii in range(k_vals.shape[1]):
                time_2D[:, ii] = times[:]
                ti2d[   :, ii] = np.arange(times.shape[0])
            
            # Interpolate the frequency space values or load from file
            fGR_path    = save_dir + 'fGRw_{}_cc_{:03}_{:03}_{:03}_{}sec.npz'.format(
                             save_string, int(10*cmp[0]), int(10*cmp[1]),
                                          int(10*cmp[2]), nsec)
            
            if os.path.exists(fGR_path) == False:
                time_interp = np.arange(times.shape[0], dtype=float)
                freq_interp = np.linspace(0.0, max_f, 1000)
                xi, yi      = np.meshgrid(time_interp, freq_interp)
        
                TGRi  = np.zeros(k_vals.shape).flatten()
                CGRi  = np.zeros(k_vals.shape).flatten()
                
                try:
                    for ii in range(TGR.shape[2]):
                        x = ti2d.flatten()
                        y = freqs[:, :, ii].flatten()
                        z = TGR[:, :, ii].flatten()
                        print('Interpolating species TGR', ii)
                        TGRi[:] += griddata((x, y), z, (xi.flatten(), yi.flatten()),
                                            method='cubic', fill_value=0.0)
                    TGRi = TGRi.reshape(xi.shape)
                    
                    for ii in range(CGR.shape[2]):
                        x = ti2d.flatten()
                        y = freqs[:, :, ii].flatten()
                        z = CGR[:, :, ii].flatten()
                        print('Interpolating species CGR', ii)
                        CGRi[:] += griddata((x, y), z, (xi.flatten(), yi.flatten()),
                                            method='cubic', fill_value=0.0)
                    CGRi = CGRi.reshape(xi.shape)
                    print('Saving growth rate interpolation...')
                    np.savez(fGR_path, times=times, freq_interp=freq_interp, TGRi=TGRi, CGRi=CGRi)
                except:
                    pass
            else:
                print('Growth rate interpolation already exist, loading from file...')
                fGR_file    = np.load(fGR_path)
                times       = fGR_file['times']
                freq_interp = fGR_file['freq_interp']
                TGRi        = fGR_file['TGRi']
                CGRi        = fGR_file['CGRi']
            
            # Set limits        
            kmin     = 0.0
            kmax     = 35.0#k_vals.max()
        
            TGR_max  = TGR[np.isnan(TGR) == False].max()
            #CGR_max  = CGR[np.isnan(CGR) == False].max()
            #TGR_min  = None
            
            ##########
            ## PLOT ##
            ##########
            plt.ioff()
            fsize = 10; X=-0.10; Y=0.00
            ## TEMPORAL GROWTH RATE ##
            fig1, axes1 = plt.subplots(5, 2, figsize=(8.0, 0.5*11.0), 
                                       gridspec_kw={'width_ratios':[1, 0.01]})
            
            axes1[0, 0].set_title('Temporal Growth Rate :: {} :: Ion Composition {}/{}/{}'.format(
                                        title_string, (cmp[0]), (cmp[1]),
                                          (cmp[2])), fontsize=fsize)
            with warnings.catch_warnings():
                warnings.simplefilter("ignore")
                
                # Pc1 Spectra
                im1a = axes1[0, 0].pcolormesh(pc1_xtimes, pc1_xfreq, pc1_power,
                                           norm=colors.LogNorm(vmin=1e-4, vmax=1e1),
                                           cmap='jet')
                axes1[0, 0].set_ylabel('f\n(Hz)', fontsize=fsize, rotation=0)
                axes1[0, 0].yaxis.set_label_coords(X, 0.4)
                cbar1a = fig1.colorbar(im1a, cax=axes1[0, 1], orientation='vertical')
                cbar1a.set_label(label='$P_\perp$\n$(nT^2/Hz)$', rotation=0, 
                               labelpad=20, fontsize=fsize)
                #cbar1a.ax.yaxis.set_major_locator(MaxNLocator(prune='lower'))
                
                # H+ Growth Rate
                im2a = axes1[1, 0].pcolormesh(time_2D, k_vals, TGR[:, :, 0], cmap='viridis', vmin=0.0)
                axes1[1, 0].set_ylabel('$H^+ k_\parallel$\n($10^6$/m)\n', fontsize=fsize, rotation=0)
                axes1[1, 0].yaxis.set_label_coords(X, Y)
                cbar2a = fig1.colorbar(im2a, cax=axes1[1, 1], orientation='vertical')
                cbar2a.set_label(label='$\gamma$\n$(\\times 10^3 s^{-1})$', rotation=0,
                                 labelpad=20, fontsize=fsize)
                
                # He+ Growth Rate
                im3a = axes1[2, 0].pcolormesh(time_2D, k_vals, TGR[:, :, 1], cmap='viridis', vmin=0.0)
                axes1[2, 0].set_ylabel('$He^+ k_\parallel$\n($10^6$/m)\n', fontsize=fsize, rotation=0)
                axes1[2, 0].yaxis.set_label_coords(X, Y)
                cbar3a = fig1.colorbar(im3a, cax=axes1[2, 1], orientation='vertical')
                cbar3a.set_label(label='$\gamma$\n$(\\times 10^3 s^{-1})$', rotation=0,
                                 labelpad=20, fontsize=fsize)
                
                # O+ Growth Rate
                im4a = axes1[3, 0].pcolormesh(time_2D, k_vals, TGR[:, :, 2], cmap='viridis', vmin=0.0)
                axes1[3, 0].set_ylabel('$O^+ k_\parallel$\n($10^6$/m)\n', fontsize=fsize, rotation=0)
                axes1[3, 0].yaxis.set_label_coords(X, Y)
                cbar4a = fig1.colorbar(im4a, cax=axes1[3, 1], orientation='vertical')
                cbar4a.set_label(label='$\gamma$\n$(\\times 10^3 s^{-1})$', rotation=0,
                                 labelpad=20, fontsize=fsize)
                
                im5a = axes1[4, 0].pcolormesh(times, freq_interp, TGRi, cmap='viridis', vmin=0.0, vmax=TGR_max)
                axes1[4, 0].set_ylabel('f\n(Hz)', fontsize=fsize, rotation=0)
                axes1[4, 0].yaxis.set_label_coords(X, 0.4)
                cbar5a = fig1.colorbar(im5a, cax=axes1[4, 1], orientation='vertical')
                cbar5a.set_label(label='$\gamma$\n$(\\times 10^3 s^{-1})$', rotation=0,
                                 labelpad=20, fontsize=fsize)
                
            for ax in axes1[:, 0]:                   
                ax.set_xlim(_time_start, _time_end)
                
                if ax != axes1[-1, 0]:
                    ax.set_xticklabels([])
                else:
                    ax.set_xlabel('Time (UT)')
                    ax.xaxis.set_major_locator(mdates.MinuteLocator(interval=3))
                    ax.xaxis.set_major_formatter(mdates.DateFormatter('%H:%M'))
                
                if ax == axes1[0, 0] or ax == axes1[-1, 0]:
                    ax.set_ylim(0.0, f_max)
                else:
                    ax.set_ylim(kmin, kmax)
                    
            fig1.tight_layout()
            fig1.align_ylabels()
            fig1.subplots_adjust(hspace=0.15, wspace=0.05)
        
            if save==False:
                plt.show()
            else:
                fig1.savefig(plot_dir + 'TGR_{}_cc_{:03}_{:03}_{:03}_{}sec.png'.format(
                             save_string, int(10*cmp[0]), int(10*cmp[1]),
                                          int(10*cmp[2]), nsec), dpi=200)
                print('Plot saved.')
                plt.close('all')
    return


def thesis_plot_single_2D_growth_timeseries(_rbsp_path, _time_start, _time_end, _probe,
                                 approx='warm', save=True, nsec=None, log=False):
    '''
    plot_start and plot_end used because linear calculations done with these
    (time_start, time_end are less than these, so filenames go with the bigger)
    These define the files and existing calculations
    Use time_start, time_end to define limitations of plot (probably just event)
    
    Plot growth rates and show with magnetic field data.
     - Plot the Pc1
     - Want to have the 3 species k vs. t
    
    How best to illustrate for multiple species?
        - Get max k, find w(k_max)
        - Plot with time
        - Maybe overlay contour spectra above some power (that shows the packeting)
    '''
    save_dir = '{}//2D_LINEAR_THEORY//EVENT_{}//'.format(ext_drive, date_string)
    plot_dir = plot_path
    if not os.path.exists(plot_dir): os.makedirs(plot_dir)
        
    # Load/Calculate each growth rate in sweep            
    k_vals,  WPDR, CGR, Vg, times, B0, name, mass, charge, density,\
        tper, ani, cold_dens = get_all_DRs_with_cutoffs(
                     save_dir, time_start, time_end, 'a',  
                     kmin=0.0, kmax=1.0, Nk=1000, knorm=True,
                     _nsec=5, HM_filter_mhz=50, N_procs=7,
                     he_frac=0.30, data_path=f'{ext_drive}//DATA//RBSP//',
                     output=True, approx='hot', load_only=True)

    pc1_xtimes, pc1_xfreq, pc1_power, _, _ = get_mag_data(_time_start, _time_end, _probe,
                 _olap=0.95, _res=25.0,
                 _HM=False, HM_LP=50.0, HM_HP=None, 
                 _split_HM=False, _split_freq = 7.0,
                 transverse_only=False)
    
    # Remove nan's at start of arrays (Just copy for now, do smarter later)
    WPDR[:, 0, :] = WPDR[:, 1, :]
    CGR[ :, 0, :] = CGR[ :, 1, :]
    Vg[  :, 0, :] = Vg[  :, 1, :]
            
    # Scale values for plotting/interpolating
    k_vals  *= 1e6
    TGR      = WPDR.imag*1e3
    CGR      = CGR*1e9
    freqs    = WPDR.real / (2*np.pi)
    max_f    = freqs[np.isnan(freqs) == False].max()
            
    time_2D = np.zeros(k_vals.shape, dtype=times.dtype)
    ti2d    = np.zeros(k_vals.shape, dtype=float)
    for ii in range(k_vals.shape[1]):
        time_2D[:, ii] = times[:]
        ti2d[   :, ii] = np.arange(times.shape[0])

    time_interp = np.arange(times.shape[0], dtype=float)
    freq_interp = np.linspace(0.0, max_f, 1000)
    xi, yi      = np.meshgrid(time_interp, freq_interp)

    TGRi  = np.zeros(k_vals.shape).flatten()
    try:
        for ii in range(TGR.shape[2]):
            x = ti2d.flatten()
            y = freqs[:, :, ii].flatten()
            z = TGR[:, :, ii].flatten()
            print('Interpolating species TGR', ii)
            TGRi[:] += griddata((x, y), z, (xi.flatten(), yi.flatten()),
                                method='cubic', fill_value=0.0)
        TGRi = TGRi.reshape(xi.shape)
    except:
        print('Error, no interpolation possible.')
        pass
            
    # Set limits        
    kmin     = 0.0
    kmax     = 35.0#k_vals.max()
    TGR_max  = TGR[np.isnan(TGR) == False].max()

            
    ##########
    ## PLOT ##
    ##########
    plt.ioff()
    fsize = 10; X=-0.10; Y=0.00
    ## TEMPORAL GROWTH RATE ##
    fig1, axes1 = plt.subplots(5, 2, figsize=(8.0, 0.5*11.0), 
                               gridspec_kw={'width_ratios':[1, 0.01]})
    
    axes1[0, 0].set_title('Temporal Growth Rate :: {} :: 30% $He^+$'.format(
                                title_string), fontsize=fsize)
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        
        # Pc1 Spectra
        im1a = axes1[0, 0].pcolormesh(pc1_xtimes, pc1_xfreq, pc1_power,
                                   norm=colors.LogNorm(vmin=1e-4, vmax=1e1),
                                   cmap='jet')
        axes1[0, 0].set_ylabel('f\n(Hz)', fontsize=fsize, rotation=0)
        axes1[0, 0].yaxis.set_label_coords(X, 0.4)
        cbar1a = fig1.colorbar(im1a, cax=axes1[0, 1], orientation='vertical')
        cbar1a.set_label(label='$P_\perp$\n$(nT^2/Hz)$', rotation=0, 
                       labelpad=20, fontsize=fsize)
        #cbar1a.ax.yaxis.set_major_locator(MaxNLocator(prune='lower'))
        
        # H+ Growth Rate
        im2a = axes1[1, 0].pcolormesh(time_2D, k_vals, TGR[:, :, 0], cmap='viridis', vmin=0.0)
        axes1[1, 0].set_ylabel('$H^+ k_\parallel$\n($10^6$/m)\n', fontsize=fsize, rotation=0)
        axes1[1, 0].yaxis.set_label_coords(X, Y)
        cbar2a = fig1.colorbar(im2a, cax=axes1[1, 1], orientation='vertical')
        cbar2a.set_label(label='$\gamma$\n$(\\times 10^3 s^{-1})$', rotation=0,
                         labelpad=20, fontsize=fsize)
        
        # He+ Growth Rate
        im3a = axes1[2, 0].pcolormesh(time_2D, k_vals, TGR[:, :, 1], cmap='viridis', vmin=0.0)
        axes1[2, 0].set_ylabel('$He^+ k_\parallel$\n($10^6$/m)\n', fontsize=fsize, rotation=0)
        axes1[2, 0].yaxis.set_label_coords(X, Y)
        cbar3a = fig1.colorbar(im3a, cax=axes1[2, 1], orientation='vertical')
        cbar3a.set_label(label='$\gamma$\n$(\\times 10^3 s^{-1})$', rotation=0,
                         labelpad=20, fontsize=fsize)
        
        # O+ Growth Rate
        im4a = axes1[3, 0].pcolormesh(time_2D, k_vals, TGR[:, :, 2], cmap='viridis', vmin=0.0)
        axes1[3, 0].set_ylabel('$O^+ k_\parallel$\n($10^6$/m)\n', fontsize=fsize, rotation=0)
        axes1[3, 0].yaxis.set_label_coords(X, Y)
        cbar4a = fig1.colorbar(im4a, cax=axes1[3, 1], orientation='vertical')
        cbar4a.set_label(label='$\gamma$\n$(\\times 10^3 s^{-1})$', rotation=0,
                         labelpad=20, fontsize=fsize)
        
        try:
            im5a = axes1[4, 0].pcolormesh(times, freq_interp, TGRi, cmap='viridis', vmin=0.0, vmax=TGR_max)
            axes1[4, 0].set_ylabel('f\n(Hz)', fontsize=fsize, rotation=0)
            axes1[4, 0].yaxis.set_label_coords(X, 0.4)
            cbar5a = fig1.colorbar(im5a, cax=axes1[4, 1], orientation='vertical')
            cbar5a.set_label(label='$\gamma$\n$(\\times 10^3 s^{-1})$', rotation=0,
                             labelpad=20, fontsize=fsize)
        except:
            pass
        
    for ax in axes1[:, 0]:                   
        ax.set_xlim(_time_start, _time_end)
        
        if ax != axes1[-1, 0]:
            ax.set_xticklabels([])
        else:
            ax.set_xlabel('Time (UT)')
            ax.xaxis.set_major_locator(mdates.MinuteLocator(interval=3))
            ax.xaxis.set_major_formatter(mdates.DateFormatter('%H:%M'))
        
        if ax == axes1[0, 0] or ax == axes1[-1, 0]:
            ax.set_ylim(0.0, f_max)
        else:
            ax.set_ylim(kmin, kmax)
            
    fig1.tight_layout()
    fig1.align_ylabels()
    fig1.subplots_adjust(hspace=0.15, wspace=0.05)

    if save==False:
        plt.show()
    else:
        fig1.savefig(plot_dir + 'TGR_CUTOFFS_{}sec.png'.format(nsec), dpi=200)
        print('Plot saved.')
        plt.close('all')
    return


def check_concentration_files(_date):
    '''
    For event date, check whether or not a certain concentration has been done
    Sweep function went in incremenets of 0.5%
    '''
    import openpyxl as pyxl
    
    _date         = _date.astype(object)
    daystring     = '{}{:02}{:02}'.format(_date.year, _date.month, _date.day)
    linear_folder = 'E://2D_LINEAR_THEORY//EVENT_{}//'.format(daystring)
    wbook         = 'D://Google Drive//Uni//PhD 2017//Data//Events//events_list_all_in.xlsx'
    sheet         = 'LINEAR_{}'.format(daystring)
    
    wb    = pyxl.load_workbook(wbook)
    sheet = wb[sheet]
    
    print('Checking files...')
    row = 2
    for ii in np.arange(0.0, 30.5, 0.5):
        col = 2
        for jj in np.arange(0.0, 10.5, 0.5):
            # Define ion concentrations
            H_comp  = '{:03}'.format(int(10*(100. - ii - jj)))
            He_comp = '{:03}'.format(int(10*ii))
            O_comp  = '{:03}'.format(int(10*jj))
            
            # Look for file containing these concentrations
            found = 0
            keywords = ['DISPw', daystring, H_comp, He_comp, O_comp]
            for file in os.listdir(linear_folder):
                if all(keyword in file for keyword in keywords):
                    found = 1
                    break
            
            # Return result
            if found == 1:
                sheet.cell(row=row, column=col).value = 'Y'
            else:
                sheet.cell(row=row, column=col).value = 'N'
                
            col += 1
        row += 1
        
    print('Saving workbook...')
    wb.save(wbook)
    print('Workbook saved.')
    return


def plot_single_kCGR():
    '''
    Remember units are nT, /cm3, eV for this function
    
    Lower energy gives higher peak frequency
    '''
    B0         = 161.0
    ndensc     = np.array([70.0, 0.0, 0.0])
    ndensw     = np.array([0.7, 0.0, 0.0])
    anisotropy = np.array([0.5, 0.0, 0.0])
    tperp      = np.array([20., 0.0, 0.0])
    
    plt.ioff()
    plt.figure()
    
    freq, growth, stop = convective_growth_rate_kozyra(B0, ndensc, ndensw, anisotropy, temperp=tperp,
                                                       norm_ampl=0, NPTS=1000, maxfreq=0.5)
    
    plt.plot(freq, growth)
    plt.xlabel('Frequency (Hz)')
    plt.ylabel('S (/cm)')
    plt.xlim(0.0, None)
    plt.show()
    sys.exit()
    return


def plot_density_sweep_kCGR():
    '''
    Remember units are nT, /cm3, eV for this function
    
    Lower energy gives higher peak frequency
    
    For varying density:
        -- Increase hot protons, keep cold constant
        -- Increase hot ion fraction of H+
    '''
    # Starting parameters
    B0         = 250
    nH         = 190.49
    ndensc     = np.array([184.142 , 88.038 , 14.673])
    ndensw     = np.array([6.607   , 0.0    , 0.0   ])
    anisotropy = np.array([0.35    , 0.0    , 0.0   ])
    tperp      = np.array([11.5    , 0.0    , 0.0   ])*1e3
    
    wdens_abs  = ndensw.copy()
    wdens_frac = ndensw.copy()
    cdens_frac = ndensc.copy()
    
    plt.ioff()
    fig, axes = plt.subplots(2, figsize=(8.0, 0.5*11.00))
    
    gyfreq = qp * B0*1e-9 / (2*np.pi*mp * np.array([1.0, 4.0, 16.0]))
    for factor in [3.5, 5.0, 7.0, 10.0, 15.0, 25.0, 50.0]:
        
        # Fractional variation
        wdens_frac[0] = nH*factor*1e-2
        cdens_frac[0] = nH - wdens_frac[0]
        
        freq, growth, stop = convective_growth_rate_kozyra(B0, ndensc, wdens_frac, anisotropy, temperp=tperp,
                                                           norm_ampl=0, NPTS=1000, maxfreq=0.5)
        axes[0].plot(freq, growth, label='%.1f%% $n_{H^+, h}$'%factor)
        
        # Add line to follow frequency
        he_low_idx, he_hi_idx = ascr.boundary_idx64(freq, gyfreq[2], gyfreq[1]) 
        peak_idx  = growth[he_low_idx: he_hi_idx].argmax() + he_low_idx
        peak_freq = freq[peak_idx]
        axes[0].axvline(peak_freq, c='k', ls='--', alpha=0.5)
        
    axes[0].set_title('Convective Growth Rate :: Varying hot proton fraction')
    axes[0].legend()
    


    for factor in [0.1, 0.25, 0.5, 1.0, 1.2, 1.5, 2.0, 3.0, 5.0, 10.0]:
        # Absolute variation
        wdens_abs[0] = ndensw[0]*factor
        freq, growth, stop = convective_growth_rate_kozyra(B0, ndensc, wdens_abs, anisotropy, temperp=tperp,
                                                           norm_ampl=0, NPTS=1000, maxfreq=0.5)
        axes[1].plot(freq, growth, label='$n_{H^+, h} = %.2f cc$'%wdens_abs[0])
        
        # Add line to follow frequency
        he_low_idx, he_hi_idx = ascr.boundary_idx64(freq, gyfreq[2], gyfreq[1]) 
        peak_idx  = growth[he_low_idx: he_hi_idx].argmax() + he_low_idx
        peak_freq = freq[peak_idx]
        axes[1].axvline(peak_freq, c='k', ls='--', alpha=0.5)
        
    axes[1].set_title('Convective Growth Rate :: Varying hot proton density') 
    axes[1].legend()
    
    for ax in axes:
        ax.set_xlabel('Frequency (Hz)')
        ax.set_ylabel('S (/cm)')
        ax.set_xlim(0.0, None)
    
    fig.show()
    return


def plot_temperature_sweep_kCGR():
    '''
    Remember units are nT, /cm3, eV for this function
    
    Lower energy gives higher peak frequency
    
    For varying density:
        -- Increase hot protons, keep cold constant
        -- Increase hot ion fraction of H+
    '''
    B0         = 249.6
    ndensc     = np.array([184.142 , 88.038 , 14.673])
    ndensw     = np.array([6.607   , 0.0    , 0.0   ])
    anisotropy = np.array([0.35    , 0.0    , 0.0   ])
    tperp      = np.array([11.5    , 0.0    , 0.0   ])*1e3
    
    plt.ioff()
    fig, ax = plt.subplots(figsize=(8.0, 0.5*11.00))
    
    gyfreq = qp * B0*1e-9 / (2*np.pi*mp * np.array([1.0, 4.0, 16.0]))
    for factor in [0.5, 0.75, 1.0, 1.2, 1.5, 2.0, 2.5, 3.0, 4.0, 5.0, 7.5, 10.0]:  
        tp = tperp*factor
        freq, growth, stop = convective_growth_rate_kozyra(B0, ndensc, ndensw, anisotropy, temperp=tp,
                                                           norm_ampl=0, NPTS=1000, maxfreq=0.5)
        he_low_idx, he_hi_idx = ascr.boundary_idx64(freq, gyfreq[2], gyfreq[1]) 
        peak_idx  = growth[he_low_idx: he_hi_idx].argmax() + he_low_idx
        peak_freq = freq[peak_idx]
        
        ax.plot(freq, growth, label='%.1f%% $n_{H^+, h}$'%factor)
        ax.axvline(peak_freq, c='k', ls='--', alpha=0.5)
    
    ax.set_title('Convective Growth Rate :: Varying hot proton fraction')
    ax.legend()
    ax.set_xlabel('Frequency (Hz)')
    ax.set_ylabel('S (/cm)')
    ax.set_xlim(0.0, None)
    
    fig.show()
    return


def investigate_anisotropy(save=True):
    '''
    To look at the amount of variation in the growth rate from small changes
    in the anisotropy.
    Use both Kozyra and multiapprox methods to look at both the temporal and
    convective changes.
    
    Except it's not just the change in the anisotropy, but also the change in |B|.
    
    Test: 
        -- Do a sweep, just changing A
        -- Do another sweep, but change B also
        -- Is there an appreciable difference? How much to change B by?
        
    16/1 Average Parameters:
        |B| ~ 90 nT
        ne  ~ 140 cc (150-120cc)
        nh  ~ 1.2, 0.07, 0.35 (H, He, O)
        Th  ~ 7.8, 7.2, 4.8 keV
        Ah  ~ 0.38, 0.35, 0.0
        
    Try A 0.6 -> 0.5, B 86 - 92
    
    DO LATER
    '''
    n_cold     = 90.0
    B0         = 150.0
    ndensc     = np.array([0.65,  0.05, 0.05]) * n_cold
    ndensw     = np.array([1.20,  0.00, 0.00])
    anisotropy = np.array([0.38,  0.00, 0.00])
    tperp      = np.array([20.0,  0.00, 4.80])*1e3
    
    plt.ioff()
    fig, axes = plt.subplots(nrows=1, ncols=2, figsize=(8.0, 0.5*11.0),
                     gridspec_kw={'width_ratios':[1, 0.2]})
    
    gyfreq = qp * B0*1e-9 / (2*np.pi*mp * np.array([1.0, 4.0, 16.0]))
    all_anis = np.array([0.2, 0.4, 0.6, 0.8, 1.0, 1.2, 1.4])
    
    for anis in all_anis: 
        alpha = anis / all_anis.max()
        anisotropy[0] = anis
        freq, growth, stop = convective_growth_rate_kozyra(B0, ndensc, ndensw, anisotropy, temperp=tperp,
                                                           norm_ampl=0, NPTS=1000, maxfreq=1.0)
        
        axes[0].plot(freq, growth, label='A = {}'.format(anis), alpha=alpha, c='r')
    
    axes[0].set_title('Convective Growth Rate :: 1991-08-12 Parameters')
    axes[0].legend()
    axes[0].set_xlabel('Frequency (Hz)')
    axes[0].set_ylabel('S $(cm^{-1})$')
    axes[0].set_xlim(0.0, 1.0)
    axes[1].set_visible(False)
    
    Lx = 0.835; Ly = 0.90; dy = 0.03
    fig.text(Lx, Ly - 0*dy, '$B_0$ = %.1f nT'%B0)    
    fig.text(Lx, Ly - 1*dy, '$n_c$ = %.1f cc'%n_cold)    
    fig.text(Lx, Ly - 2*dy, '$n_h$  = %.1f cc' % (ndensw[0]))
    fig.text(Lx, Ly - 3*dy, '$T_\perp$  = %.1f keV' % (tperp[0]*1e-3))
    
    fig.text(Lx, Ly - 5*dy, '$H^+$  = %5.2f %%' % (ndensc[0]/n_cold*100.))    
    fig.text(Lx, Ly - 6*dy, '$He^+$ = %5.2f %%' % (ndensc[1]/n_cold*100.))
    fig.text(Lx, Ly - 7*dy, '$O^+$  = %5.2f %%' % (ndensc[2]/n_cold*100.))
    
    fig.tight_layout()
    if save:
        save_path = plot_path
        if not os.path.exists(save_path): os.makedirs(save_path)
        save_name = save_path + '19910812_anisotropy.png'
        print('Saving {}'.format(save_name))
        fig.savefig(save_name)
        plt.close('all')
    else:
        plt.show()
    return


def get_CRRES_kCGR_timeseries():
    '''
    To do: 
        -- Incorporate cutoffs for variable cold plasma composition
    
    Normalize for now, but use the same tactic as for variable k (using time_2D)
    '''
    _crres_path = '%s//DATA//CRRES//' % ext_drive
    if False:
        _time_start = np.datetime64('1991-07-17T20:15:00.000000')
        _time_end   = np.datetime64('1991-07-17T21:00:00.000000')
        f_event     = 0.165     # Event mean/strongest frequency
        f_max       = 0.40
    else:
        # Hot proton fluxes at 15-20keV on E_perp flux plot (90 deg PA bin)
        _time_start = np.datetime64('1991-08-12T22:10:00.000000')
        _time_end   = np.datetime64('1991-08-12T23:15:00.000000')
    title_string= _time_start.astype(object).strftime('%Y-%m-%d')
    #date_string = _time_start.astype(object).strftime('%Y%m%d')
    #save_dir    = '{}//2D_LINEAR_THEORY//EVENT_{}//'.format(ext_drive, date_string)
    #plot_path   = 'D://Google Drive//Uni//PhD 2017//Josh PhD Share Folder//Thesis//Data_Plots//'
    
    # Load CRRES data
    times, B0, ne = data.load_CRRES_data(_time_start, _time_end, crres_path=_crres_path)

    mag_times, _B0, HM, dB, E0, HMe, dE, S, B, E = cfr.get_crres_fields(_crres_path,
                  _time_start, _time_end, pad=0, E_ratio=5.0, rotation_method='vector', 
                  output_raw_B=True, interpolate_nan=None, B0_LP=1.0,
                  Pc1_LP=5000, Pc1_HP=100, Pc5_LP=30, Pc5_HP=None, dEx_LP=None)

    ntimes   = times.shape[0]
    npts     = 1000
    all_f    = np.zeros((ntimes, npts), dtype=np.float64)
    all_CGR  = np.zeros((ntimes, npts), dtype=np.float64)
    all_stop = np.zeros((ntimes, npts), dtype=np.float64)
    
    for ii in range(ntimes): 
        print('Calculating time', times[ii])
        ndensc     = np.array([0.88, 0.10, 0.01]) * ne[ii]
        ndensw     = np.array([0.01,  0.0,  0.0]) * ne[ii]
        anisotropy = np.array([0.70,  0.0,  0.0])
        tperp      = np.array([20.0,  0.0,  0.0]) * 1e3
    
        all_f[ii], all_CGR[ii], all_stop[ii] = convective_growth_rate_kozyra(
                            B0[ii], ndensc, ndensw, anisotropy, temperp=tperp,
                            norm_ampl=0, norm_freq=0, NPTS=npts, maxfreq=1.0)
    
    ###
    ### DO PLOT 
    ###
    time_2D = np.zeros((ntimes, npts), dtype=times.dtype)
    for ii in range(npts):
        time_2D[:, ii] = times[:]
    
    plt.ioff()
    fig, axes = plt.subplots(nrows=3, figsize=(16, 10), sharex=True) 
                               #gridspec_kw={'width_ratios':[2, 0.05]})
    
    axes[0].set_title('Convective Growth Rate :: CRRES :: {}'.format(title_string))

# =============================================================================
#     im1a = axes[0].pcolormesh(times, pc1_xfreq, pc1_perp_power,
#                                vmin=-5, vmax=0, cmap='jet', shading='auto')
# =============================================================================
# =============================================================================
#     im2 = axes[1].pcolormesh(times, all_f[0], all_CGR.T,
#                     cmap='jet')
# =============================================================================
    
    im2 = axes[1].pcolormesh(time_2D, all_f, all_CGR, cmap='viridis')
# =============================================================================
#     cbar = fig.colorbar(im2, cax=axes[1, 1], extend='both').set_label(
#             'Growth Rate', fontsize=10, rotation=0, labelpad=30)
# =============================================================================
    pdb.set_trace()
    axes[2].plot(times, ne, c='k')
    axes[2].set_ylabel('cc')
    ax3 = axes[2].twinx()
    ax3.plot(times, B0, c='b')
    ax3.set_ylabel('nT', color='b')
    plt.show()
    return


def plot_CGR_heavyions_search(B0, ne, rc_tperp, rc_anis, rc_frac=0.01, band=None,
                              he_min=0.0, he_max=50.0, nHe=500,
                              o_min=0.0, o_max=30.0, nO=500,
                              f_max=0.7, target_freqs=None, save=False):
    '''
    Band either 1, 2, 3 (for H, He, O band) or None for absolute max
    
    RC currently fraction of ne. Should be fraction of nH but that's harder.
    Add option later
    
    Aug12
    |B| = 161 \pm ~1nT for whole event
    ne  = 20-100cc (big variation). Doesn't seem to change frequency. Set at ~70cc
    Tperp ~ 15-20 keV
    A ~ 0.5
    ni ~ 1% or so
    '''    
    # Gyrofrequencies
    gyfreq = qp * B0*1e-9 / (2*np.pi*mp * np.array([1.0, 4.0, 16.0]))
    
    # Do parameter search in heavy ion space
    nHe_axis = np.linspace(he_min, he_max, nHe)*1e-2
    nO_axis  = np.linspace( o_min,  o_max, nO)*1e-2
    max_freq   = np.zeros((nHe, nO), dtype=np.float64)
    max_growth = np.zeros((nHe, nO), dtype=np.float64)
    
    ANI    = np.array([rc_anis, 0.0, 0.0])
    tperp  = np.array([rc_tperp, 0.0, 0.0])
    ndensw = np.array([rc_frac, 0.0   , 0.0 ])*ne
            
    for ii in range(nHe):
        #print(f'Calculating for {nHe_axis[ii]}% He')
        for jj in range(nO):
            cHe = nHe_axis[ii]
            cO  = nO_axis[jj]
            cH  = 1. - cHe - cO - rc_frac
            ndensc = np.array([cH, cHe, cO])*ne

            # Calculate growth rate
            try:
                freq, growth, stop = convective_growth_rate_kozyra(
                    B0, ndensc, ndensw, ANI, tperp,
                    norm_ampl=0, norm_freq=0, NPTS=1000, maxfreq=1.0)
                
                he_low_idx, he_hi_idx = ascr.boundary_idx64(freq, gyfreq[2], gyfreq[1]) 
                peak_idx  = growth[he_low_idx: he_hi_idx].argmax() + he_low_idx
    
                max_freq[ii, jj] = freq[peak_idx]
                max_growth[ii, jj] = growth[peak_idx]
                    
            except:
                print(f'ERROR: {cHe*100.}% He, {cO*100.}% O')
                max_freq[ii, jj] = np.nan
                max_growth[ii, jj] = np.nan
            
    
    # Plot results:
    plt.ioff()
    fig, axes = plt.subplots(nrows=2, ncols=3, figsize=(8.0, 11.0),
                 gridspec_kw={
                 'width_ratios':[1, 0.01, 0.4],
                 'height_ratios':[1, 1]})
    
    axes[0, 0].set_title('Heavy Ion Parameter Search :: Aug 12 1991')
    
    im = axes[0, 0].pcolormesh(nHe_axis, nO_axis, max_freq.T, shading='auto', 
                          norm=colors.Normalize(vmin=0.0, vmax=f_max), cmap='jet')
    fig.colorbar(im, cax=axes[0, 1], extend='both').set_label(
            'Frequency of max S')
    
    if target_freqs is not None:
        cs1 = axes[0, 0].contour(nHe_axis, nO_axis, max_freq.T, target_freqs, colors='k')
        cs2 = axes[1, 0].contour(nHe_axis, nO_axis, max_freq.T, target_freqs, colors='k')
    
        axes[0, 0].clabel(cs1, inline=True)
        axes[1, 0].clabel(cs2, inline=True)
    
    im = axes[1, 0].pcolormesh(nHe_axis, nO_axis, max_growth.T, shading='auto', 
                          norm=colors.Normalize(), cmap='jet')
    fig.colorbar(im, cax=axes[1, 1], extend='both').set_label(
            'Growth Rate at max Frequency')
    
    axes[0, 2].set_visible(False)
    axes[1, 2].set_visible(False)
    
    # Figure text for parameters
    Lx = 0.84; Ly = 0.95; dy = 0.02
    fig.text(Lx, Ly - 0*dy, '$B_0$ = %.1f nT'%B0)    
    fig.text(Lx, Ly - 1*dy, '$n_e$ = %.1f cc'%ne)    
    
    fig.text(Lx, Ly - 2*dy, '$n_h$ = %.2f ne'%rc_frac)    
    fig.text(Lx, Ly - 3*dy, '$T_h$ = %.2f keV'%(rc_tperp*1e-3))    
    fig.text(Lx, Ly - 4*dy, '$A_h$ = %.2f'%rc_anis)
    
    fig.tight_layout()
    fig.subplots_adjust(wspace=0.01)
    
    for ax in axes[:, 0]:
        ax.set_xlim(nHe_axis[0], nHe_axis[-1])
        ax.set_ylim(nO_axis[0], nO_axis[-1])

    if save:
        save_path = plot_path + '//CRRES_HEAVYION_19910812//'
        if not os.path.exists(save_path): os.makedirs(save_path)
        n_plot = len(os.listdir(save_path))
        
        save_name = save_path + '//CRRES_HEAVYION_PLOT{:03}.png'.format(n_plot)
        print('Saving {}'.format(save_name))
        fig.savefig(save_name)
        plt.close('all')
    else:
        # Only shows the first one
        figManager = plt.get_current_fig_manager()
        figManager.window.showMaximized()
    return


def search_TA_space(B0, ne, cHe, cO, rc_frac=0.01, band=None,
                    T_min=0.0, T_max=50.0, nT=100,
                    A_min=0.0, A_max=2.00, nA=100,
                    f_max=0.7, target_freqs=None, save=False):
    '''
    For some combination of cHe+ and cO+ heavy ions, calculate the max frequency
    and growth rate in the He band for T between 1keV-50keV and A - 0.0, 2.0
    
    Concentrations in fractions
    Temperature in keV
    B0 in nT
    ne in cc
    '''    
    # Gyrofrequencies
    gyfreq = qp * B0*1e-9 / (2*np.pi*mp * np.array([1.0, 4.0, 16.0]))
    
    # Do parameter search in heavy ion space
    T_axis = np.linspace(T_min, T_max, nT)*1e3
    A_axis = np.linspace(A_min, A_max, nA)
    max_freq   = np.zeros((nT, nA), dtype=np.float64)
    max_growth = np.zeros((nT, nA), dtype=np.float64)
    
    cH     = 1. - cHe - cO
    ndensc = np.array([cH*(1 - rc_frac), cHe, cO ])*ne
    ndensw = np.array([cH*     rc_frac , 0.0, 0.0])*ne
    
    print('Calculating...')
    for ii in range(1, nT):
        for jj in range(1, nA):
            tperp = np.array([T_axis[ii], 0.0, 0.0])
            ANI   = np.array([A_axis[jj], 0.0, 0.0])

            # Calculate growth rate
            try:
                freq, growth, stop = convective_growth_rate_kozyra(
                    B0, ndensc, ndensw, ANI, tperp,
                    norm_ampl=0, norm_freq=0, NPTS=1000, maxfreq=1.0)
                
                he_low_idx, he_hi_idx = ascr.boundary_idx64(freq, gyfreq[2], gyfreq[1]) 
                peak_idx  = growth[he_low_idx: he_hi_idx].argmax() + he_low_idx
    
                max_freq[ii, jj] = freq[peak_idx]
                max_growth[ii, jj] = growth[peak_idx]
                    
            except:
                max_freq[ii, jj] = np.nan
                max_growth[ii, jj] = np.nan
    
    try:
    #if True:
        print('Plotting')
        # Plot results:
        plt.ioff()
        fig, axes = plt.subplots(nrows=1, ncols=3, figsize=(8.0, 0.5*11.0),
                     gridspec_kw={'width_ratios':[1, 0.01, 0.4]})
        
        axes[0].set_title('$(T_\perp, A)$ Parameter Search')
        
        im = axes[0].pcolormesh(T_axis*1e-3, A_axis, max_growth.T, shading='auto', 
                              norm=colors.Normalize(), cmap='jet')
        fig.colorbar(im, cax=axes[1], extend='both').set_label(
                'Growth Rate at max Frequency')
        
        if target_freqs is not None:
            cs1 = axes[0].contour(T_axis*1e-3, A_axis, max_freq.T, target_freqs,
                                     colors='k', linewidths=0.75)
            axes[0].clabel(cs1, inline=True, manual=([(20., 1.), (40., 1.1), (60, 1.2)]))
        axes[0].set_xlim(T_axis[0]*1e-3, T_axis[-1]*1e-3)
        axes[0].set_ylim(A_axis[0], A_axis[-1])
        axes[0].set_xlabel('$T_\perp$ (keV)')
        axes[0].set_ylabel('A', rotation=0, labelpad=10)
        axes[2].set_visible(False)
        
        # Figure text for parameters
        Lx = 0.84; Ly = 0.90; dy = 0.03
        fig.text(Lx, Ly - 0*dy, '$B_0$ = %.1f nT'%B0)    
        fig.text(Lx, Ly - 1*dy, '$n_e$ = %.1f cc'%ne)    
        
        fig.text(Lx, Ly - 2*dy, '$H^+$  = %.1f %%' % (cH*100.))    
        fig.text(Lx, Ly - 3*dy, '$He^+$ = %.1f %%' % (cHe*100.))
        fig.text(Lx, Ly - 4*dy, '$O^+$  = %.1f %%' % (cO*100.))
        fig.text(Lx, Ly - 5*dy, '$n_h$  = %.1f %%' % (rc_frac*100.))
            
        fig.tight_layout(rect=[0.02, 0.02, 1.0, 1.0])
        fig.subplots_adjust(wspace=0.03, hspace=0.15)
        
        if save:
            save_path = plot_path + '//CRRES_TA_SEARCH_REFINED//'
            if not os.path.exists(save_path): os.makedirs(save_path)
            suff = f'B{B0*10:.0f}_n{ne:.0f}_He{int(cHe*1e3):03}O{int(cO*1e3):03}_rc{int(rc_frac*1e3):03}'
            save_name = save_path + f'//CRRES_TA_SEARCH_{suff}.png'
            print('Saving {}'.format(save_name))
            fig.savefig(save_name)
            plt.close('all')
        else:
            plt.show()
    except:
        print('Plot failed.')
        pass
    return


#%% -- MAIN --
if __name__ == '__main__':
    plot_path = 'D://Google Drive//Uni//PhD 2017//Josh PhD Share Folder//Thesis//Data_Plots//'
    ext_drive = 'E:'
    rbsp_path = '%s//DATA//RBSP//' % ext_drive
    
    # Sweeps for CRRES: 3 combinations
    # 17/7: 80nT/45cc, 70nT/20cc, 60nT/10cc 
    
# =============================================================================
#     for _mag, _dens in zip([80.0, 70.0, 60.0], [45., 20., 10.]):
#             for _nHe in [0.02, 0.05, 0.10]:
#                 for _nO in [0.02, 0.05, 0.10]:
#                     search_TA_space(_mag, _dens, _nHe, _nO, rc_frac=0.05, band=None,
#                                         T_min=0.0, T_max=100.0, nT=250,
#                                         A_min=0.0, A_max=2.00, nA=250,
#                                         f_max=0.7, target_freqs=[0.14, 0.16, 0.18],
#                                         save=True)
# =============================================================================
# =============================================================================
#     # 8/12: 165nT/100cc, 140nT/85cc, 120nT/50cc         
#     for _mag, _dens in zip([165., 140., 120.], [100., 85., 50.]):
#             for _nHe in [0.02, 0.05, 0.10]:
#                 for _nO in [0.02, 0.05, 0.10]:
#                     search_TA_space(_mag, _dens, _nHe, _nO, rc_frac=0.05, band=None,
#                                         T_min=0.0, T_max=100.0, nT=250,
#                                         A_min=0.0, A_max=2.00, nA=250,
#                                         f_max=0.7, target_freqs=[0.25, 0.30, 0.35],
#                                         save=True)
# =============================================================================
                    
    #validation_plots_fraser_1996()
    #validation_plots_omura2010()
    #validation_plots_wang_2016()
    #hybrid_test_plot()
    #power_spectrum_CGR_comparison()
    
    #plot_single_kCGR()
    #plot_parameter_sweep_kCGR()
    #plot_density_sweep_kCGR()
    #investigate_anisotropy()
    #plot_temperature_sweep_kCGR()
    #parameter_search_2D()
    
    #get_CRRES_kCGR_timeseries()
    
# =============================================================================
#     time_start  = np.datetime64('2013-07-25T21:25:00')
#     time_end    = np.datetime64('2013-07-25T21:47:00')
#     cutoff_filename = 'D://Google Drive//Uni//PhD 2017//Josh PhD Share Folder//Thesis//Data_Plots//20130725_RBSP-A//cutoffs_only.txt'
#     title_string= time_start.astype(object).strftime('%Y-%m-%d')
#     date_string = time_start.astype(object).strftime('%Y%m%d')
#     save_dir    = '{}//2D_LINEAR_THEORY//EVENT_{}//'.format(ext_drive, date_string)
#     
#     f_max = 1.2
#     glim  = 28
# =============================================================================
    
# For calculation
# =============================================================================
#     for _approx in ['cold', 'warm', 'hot']:
#         print(f'Doing {_approx} approximation...')
#         get_all_DRs_with_cutoffs(save_dir, time_start, time_end, 'a',  
#                         kmin=0.0, kmax=1.0, Nk=1000, knorm=True,
#                         _nsec=5, HM_filter_mhz=50, N_procs=7,
#                         he_frac=0.30, data_path=f'{ext_drive}//DATA//RBSP//',
#                         output=False, approx=_approx, load_only=False)
# =============================================================================

# =============================================================================
#     # For plotting
#     thesis_plot_single_2D_growth_timeseries(rbsp_path, time_start, time_end, 'a',
#                                  approx='warm', save=True, nsec=5, log=False)
# =============================================================================
    

    
# =============================================================================
#     for _mag in [159.0, 161.0, 163.0]:
#         for _dens in [50.0, 60.0, 70.0, 80.0, 90.0, 100., 110.]:
#             for _eV in [15e3, 20e3, 25e3, 30e3, 40e3, 60e3, 80e3, 100e3]:
#                 for _anis in [0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9, 1.0]:
#                     plot_CGR_heavyions_search(_mag, _dens, _eV, _anis, rc_frac=0.01, band=None,
#                                     he_min=0.0, he_max=50.0, nHe=500,
#                                     o_min=0.0, o_max=30.0, nO=500,
#                                     f_max=0.7, target_freqs=[0.20, 0.4, 0.6], save=True)
# =============================================================================
    
# =============================================================================
#     plot_CGR_heavyions_search(161.0, 70.0, 30e3, 0.5, rc_frac=0.01, band=None,
#                     he_min=0.0, he_max=50.0, nHe=500,
#                     o_min=0.0, o_max=30.0, nO=500,
#                     f_max=0.7, target_freqs=[0.20, 0.4, 0.6], save=True)
# =============================================================================
    #sys.exit()
    
    #### Read in command-line arguments, if present
    import argparse as ap
    parser = ap.ArgumentParser()
    parser.add_argument('-n', '--N_procs'   , default=1, type=int)
    args = vars(parser.parse_args())
    n_processes = args['N_procs']
    
    ### Do Sweep of growth rates to create save files ### 
    rbsp_path   = '%s//DATA//RBSP//' % ext_drive
    probe       = 'a'
    
    if False:
        time_start  = np.datetime64('2013-07-25T21:25:00')
        time_end    = np.datetime64('2013-07-25T21:47:00')
        plot_start  = np.datetime64('2013-07-25T21:00:00')
        plot_end    = np.datetime64('2013-07-25T22:00:00')
        f_max       = 1.2
        glim        = 28
    else:
        time_start  = np.datetime64('2015-01-16T04:25:00')
        time_end    = np.datetime64('2015-01-16T05:10:00')
        plot_start  = np.datetime64('2015-01-16T04:05:00')
        plot_end    = np.datetime64('2015-01-16T05:15:00')
        f_max       = 0.5
        glim        = 75
        
    #check_concentration_files(time_start)
    title_string= time_start.astype(object).strftime('%Y-%m-%d')
    date_string = time_start.astype(object).strftime('%Y%m%d')
    save_dir    = '{}//2D_LINEAR_THEORY//EVENT_{}//'.format(ext_drive, date_string)
    
    #thesis_plot_summaries(rbsp_path, plot_start, plot_end, probe)
    
    #plot_cold_ion_via_cutoff([0.05, 0.055, 0.06, 0.065, 0.07, 0.075, 0.08, 0.085, 0.09], title_suff=' :: 25/07/2013 Event')
    #plot_normalized_pc1(time_start, time_end, probe)
    
# =============================================================================
#     calculate_warm_sweep(rbsp_path, save_dir, plot_start, plot_end, probe,
#                           _nsec=5, N_procs=n_processes)
# =============================================================================

    thesis_plot_2D_growth_rates_with_time(rbsp_path, plot_start, plot_end, probe,
                             approx='warm', save=True, nsec=5, log=False,
                             _time_start=time_start, _time_end=time_end)
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
# =============================================================================
#     #%% -- TEST --
#     # Question: Do I zero the growth rates of species with no cold counterpart?
#     # Not quite terribly accurate for growth rates, but warm component alone probably
#     # doesn't have enough to really support it.
#     _cmp = [100., 0.0, 0.0]
#     times, B0, name, mass, charge, density, tper, ani, cold_dens = \
#         extract_species_arrays(plot_start, plot_end, probe, cmp=np.asarray(_cmp), 
#                                return_raw_ne=True, nsec=5, HM_filter_mhz=50.,
#                                rbsp_path=rbsp_path, HOPE_only=True)
#     
#     Species, PP = create_species_array(B0[0], name, mass, charge, density[:, 0], tper[:, 0], ani[:, 0],
#                                        remove_zero_density_species=False)
#     
#     this_k  = np.linspace(0.0, 2.0, 1000, endpoint=False)
#     this_k *= PP['pcyc_rad'] / PP['va']
#     
#     OUT_solns, CGR_solns, VEL_solns = get_dispersion_relation(Species, this_k, approx='warm',
#                             guesses=None, complex_out=True,
#                             print_filtered=True, return_vg=True, force_3soln=True)
# =============================================================================
    
    
# =============================================================================
#     if False:
#         rbsp_path   = '%s//DATA//RBSP//' % ext_drive
#         save_drive  = ext_drive
#         
#         time_start  = np.datetime64('2015-01-16T04:05:00')
#         time_end    = np.datetime64('2015-01-16T05:15:00')
#         probe       = 'a'
#         pad         = 0
#         fmax        = 0.5
#         
#         plot_start  = np.datetime64('2015-01-16T04:25:00')
#         plot_end    = np.datetime64('2015-01-16T05:10:00')
#         
#         # Test/Check plasma params from files
#         if False:
#             TIMES, MAG, NAME, MASS, CHARGE, DENS, TPER, ANI, COLD_DENS = \
#             extract_species_arrays(time_start, time_end, probe, cmp=np.asarray([70, 20, 10]), 
#                                    return_raw_ne=True, nsec=None, HM_filter_mhz=50)
#             sys.exit()
#         
#         date_string = time_start.astype(object).strftime('%Y%m%d')
#         save_string = time_start.astype(object).strftime('%Y%m%d_%H%M_') + time_end.astype(object).strftime('%H%M')
#         save_dir    = '{}//NEW_LT//EVENT_{}//CHEN_DR_CODE//'.format(save_drive, date_string)
#         
#     # =============================================================================
#     #     all_f, all_CGR_HOPE, all_stop_HOPE, all_CGR_SPICE, all_stop_SPICE,          \
#     #            times, B0, name, mass, charge, density, tper, anisotropy, cold_dens =\
#     #     get_all_CGRs_kozyra(time_start, time_end, probe, pad, cmp=np.array([70, 20, 10]), 
#     #                         fmax_pcyc=1.0, Nf=1000, nsec=None, HM_filter_mhz=50, instr='RBSPICE')
#     # =============================================================================
#         
#     
#         if False:
#             vlines = [  '2015-01-16T04:32:53.574540',
#                         '2015-01-16T04:35:36.689700',
#                         '2015-01-16T04:44:53.200200',
#                         '2015-01-16T04:47:48.309120',
#                         '2015-01-16T04:48:31.486660',
#                         '2015-01-16T04:49:17.062940',
#                         '2015-01-16T04:49:59.041120',
#                         '2015-01-16T04:51:11.003680',
#                         '2015-01-16T04:52:03.776220',
#                         '2015-01-16T04:53:38.526940',
#                         '2015-01-16T04:55:24.072040',
#                         '2015-01-16T04:36:27.063500',
#                         '2015-01-16T04:28:50.101200',
#                         '2015-01-16T04:29:30.879980',
#                         '2015-01-16T04:38:23.402980',
#                         '2015-01-16T04:39:11.378020',
#                         '2015-01-16T04:42:07.686300',
#                         '2015-01-16T04:56:19.461080',
#                         '2015-01-16T04:57:08.028780',
#                         '2015-01-16T04:59:38.756120',
#                         '2015-01-16T05:00:42.954780',
#                         '2015-01-16T05:02:08.925200',
#                         '2015-01-16T05:03:24.847120',
#                         '2015-01-16T05:03:58.342080',
#                         '2015-01-16T05:05:06.448500',
#                         '2015-01-16T05:07:41.083580',
#                         '2015-01-16T05:05:53.341440',
#                         '2015-01-16T04:40:01.703620',
#                         '2015-01-16T05:10:37.874220',
#                         '2015-01-16T05:34:50.015900',
#                         '2015-01-16T05:37:30.425800',
#                         '2015-01-16T05:40:04.314980',
#                         '2015-01-16T05:40:30.397900',
#                         '2015-01-16T05:44:50.574940',
#                         '2015-01-16T05:44:23.839960',
#                         '2015-01-16T04:58:06.574340',
#                         '2015-01-16T04:58:43.263980',
#                         '2015-01-16T04:54:14.206560',
#                         '2015-01-16T04:40:54.027960',
#                         '2015-01-16T04:33:50.763680']
#             
#             vspan = [
#                     ['2015-01-16T04:47:19.000000', '2015-01-16T04:48:53.000000'],
#                     ['2015-01-16T04:50:35.000000', '2015-01-16T04:52:47.000000'],
#                     ['2015-01-16T04:44:04.000000', '2015-01-16T04:45:39.000000'],
#                     ['2015-01-16T04:34:42.000000', '2015-01-16T04:36:00.000000'],
#                     ['2015-01-16T04:32:00.000000', '2015-01-16T04:33:00.000000'],
#                     
#                     ['2015-01-16T04:55:30.000000', '2015-01-16T04:57:37.000000'],
#                     
#                     ['2015-01-16T04:59:04.000000', '2015-01-16T05:01:22.000000'],
#                     ['2015-01-16T05:04:07.000000', '2015-01-16T05:06:04.000000'],
#                     ['2015-01-16T05:07:07.000000', '2015-01-16T05:08:43.000000'],
#                     ]
#             
#             
#             _K, _CPDR, _WPDR, _HPDR, _cCGR, _wCGR, _hCGR, _cVG, _wVG, _hVG,        \
#             TIMES, MAG, NAME, MASS, CHARGE, DENS, TPER, ANI, COLD_NE =             \
#             get_all_DRs_parallel(time_start, time_end, probe, [70, 20, 10], 
#                              kmin=0.0, kmax=1.0, Nk=5000, knorm=True,
#                              nsec=None, HM_filter_mhz=50, N_procs=6)
#             
#             #plot_all_DRs(_K, _CPDR, _WPDR, _HPDR, TIMES, MAG, NAME, MASS, CHARGE, DENS, TPER, ANI, COLD_NE,
#             #             suff='')
#             
#             #plot_all_CGRs(_K, _cCGR, _wCGR, _hCGR, TIMES, MAG, NAME, MASS, CHARGE, DENS, TPER, ANI, COLD_NE,
#             #              HM_filter_mhz=50, overwrite=False, save=True, figtext=True, suff='')
#             
#             plot_max_growth_rate_with_time(TIMES, _K, _CPDR, _WPDR, _HPDR,
#                                            save=True, norm_w=False, B0=None, plot_pc1=True,
#                                            ccomp=[70, 20, 10], suff='_withPc1', plot_pearls=False)
#             
#             #plot_max_CGR_with_time(TIMES, _K, _cCGR, _wCGR, _hCGR,  
#             #                        save=True, norm_w=False, B0=None, plot_pc1=True,
#             #                        ccomp=[70, 20, 10], suff='_withPc1', plot_pearls=True)
#     
# 
# =============================================================================
